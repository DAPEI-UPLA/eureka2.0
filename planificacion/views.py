"""Views de planificación. Acciones POST separadas en endpoints propios."""
import io
import json
import logging
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Count, Q
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_GET, require_POST, require_http_methods

from .forms import EstrategiaForm, IndicadorForm, ObjetivoForm, ProgramaForm
from .helpers import (
    DEFAULT_ANIOS,
    _safe_int,
    avance_global_ponderado,
    avance_indicador,
    color_avance,
    es_planificacion as _es_planificacion,
    get_user_role,
    grado_cumplimiento_plan,
    parse_anio_semestre,
    parse_decimal,
    planificacion_required,
)
from .models import (
    CierreAnual,
    Estrategia,
    Indicador,
    MetaIndicador,
    Objetivo,
    Programa,
    ProgramaIndicador,
    SeguimientoIndicador,
)

logger = logging.getLogger(__name__)


# =============================================================
# HOME
# =============================================================

@login_required
def planificacion_home(request):
    role = get_user_role(request)

    conteos = dict(
        Programa.objects.values_list("tipo")
        .annotate(total=Count("id"))
        .values_list("tipo", "total")
    )

    tipos = [
        {
            "codigo": Programa.TipoPrograma.PEI,
            "label": "PEI",
            "descripcion": "Plan Estratégico Institucional. Define el rumbo plurianual de la universidad.",
            "color": "#2563eb",
            "icono": "🧭",
            "total": conteos.get(Programa.TipoPrograma.PEI, 0),
        },
        {
            "codigo": Programa.TipoPrograma.PD_UNIDADES,
            "label": "PD Unidades Académicas",
            "descripcion": "Planes de Desarrollo de facultades, escuelas y carreras.",
            "color": "#7c3aed",
            "icono": "🏛️",
            "total": conteos.get(Programa.TipoPrograma.PD_UNIDADES, 0),
        },
        {
            "codigo": Programa.TipoPrograma.PD_FUNCIONES,
            "label": "PD Funciones Institucionales",
            "descripcion": "Planes de Desarrollo de funciones transversales.",
            "color": "#16a34a",
            "icono": "🧩",
            "total": conteos.get(Programa.TipoPrograma.PD_FUNCIONES, 0),
        },
    ]

    totales = {
        "objetivos": Objetivo.objects.count(),
        "indicadores": Indicador.objects.count(),
        "estrategias": Estrategia.objects.count(),
    }

    return render(request, "planificacion/home.html", {
        "tipos": tipos,
        "totales": totales,
        "es_planificacion": role["es_planificacion"],
    })


# =============================================================
# BÚSQUEDA GLOBAL
# =============================================================

@login_required
def busqueda_global(request):
    q = (request.GET.get("q") or "").strip()
    role = get_user_role(request)

    resultados = {"programas": [], "objetivos": [], "indicadores": [], "estrategias": []}
    if q:
        resultados["programas"] = list(
            Programa.objects.filter(nombre__icontains=q)[:10]
        )
        resultados["objetivos"] = list(
            Objetivo.objects.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))[:10]
        )
        resultados["indicadores"] = list(
            Indicador.objects.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))[:10]
        )
        resultados["estrategias"] = list(
            Estrategia.objects.filter(Q(nombre__icontains=q) | Q(descripcion__icontains=q))[:10]
        )

    total = sum(len(v) for v in resultados.values())
    return render(request, "planificacion/busqueda.html", {
        "q": q,
        "resultados": resultados,
        "total": total,
        "es_planificacion": role["es_planificacion"],
    })


# =============================================================
# INDICADORES
# =============================================================

@login_required
def lista_indicadores(request):
    role = get_user_role(request)

    indicadores = Indicador.objects.select_related("objetivo", "creado_por")
    form = IndicadorForm()

    if request.method == "POST":
        if not role["es_planificacion"]:
            raise PermissionDenied
        form = IndicadorForm(request.POST)
        if form.is_valid():
            indicador = form.save(commit=False)
            indicador.creado_por = request.user
            indicador.save()
            messages.success(request, "Indicador creado correctamente.")
            return redirect("planificacion:indicadores")

    return render(request, "planificacion/lista_indicadores.html", {
        "indicadores": indicadores,
        "form": form,
        "es_planificacion": role["es_planificacion"],
    })


@planificacion_required
@require_POST
def editar_indicador(request, pk):
    indicador = get_object_or_404(Indicador, pk=pk)
    form = IndicadorForm(request.POST, instance=indicador)
    if form.is_valid():
        form.save()
        messages.success(request, "Indicador actualizado.")
    else:
        logger.warning("Errores al editar indicador %s: %s", pk, form.errors)
        messages.error(request, "Errores al actualizar el indicador.")
    return redirect("planificacion:indicadores")


@planificacion_required
@require_POST
def eliminar_indicador(request, pk):
    indicador = get_object_or_404(Indicador, pk=pk)
    indicador.delete()
    messages.success(request, "Indicador eliminado.")
    return redirect("planificacion:indicadores")


# =============================================================
# OBJETIVOS
# =============================================================

@login_required
def lista_objetivos(request):
    role = get_user_role(request)
    objetivos = Objetivo.objects.prefetch_related("programas_asociados").select_related("creado_por")
    estrategicos = objetivos.filter(tipo=Objetivo.TipoObjetivo.ESTRATEGICO)
    especificos = objetivos.filter(tipo=Objetivo.TipoObjetivo.ESPECIFICO)

    form = ObjetivoForm()
    if request.method == "POST":
        if not role["es_planificacion"]:
            raise PermissionDenied
        form = ObjetivoForm(request.POST)
        if form.is_valid():
            objetivo = form.save(commit=False)
            objetivo.creado_por = request.user
            objetivo.save()
            messages.success(request, "Objetivo creado.")
            return redirect("planificacion:lista_objetivos")

    return render(request, "planificacion/lista_objetivos.html", {
        "estrategicos": estrategicos,
        "especificos": especificos,
        "form": form,
        "es_planificacion": role["es_planificacion"],
    })


@planificacion_required
@require_POST
def editar_objetivo(request, pk):
    objetivo = get_object_or_404(Objetivo, pk=pk)
    form = ObjetivoForm(request.POST, instance=objetivo)
    if form.is_valid():
        form.save()
        messages.success(request, "Objetivo actualizado.")
    else:
        logger.warning("Errores al editar objetivo %s: %s", pk, form.errors)
        messages.error(request, "Errores al actualizar el objetivo.")
    return redirect("planificacion:lista_objetivos")


@planificacion_required
@require_POST
def eliminar_objetivo(request, pk):
    objetivo = get_object_or_404(Objetivo, pk=pk)
    objetivo.eliminado = True
    objetivo.save(update_fields=["eliminado"])
    messages.success(request, "Objetivo eliminado.")
    return redirect("planificacion:lista_objetivos")


# =============================================================
# ESTRATEGIAS
# =============================================================

def _aplicar_indicadores_post(request, estrategia):
    """Si el POST trae el flag manage_indicadores=1, sincroniza la M2M
    desde la lista 'indicadores'. Si no viene el flag, deja la relación intacta
    (el modal de edición rápida no envía ese flag)."""
    if request.POST.get("manage_indicadores") != "1":
        return
    raw_ids = request.POST.getlist("indicadores")
    ids = []
    for x in raw_ids:
        x = (x or "").strip()
        if not x:
            continue
        try:
            ids.append(int(x))
        except ValueError:
            continue
    estrategia.indicadores.set(ids)


@login_required
def lista_estrategias(request):
    role = get_user_role(request)

    estrategias = (
        Estrategia.objects
        .prefetch_related("indicadores")
        .select_related("creado_por")
    )
    form = EstrategiaForm()

    if request.method == "POST":
        if not role["es_planificacion"]:
            raise PermissionDenied
        form = EstrategiaForm(request.POST)
        if form.is_valid():
            estrategia = form.save(commit=False)
            estrategia.creado_por = request.user
            estrategia.save()
            _aplicar_indicadores_post(request, estrategia)
            messages.success(request, "Estrategia creada.")
            return redirect("planificacion:estrategias")

    return render(request, "planificacion/lista_estrategias.html", {
        "estrategias": estrategias,
        "form": form,
        "es_planificacion": role["es_planificacion"],
        "indicadores_disponibles": Indicador.objects.order_by("nombre"),
    })


@planificacion_required
def editar_estrategia(request, pk):
    estrategia = get_object_or_404(Estrategia, pk=pk)
    if request.method == "POST":
        form = EstrategiaForm(request.POST, instance=estrategia)
        if form.is_valid():
            form.save()
            _aplicar_indicadores_post(request, estrategia)
            messages.success(request, "Estrategia actualizada.")
            return redirect("planificacion:estrategias")
    else:
        form = EstrategiaForm(instance=estrategia)
    return render(request, "planificacion/editar_estrategia.html", {
        "form": form,
        "estrategia": estrategia,
        "indicadores_disponibles": Indicador.objects.order_by("nombre"),
        "indicadores_actuales": list(
            estrategia.indicadores.values_list("id", flat=True)
        ),
    })


@planificacion_required
@require_POST
def eliminar_estrategia(request, pk):
    estrategia = get_object_or_404(Estrategia, pk=pk)
    estrategia.delete()
    messages.success(request, "Estrategia eliminada.")
    return redirect("planificacion:estrategias")


# =============================================================
# PROGRAMAS - LISTAS
# =============================================================

@login_required
def lista_programas(request):
    role = get_user_role(request)

    programas = (
        Programa.objects.prefetch_related("objetivos")
        .annotate(
            total_objetivos_db=Count("objetivos", distinct=True),
            total_indicadores=Count("programas_indicadores", distinct=True),
        )
    )

    form = ProgramaForm()
    if request.method == "POST":
        if not role["es_planificacion"]:
            raise PermissionDenied
        form = ProgramaForm(request.POST)
        if form.is_valid():
            programa = form.save(commit=False)
            programa.creado_por = request.user
            programa.save()
            form.save_m2m()
            messages.success(request, "Plan creado.")
            return redirect("planificacion:lista_programas")

    return render(request, "planificacion/lista_programas.html", {
        "programas": programas,
        "form": form,
        "tipo": None,
        "tipo_label": "Todos los planes",
        "es_planificacion": role["es_planificacion"],
    })


@login_required
def lista_programas_tipo(request, tipo):
    role = get_user_role(request)

    valid_types = {choice[0] for choice in Programa.TipoPrograma.choices}
    if tipo not in valid_types:
        programas = Programa.objects.none()
    else:
        programas = (
            Programa.objects.filter(tipo=tipo)
            .prefetch_related("objetivos")
            .annotate(
                total_objetivos_db=Count("objetivos", distinct=True),
                total_indicadores=Count("programas_indicadores", distinct=True),
            )
        )

    form = ProgramaForm()
    if request.method == "POST" and role["es_planificacion"] and tipo in valid_types:
        form = ProgramaForm(request.POST)
        if form.is_valid():
            programa = form.save(commit=False)
            programa.tipo = tipo
            programa.creado_por = request.user
            programa.save()
            messages.success(request, "Plan creado.")
            return redirect("planificacion:lista_programas_tipo", tipo=tipo)

    tipo_label = dict(Programa.TipoPrograma.choices).get(tipo, "Planes")
    return render(request, "planificacion/lista_programas.html", {
        "programas": programas,
        "form": form,
        "tipo": tipo,
        "tipo_label": tipo_label,
        "es_planificacion": role["es_planificacion"],
    })


@planificacion_required
@require_POST
def crear_programa(request):
    form = ProgramaForm(request.POST)
    if form.is_valid():
        p = form.save(commit=False)
        p.creado_por = request.user
        p.save()
    return redirect("planificacion:home")


@planificacion_required
@require_POST
def editar_programa(request, id):
    programa = get_object_or_404(Programa, id=id)
    nombre = (request.POST.get("nombre") or "").strip()
    if not nombre:
        messages.error(request, "El nombre no puede estar vacío.")
    elif (
        Programa.objects.filter(nombre=nombre, tipo=programa.tipo)
        .exclude(pk=programa.pk)
        .exists()
    ):
        messages.error(request, "Ya existe un plan con ese nombre en esta sección.")
    elif nombre != programa.nombre:
        programa.nombre = nombre
        programa.save(update_fields=["nombre"])
        messages.success(request, "Nombre del plan actualizado.")
    return redirect("planificacion:lista_programas_tipo", tipo=programa.tipo)


@planificacion_required
@require_POST
def eliminar_programa(request, pk):
    programa = get_object_or_404(Programa, pk=pk)
    # soft delete
    programa.eliminado = True
    programa.save()
    messages.success(request, "Plan eliminado.")
    return redirect("planificacion:lista_programas_tipo", tipo=programa.tipo)


@planificacion_required
@require_POST
def cambiar_estado_programa(request, pk):
    programa = get_object_or_404(Programa, pk=pk)
    nuevo = request.POST.get("estado")
    valid = {choice[0] for choice in Programa.Estado.choices}
    if nuevo not in valid:
        messages.error(request, "Estado inválido.")
    else:
        programa.estado = nuevo
        programa.save()
        messages.success(request, f"Plan marcado como {programa.get_estado_display()}.")
    return redirect("planificacion:seguimiento_programa", pk=pk)


# =============================================================
# SEGUIMIENTO PROGRAMA — LECTURA
# =============================================================

def _construir_estructura(programa, anio, semestre):
    estructura = []
    items_para_global = []

    for objetivo in programa.objetivos.order_by("fecha_creacion"):
        indicadores = []
        items_objetivo = []
        n_con_dato = 0

        # Orden por id ASC del ProgramaIndicador = orden en que se agregaron al plan.
        pis_objetivo = (
            programa.programas_indicadores
            .filter(indicador__objetivo=objetivo)
            .select_related("indicador")
            .prefetch_related("seguimientos", "metas_anuales", "indicador__estrategias")
            .order_by("id")
        )

        for pi in pis_objetivo:
            indicador = pi.indicador
            data = avance_indicador(pi, anio, semestre)
            seguimientos = list(pi.seguimientos.all())

            actual_seg = data.get("seguimiento")
            indicadores.append({
                "pi": pi,
                "indicador": indicador,
                "meta": data["meta"],
                "meta_display": data["meta_display"],
                "estado_meta": data["estado_meta"],
                "valor": data["valor"],
                "valor_display": data["valor_display"],
                "estado_medicion": data["estado_medicion"],
                "avance": data["avance"],
                "color": data["color"],
                "color_hex": data["color_hex"],
                "seguimientos": seguimientos,
                "estrategias": indicador.estrategias.all(),
                "comentario_actual": actual_seg.comentario if actual_seg else "",
                "evidencia_actual": actual_seg.evidencia if actual_seg and actual_seg.evidencia else None,
            })

            item_score = {"color": data["color"], "peso": pi.peso}
            items_objetivo.append(item_score)

            if data["avance"] is not None:
                n_con_dato += 1
                items_para_global.append(item_score)

        # Cumplimiento del objetivo con la misma fórmula que el global,
        # restringido a los indicadores del objetivo.
        cumplimiento_obj = grado_cumplimiento_plan(items_objetivo)

        estructura.append({
            "objetivo": objetivo,
            "indicadores": indicadores,
            "avance": cumplimiento_obj,
            "color_obj": color_avance(cumplimiento_obj),
            "indicadores_total": len(indicadores),
            "indicadores_con_dato": n_con_dato,
        })

    avance_global = grado_cumplimiento_plan(items_para_global)
    return estructura, avance_global


@login_required
def seguimiento_programa(request, pk):
    role = get_user_role(request)

    programa = get_object_or_404(
        Programa.objects.prefetch_related(
            "objetivos",
            "programas_indicadores__indicador__estrategias",
            "programas_indicadores__seguimientos",
            "programas_indicadores__metas_anuales",
        ),
        pk=pk,
    )

    anio, semestre = parse_anio_semestre(request, default_sem=2)

    objetivos_disponibles = Objetivo.objects.exclude(programas_asociados=programa)
    indicadores_disponibles = Indicador.objects.exclude(
        id__in=programa.programas_indicadores.values_list("indicador_id", flat=True)
    )

    estructura, avance_global = _construir_estructura(programa, anio, semestre)

    cierre_existente = programa.cierres.filter(anio=anio).exists()

    return render(request, "planificacion/seguimiento_programa.html", {
        "programa": programa,
        "estructura": estructura,
        "avance_global": avance_global,
        "anio": anio,
        "semestre": semestre,
        "anios": DEFAULT_ANIOS,
        "objetivos_disponibles": objetivos_disponibles,
        "indicadores_disponibles": indicadores_disponibles,
        "es_planificacion": role["es_planificacion"],
        "es_publico": False,
        "cierre_existente": cierre_existente,
        "total_objetivos": len(estructura),
        "total_indicadores": sum(b["indicadores_total"] for b in estructura),
    })


@login_required
def seguimiento_programa_publico(request, pk):
    """Vista de solo lectura para autoridades."""
    programa = get_object_or_404(Programa, pk=pk)
    anio, semestre = parse_anio_semestre(request, default_sem=2)
    estructura, avance_global = _construir_estructura(programa, anio, semestre)
    return render(request, "planificacion/seguimiento_programa.html", {
        "programa": programa,
        "estructura": estructura,
        "avance_global": avance_global,
        "anio": anio,
        "semestre": semestre,
        "anios": DEFAULT_ANIOS,
        "objetivos_disponibles": [],
        "indicadores_disponibles": [],
        "es_planificacion": False,
        "es_publico": True,
        "cierre_existente": True,
        "total_objetivos": len(estructura),
        "total_indicadores": sum(b["indicadores_total"] for b in estructura),
    })


# =============================================================
# HEATMAP MULTI-AÑO
# =============================================================

@login_required
def heatmap_programa(request, pk):
    programa = get_object_or_404(
        Programa.objects.prefetch_related(
            "programas_indicadores__seguimientos",
            "programas_indicadores__metas_anuales",
        ),
        pk=pk,
    )

    semestre = _safe_int(request.GET.get("semestre"), 2)
    if semestre not in (1, 2):
        semestre = 2

    # Cargamos los PIs una sola vez, ordenados por id (orden de inserción al plan)
    # y los agrupamos por objetivo para iterar sin disparar queries extra.
    todos_pis = list(
        programa.programas_indicadores
        .select_related("indicador")
        .order_by("id")
    )
    pis_por_objetivo_id = {}
    for pi in todos_pis:
        pis_por_objetivo_id.setdefault(pi.indicador.objetivo_id, []).append(pi)

    grupos = []
    for objetivo in programa.objetivos.order_by("fecha_creacion"):
        indicadores_filas = []
        promedio_celdas = []
        pis_objetivo = pis_por_objetivo_id.get(objetivo.id, [])

        for pi in pis_objetivo:
            ind = pi.indicador
            celdas = []
            for anio in DEFAULT_ANIOS:
                data = avance_indicador(pi, anio, semestre=semestre)
                celdas.append({
                    "anio": anio,
                    "avance": data["avance"],
                    "color": data["color"],
                    "valor_display": data["valor_display"],
                    "meta_display": data["meta_display"],
                })
            indicadores_filas.append({"indicador": ind, "celdas": celdas})

        for idx, anio in enumerate(DEFAULT_ANIOS):
            valores = [
                fila["celdas"][idx]["avance"]
                for fila in indicadores_filas
                if fila["celdas"][idx]["avance"] is not None
            ]
            avg = (sum(valores) / len(valores)) if valores else None
            promedio_celdas.append({
                "anio": anio,
                "avance": avg,
                "color": color_avance(avg),
            })

        grupos.append({
            "objetivo": objetivo,
            "indicadores": indicadores_filas,
            "promedio": promedio_celdas,
            "total_indicadores": len(indicadores_filas),
        })

    return render(request, "planificacion/heatmap.html", {
        "programa": programa,
        "grupos": grupos,
        "anios": DEFAULT_ANIOS,
        "semestre": semestre,
        "semestre_label": "1° semestre" if semestre == 1 else "Anual (cierre 2° sem.)",
    })


# =============================================================
# COMPARAR PROGRAMAS
# =============================================================

@login_required
def comparar_programas(request):
    raw_ids = request.GET.getlist("p")
    ids = []
    for x in raw_ids:
        v = _safe_int(x, None)
        if v is not None:
            ids.append(v)
    ids = ids[:3]

    seleccionados = list(Programa.objects.filter(id__in=ids))
    anio = _safe_int(request.GET.get("anio"), DEFAULT_ANIOS[0])
    semestre = _safe_int(request.GET.get("semestre"), 2)
    if semestre not in (1, 2):
        semestre = 2

    columnas = []
    for p in seleccionados:
        estructura, avance = _construir_estructura(p, anio, semestre)
        columnas.append({"programa": p, "estructura": estructura, "avance": avance})

    todos = Programa.objects.all().order_by("tipo", "nombre")
    return render(request, "planificacion/comparar.html", {
        "todos": todos,
        "seleccionados": seleccionados,
        "columnas": columnas,
        "anio": anio,
        "semestre": semestre,
        "anios": DEFAULT_ANIOS,
    })


# =============================================================
# ENDPOINTS DEL PROGRAMA — POST específicos
# =============================================================

def _get_programa_editable(request, pk):
    programa = get_object_or_404(Programa, pk=pk)
    if not _es_planificacion(request.user):
        raise PermissionDenied
    if not programa.es_editable:
        raise PermissionDenied
    return programa


@planificacion_required
@require_POST
def agregar_objetivo(request, pk):
    programa = _get_programa_editable(request, pk)
    obj_id = request.POST.get("objetivo_id")
    if obj_id:
        programa.objetivos.add(obj_id)
        messages.success(request, "Objetivo agregado al plan.")
    return _redirect_seguimiento(request, programa.pk)


@planificacion_required
@require_POST
def quitar_objetivo(request, pk):
    programa = _get_programa_editable(request, pk)
    obj_id = request.POST.get("objetivo_id")
    if obj_id:
        programa.objetivos.remove(obj_id)
        messages.success(request, "Objetivo retirado del plan.")
    return _redirect_seguimiento(request, programa.pk)


@planificacion_required
@require_POST
def agregar_indicador(request, pk):
    programa = _get_programa_editable(request, pk)
    ind_id = request.POST.get("indicador_id")
    if ind_id:
        ind = get_object_or_404(Indicador, pk=ind_id)
        # Validación: el objetivo del indicador debe estar en el programa
        if not programa.objetivos.filter(id=ind.objetivo_id).exists():
            messages.error(request, "El objetivo del indicador no está en el plan.")
            return _redirect_seguimiento(request, programa.pk)
        ProgramaIndicador.objects.get_or_create(programa=programa, indicador=ind)
        messages.success(request, "Indicador agregado al plan.")
    return _redirect_seguimiento(request, programa.pk)


@planificacion_required
@require_POST
def quitar_indicador(request, pk):
    programa = _get_programa_editable(request, pk)
    ind_id = request.POST.get("indicador_id")
    if ind_id:
        ProgramaIndicador.objects.filter(programa=programa, indicador_id=ind_id).delete()
        messages.success(request, "Indicador retirado del plan.")
    return _redirect_seguimiento(request, programa.pk)


@planificacion_required
@require_http_methods(["POST"])
def registrar_valor(request, pi_id):
    pi = get_object_or_404(
        ProgramaIndicador.objects.select_related("programa", "indicador"), pk=pi_id
    )
    if not pi.programa.es_editable:
        messages.error(request, "Plan cerrado")
        return _redirect_seguimiento(request, pi.programa_id)

    anio = _safe_int(request.POST.get("anio"), DEFAULT_ANIOS[0])
    semestre = _safe_int(request.POST.get("semestre"), 2)
    if semestre not in (1, 2):
        semestre = 2

    estado = (request.POST.get("estado_medicion") or "VALOR").upper()
    if estado not in {"VALOR", "SM", "NA"}:
        estado = "VALOR"

    valor = parse_decimal(request.POST.get("valor"))
    if estado == "VALOR" and valor is None:
        messages.error(request, "Valor inválido. Si no aplica, usa S/M o N/A.")
        return _redirect_seguimiento(request, pi.programa_id, anio, semestre)

    # Si es S/M o N/A, forzamos valor numérico a None
    if estado != "VALOR":
        valor = None

    # Bloquear si el año está cerrado
    if pi.programa.cierres.filter(anio=anio).exists():
        messages.error(request, f"El año {anio} está cerrado para este plan.")
        return _redirect_seguimiento(request, pi.programa_id, anio, semestre)

    obj, created = SeguimientoIndicador.objects.update_or_create(
        programa_indicador=pi,
        anio=anio,
        semestre=semestre,
        defaults={
            "valor": valor,
            "estado": estado,
            "comentario": request.POST.get("comentario", ""),
            "creado_por": request.user,
        },
    )

    archivo = request.FILES.get("evidencia")
    if archivo:
        obj.evidencia = archivo
        obj.save(update_fields=["evidencia"])

    messages.success(request, "Medición registrada.")
    return _redirect_seguimiento(request, pi.programa_id, anio, semestre)


@planificacion_required
@require_POST
def guardar_meta(request, pi_id):
    pi = get_object_or_404(ProgramaIndicador.objects.select_related("programa"), pk=pi_id)
    if not pi.programa.es_editable:
        messages.error(request, "Plan cerrado")
        return _redirect_seguimiento(request, pi.programa_id)

    anio = _safe_int(request.POST.get("anio"), DEFAULT_ANIOS[0])
    estado = (request.POST.get("estado_meta") or "VALOR").upper()
    if estado not in {"VALOR", "SM", "NA"}:
        estado = "VALOR"

    meta_val = parse_decimal(request.POST.get("meta"))
    peso_int = _safe_int(request.POST.get("peso"), None)

    if estado == "VALOR" and meta_val is None:
        messages.error(request, "Meta inválida. Si no aplica, usa S/M o N/A.")
        return _redirect_seguimiento(request, pi.programa_id, anio)

    if estado != "VALOR":
        meta_val = None

    MetaIndicador.objects.update_or_create(
        programa_indicador=pi,
        anio=anio,
        defaults={"valor": meta_val, "estado": estado},
    )
    if estado == "VALOR" and meta_val is not None:
        pi.meta = meta_val

    if peso_int in (1, 2):
        pi.peso = Decimal(peso_int)
    pi.save()

    messages.success(request, "Meta guardada.")
    return _redirect_seguimiento(
        request, pi.programa_id, anio, _safe_int(request.POST.get("semestre"), 2)
    )


@planificacion_required
@require_POST
def guardar_meta_y_valor(request, pi_id):
    """Guarda meta + peso del año + medición del período en una sola pasada."""
    pi = get_object_or_404(
        ProgramaIndicador.objects.select_related("programa", "indicador"), pk=pi_id
    )
    if not pi.programa.es_editable:
        messages.error(request, "Plan cerrado")
        return _redirect_seguimiento(request, pi.programa_id)

    anio_meta = _safe_int(request.POST.get("anio_meta"), DEFAULT_ANIOS[0])
    anio_med = _safe_int(request.POST.get("anio"), anio_meta)
    semestre = _safe_int(request.POST.get("semestre"), 2)
    if semestre not in (1, 2):
        semestre = 2

    # ---- META ----
    estado_meta = (request.POST.get("estado_meta") or "VALOR").upper()
    if estado_meta not in {"VALOR", "SM", "NA"}:
        estado_meta = "VALOR"
    meta_val = parse_decimal(request.POST.get("meta"))
    peso_int = _safe_int(request.POST.get("peso"), None)

    if estado_meta == "VALOR" and meta_val is None:
        messages.error(request, "Meta inválida. Si no aplica, usa S/M o N/A.")
        return _redirect_seguimiento(request, pi.programa_id, anio_meta, semestre)
    if estado_meta != "VALOR":
        meta_val = None

    # ---- MEDICIÓN ----
    estado_med = (request.POST.get("estado_medicion") or "VALOR").upper()
    if estado_med not in {"VALOR", "SM", "NA"}:
        estado_med = "VALOR"
    valor = parse_decimal(request.POST.get("valor"))
    if estado_med == "VALOR" and valor is None:
        messages.error(request, "Medición inválida. Si no aplica, usa S/M o N/A.")
        return _redirect_seguimiento(request, pi.programa_id, anio_meta, semestre)
    if estado_med != "VALOR":
        valor = None

    # Bloquear si el año de la medición está cerrado
    if pi.programa.cierres.filter(anio=anio_med).exists():
        messages.error(request, f"El año {anio_med} está cerrado para este plan.")
        return _redirect_seguimiento(request, pi.programa_id, anio_meta, semestre)

    # ---- GUARDAR (transaccional) ----
    from django.db import transaction
    with transaction.atomic():
        MetaIndicador.objects.update_or_create(
            programa_indicador=pi,
            anio=anio_meta,
            defaults={"valor": meta_val, "estado": estado_meta},
        )
        if estado_meta == "VALOR" and meta_val is not None:
            pi.meta = meta_val
        if peso_int in (1, 2):
            pi.peso = Decimal(peso_int)
        pi.save()

        seg, _ = SeguimientoIndicador.objects.update_or_create(
            programa_indicador=pi,
            anio=anio_med,
            semestre=semestre,
            defaults={
                "valor": valor,
                "estado": estado_med,
                "comentario": request.POST.get("comentario", ""),
                "creado_por": request.user,
            },
        )
        archivo = request.FILES.get("evidencia")
        if archivo:
            seg.evidencia = archivo
            seg.save(update_fields=["evidencia"])

    messages.success(request, "Meta y medición guardadas.")
    return _redirect_seguimiento(request, pi.programa_id, anio_meta, semestre)


@planificacion_required
@require_POST
def cerrar_anio(request, pk):
    programa = get_object_or_404(Programa, pk=pk)
    anio = _safe_int(request.POST.get("anio"), DEFAULT_ANIOS[0])
    obj, created = CierreAnual.objects.get_or_create(
        programa=programa,
        anio=anio,
        defaults={
            "cerrado_por": request.user,
            "comentario": request.POST.get("comentario", ""),
        },
    )
    if created:
        messages.success(request, f"Año {anio} cerrado para {programa.nombre}.")
    else:
        messages.info(request, f"Año {anio} ya estaba cerrado.")
    return _redirect_seguimiento(request, programa.pk, anio)


@planificacion_required
@require_POST
def reabrir_anio(request, pk):
    programa = get_object_or_404(Programa, pk=pk)
    anio = _safe_int(request.POST.get("anio"), DEFAULT_ANIOS[0])
    CierreAnual.objects.filter(programa=programa, anio=anio).delete()
    messages.success(request, f"Año {anio} reabierto.")
    return _redirect_seguimiento(request, programa.pk, anio)


def _redirect_seguimiento(request, programa_id, anio=None, semestre=None):
    url = reverse("planificacion:seguimiento_programa", args=[programa_id])
    params = []
    if anio:
        params.append(f"anio={anio}")
    if semestre:
        params.append(f"semestre={semestre}")
    if params:
        url = f"{url}?{'&'.join(params)}"
    return redirect(url)


# =============================================================
# JSON / SERIES
# =============================================================

@login_required
@require_GET
def api_series_programa(request, pk):
    programa = get_object_or_404(Programa, pk=pk)
    out = []
    for pi in programa.programas_indicadores.select_related("indicador").prefetch_related("seguimientos", "metas_anuales"):
        serie = []
        for s in pi.seguimientos.all():
            serie.append({
                "anio": s.anio,
                "semestre": s.semestre,
                "valor": float(s.valor) if s.valor is not None else None,
            })
        metas = [{"anio": m.anio, "valor": float(m.valor)} for m in pi.metas_anuales.all()]
        out.append({
            "pi_id": pi.id,
            "indicador": pi.indicador.nombre,
            "calculo_invertido": pi.indicador.calculo_invertido,
            "acumulativo": pi.indicador.acumulativo,
            "meta_legacy": float(pi.meta) if pi.meta is not None else None,
            "metas_anuales": metas,
            "serie": serie,
        })
    return JsonResponse({"programa": programa.nombre, "indicadores": out})


# =============================================================
# EXCEL: PLANTILLA + IMPORT + EXPORT
# =============================================================

EXCEL_HEADERS = [
    "programa_id",
    "programa_nombre",
    "indicador_id",
    "indicador_nombre",
    "anio",
    "semestre",
    "valor",
    "comentario",
]


@planificacion_required
@require_GET
def descargar_plantilla_excel(request, pk):
    """Plantilla pre-rellenada con los indicadores del programa."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return HttpResponse(
            "openpyxl no está instalado. Ejecuta: pip install openpyxl",
            status=500,
        )

    programa = get_object_or_404(Programa, pk=pk)

    wb = Workbook()
    ws = wb.active
    ws.title = "Seguimiento"

    # Encabezados
    for col, h in enumerate(EXCEL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2563EB")

    # Filas pre-rellenadas
    row = 2
    for pi in programa.programas_indicadores.select_related("indicador"):
        for sem in (1, 2):
            ws.cell(row=row, column=1, value=programa.id)
            ws.cell(row=row, column=2, value=programa.nombre)
            ws.cell(row=row, column=3, value=pi.indicador_id)
            ws.cell(row=row, column=4, value=pi.indicador.nombre)
            ws.cell(row=row, column=5, value=DEFAULT_ANIOS[0])
            ws.cell(row=row, column=6, value=sem)
            ws.cell(row=row, column=7, value=None)
            ws.cell(row=row, column=8, value="")
            row += 1

    # Hoja de instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["Plantilla de carga — Planificación"],
        [""],
        ["Cómo usar esta plantilla:"],
        ["1. No cambies la fila 1 (encabezados)."],
        ["2. Una fila por (indicador, año, semestre). Llena 'valor' con un número."],
        ["3. 'semestre' debe ser 1 o 2 (1=primer semestre, 2=segundo)."],
        ["4. 'comentario' es opcional."],
        ["5. Las filas con valor vacío se ignoran al subir."],
        ["6. Si el año está cerrado, esa fila será rechazada."],
        ["7. Sube el archivo desde el botón 'Importar Excel' en el plan."],
    ]
    for r, fila in enumerate(instrucciones, 1):
        ws2.cell(row=r, column=1, value=fila[0])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = (
        f'attachment; filename="plantilla_seguimiento_{programa.id}.xlsx"'
    )
    return resp


@planificacion_required
@require_POST
def importar_excel(request, pk):
    programa = get_object_or_404(Programa, pk=pk)

    if not programa.es_editable:
        messages.error(request, "El plan está cerrado.")
        return redirect("planificacion:seguimiento_programa", pk=pk)

    archivo = request.FILES.get("archivo")
    if not archivo:
        messages.error(request, "Adjunta un archivo .xlsx")
        return redirect("planificacion:seguimiento_programa", pk=pk)

    try:
        from openpyxl import load_workbook
    except ImportError:
        messages.error(request, "openpyxl no está instalado.")
        return redirect("planificacion:seguimiento_programa", pk=pk)

    try:
        wb = load_workbook(archivo, data_only=True)
        ws = wb["Seguimiento"] if "Seguimiento" in wb.sheetnames else wb.active
    except Exception as e:
        messages.error(request, f"Error al abrir Excel: {e}")
        return redirect("planificacion:seguimiento_programa", pk=pk)

    headers = [str(c.value or "").strip() for c in ws[1]]
    if headers[:6] != EXCEL_HEADERS[:6]:
        messages.error(
            request,
            "Encabezados inválidos. Descarga la plantilla y úsala sin modificar la fila 1.",
        )
        return redirect("planificacion:seguimiento_programa", pk=pk)

    aniades_cerrados = set(programa.cierres.values_list("anio", flat=True))

    ok = 0
    errores = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None for c in row):
            continue
        try:
            programa_id, _pn, indicador_id, _in, anio, semestre, valor, comentario = (
                list(row) + [None] * 8
            )[:8]

            if int(programa_id) != programa.id:
                errores.append(f"Fila {i}: programa_id no coincide.")
                continue

            if valor in (None, ""):
                continue

            anio = int(anio)
            semestre = int(semestre)
            if semestre not in (1, 2):
                errores.append(f"Fila {i}: semestre debe ser 1 o 2.")
                continue
            if anio in aniades_cerrados:
                errores.append(f"Fila {i}: año {anio} cerrado.")
                continue

            valor_dec = parse_decimal(valor)
            if valor_dec is None:
                errores.append(f"Fila {i}: valor inválido.")
                continue

            pi = programa.programas_indicadores.filter(indicador_id=indicador_id).first()
            if not pi:
                errores.append(f"Fila {i}: indicador {indicador_id} no asociado.")
                continue

            SeguimientoIndicador.objects.update_or_create(
                programa_indicador=pi,
                anio=anio,
                semestre=semestre,
                defaults={
                    "valor": valor_dec,
                    "comentario": comentario or "",
                    "creado_por": request.user,
                },
            )
            ok += 1
        except (ValueError, TypeError) as e:
            errores.append(f"Fila {i}: {e}")

    if ok:
        messages.success(request, f"Se importaron {ok} valores.")
    if errores:
        messages.warning(request, "Errores: " + " | ".join(errores[:8]))
    return redirect("planificacion:seguimiento_programa", pk=pk)


@planificacion_required
@require_GET
def exportar_excel(request, pk):
    """Exporta el seguimiento actual a Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return HttpResponse("openpyxl no instalado", status=500)

    programa = get_object_or_404(Programa, pk=pk)
    wb = Workbook()
    ws = wb.active
    ws.title = "Seguimiento"

    for col, h in enumerate(EXCEL_HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2563EB")

    row = 2
    for pi in programa.programas_indicadores.select_related("indicador").prefetch_related("seguimientos"):
        for s in pi.seguimientos.all():
            ws.cell(row=row, column=1, value=programa.id)
            ws.cell(row=row, column=2, value=programa.nombre)
            ws.cell(row=row, column=3, value=pi.indicador_id)
            ws.cell(row=row, column=4, value=pi.indicador.nombre)
            ws.cell(row=row, column=5, value=s.anio)
            ws.cell(row=row, column=6, value=s.semestre)
            ws.cell(row=row, column=7, value=float(s.valor) if s.valor is not None else None)
            ws.cell(row=row, column=8, value=s.comentario or "")
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="seguimiento_{programa.id}.xlsx"'
    return resp


# =============================================================
# LEGACY (mantener para no romper imports)
# =============================================================

@login_required
def index(request):
    return render(request, "planificacion/index.html")


@login_required
def desafios(request):
    return render(request, "planificacion/desafios.html")


@login_required
def home_planificacion(request):
    return planificacion_home(request)
