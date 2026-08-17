"""Pruebas del tablero maestro de resultados OCT.

Dos cosas que hay que poder afirmar sin dudar:

1. que los números calculados **calzan con los del Excel** (por eso las cifras
   de referencia salen del archivo real, no de valores inventados), y
2. que subir la planilla otra vez **no duplica ni pisa** lo que se corrigió a
   mano en el sistema.
"""

import re
import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path

from django.contrib.auth.models import User
from django.test import TestCase, override_settings
from django.urls import reverse
import openpyxl
from openpyxl import Workbook

from . import graficos, tablero
from .models import (
    Ambito,
    EstadoGestion,
    Gestion,
    MetaAmbito,
    Origen,
    ProyeccionMensual,
)
from .planilla import ErrorImportacion, ImportadorPlanilla

RUTA_PLANILLA = Path(__file__).resolve().parent / "Planilla_Resultados_OCT_2026.xlsx"

ENCABEZADOS = [
    "COD", "Ámbito", "Tipo de iniciativa", "Nombre de la iniciativa",
    "Institución", "Fecha de ingreso", "Mes de ingreso", "Monto postulado",
    "Estado", "Fecha de resultado", "Mes de resultado", "Monto adjudicado",
    "Responsable", "Observaciones",
]


def planilla(filas, proyeccion=None, metas=None, titulo="REGISTRO DE INICIATIVAS 2026"):
    """Arma un .xlsx con la forma de la planilla real y devuelve su ruta.

    ``filas`` son diccionarios con las claves de ``ENCABEZADOS``; lo que no
    esté queda vacío, igual que en el archivo de verdad.
    """
    libro = Workbook()

    hoja = libro.active
    hoja.title = "Registro iniciativas"
    hoja.append([titulo])
    hoja.append([])
    hoja.append(ENCABEZADOS)
    for fila in filas:
        hoja.append([fila.get(c) for c in ENCABEZADOS])

    if proyeccion is not None:
        hp = libro.create_sheet("Proyección financiera")
        hp.append(["PROYECCIÓN FINANCIERA 2026"])
        hp.append([])
        hp.append(["Ámbito"] + [f"{n}-26" for n in
                                ["ene", "feb", "mar", "abr", "may", "jun",
                                 "jul", "ago", "sept", "oct", "nov", "dic"]]
                  + ["Total anual"])
        for ambito, montos in proyeccion.items():
            hp.append([Ambito(ambito).label] + list(montos))

    if metas is not None:
        ht = libro.create_sheet("Tablero de control")
        ht.append(["TABLERO DE CONTROL — RESULTADOS 2026"])
        ht.append([])
        ht.append(["Ámbito", "Meta anual de gestiones", "Gestiones realizadas"])
        for ambito, meta in metas.items():
            ht.append([Ambito(ambito).label, meta, 0])

    ruta = Path(tempfile.mkdtemp()) / "planilla.xlsx"
    libro.save(ruta)
    return ruta


class CalculoDelTableroTests(TestCase):
    """El avance mensual y el tablero de control, contra las reglas del Excel."""

    def setUp(self):
        # Una licitación presentada en abril y adjudicada en junio.
        Gestion.objects.create(
            anio=2026, ambito=Ambito.LICITACIONES, nombre="Diplomado",
            fecha_ingreso=date(2026, 4, 6), monto_postulado=Decimal("45911411"),
            estado=EstadoGestion.ADJUDICADA,
            fecha_resultado=date(2026, 6, 12), monto_adjudicado=Decimal("40000000"),
        )
        # Otra licitación en evaluación, ingresada en junio.
        Gestion.objects.create(
            anio=2026, ambito=Ambito.LICITACIONES, nombre="Curso de inglés",
            fecha_ingreso=date(2026, 6, 1), monto_postulado=Decimal("2800000"),
            estado=EstadoGestion.EN_EVALUACION,
        )
        # Una en preparación: gestionada, pero no presentada.
        Gestion.objects.create(
            anio=2026, ambito=Ambito.LICITACIONES, nombre="Curso ESCAR",
            fecha_ingreso=date(2026, 6, 22), monto_postulado=Decimal("8580000"),
            estado=EstadoGestion.EN_PREPARACION,
        )
        MetaAmbito.objects.create(
            anio=2026, ambito=Ambito.LICITACIONES, meta_gestiones=15)

    def _fila(self, clave, ambito=Ambito.LICITACIONES):
        bloques = {b["ambito"]: b for b in tablero.avance_mensual(2026)}
        return next(f for f in bloques[ambito]["filas"] if f["clave"] == clave)

    def test_lo_gestionado_se_cuenta_por_el_mes_de_ingreso(self):
        fila = self._fila("gestionadas")
        self.assertEqual(fila["meses"][3], 1)   # abril
        self.assertEqual(fila["meses"][5], 2)   # junio
        self.assertEqual(fila["total"], 3)

    def test_solo_lo_que_salio_de_la_casa_cuenta_como_presentado(self):
        # La que está "En preparación" se gestionó, pero no se presentó.
        self.assertEqual(self._fila("presentadas")["total"], 2)

    def test_lo_exitoso_se_cuenta_por_el_mes_de_resultado(self):
        fila = self._fila("exitosas")
        self.assertEqual(fila["meses"][3], 0)   # abril: ingresó, no se resolvió
        self.assertEqual(fila["meses"][5], 1)   # junio: se adjudicó
        self.assertEqual(fila["total"], 1)

    def test_el_monto_adjudicado_pesa_en_el_mes_del_resultado(self):
        fila = self._fila("monto_adjudicado")
        self.assertEqual(fila["meses"][3], Decimal("0"))
        self.assertEqual(fila["meses"][5], Decimal("40000000"))

    def test_una_gestion_sin_fecha_de_ingreso_no_cae_en_ningun_mes(self):
        """Es lo que hace la planilla con la fila de ejemplo de donaciones."""
        Gestion.objects.create(
            anio=2026, ambito=Ambito.DONACIONES, nombre="Ejemplo donación",
            estado=EstadoGestion.EN_IDENTIFICACION)
        self.assertEqual(self._fila("gestionadas", Ambito.DONACIONES)["total"], 0)
        self.assertEqual(Gestion.objects.filter(ambito=Ambito.DONACIONES).count(), 1)

    def test_un_convenio_se_cuenta_exitoso_cuando_se_suscribe_no_cuando_se_adjudica(self):
        Gestion.objects.create(
            anio=2026, ambito=Ambito.CONVENIOS, nombre="Convenio DOH",
            fecha_ingreso=date(2026, 3, 1), estado=EstadoGestion.ADJUDICADA,
            fecha_resultado=date(2026, 6, 30))
        self.assertEqual(self._fila("exitosas", Ambito.CONVENIOS)["total"], 0)

        Gestion.objects.filter(nombre="Convenio DOH").update(
            estado=EstadoGestion.SUSCRITA)
        self.assertEqual(self._fila("exitosas", Ambito.CONVENIOS)["total"], 1)

    def test_la_tasa_mensual_va_sobre_presentadas_y_la_del_tablero_sobre_gestionadas(self):
        """No es una inconsistencia mía: la planilla las calcula distinto."""
        self.assertEqual(self._fila("tasa")["total"], Decimal("1") / Decimal("2"))

        filas, _ = tablero.tablero_control(2026)
        licitaciones = next(f for f in filas if f["ambito"] == Ambito.LICITACIONES)
        self.assertEqual(licitaciones["tasa_exito"], Decimal("1") / Decimal("3"))

    def test_el_nivel_de_avance_es_gestiones_sobre_meta(self):
        filas, total = tablero.tablero_control(2026)
        licitaciones = next(f for f in filas if f["ambito"] == Ambito.LICITACIONES)
        self.assertEqual(licitaciones["avance"], Decimal("3") / Decimal("15"))
        self.assertEqual(total["meta"], 15)

    def test_sin_meta_ni_proyeccion_no_revienta_la_division(self):
        MetaAmbito.objects.all().delete()
        filas, total = tablero.tablero_control(2026)
        self.assertEqual(total["avance"], Decimal("0"))
        self.assertEqual(total["cumplimiento"], Decimal("0"))

    def test_el_cumplimiento_financiero_compara_contra_la_proyeccion(self):
        for mes in range(1, 13):
            ProyeccionMensual.objects.create(
                anio=2026, ambito=Ambito.LICITACIONES, mes=mes,
                monto=Decimal("10000000") if mes == 5 else Decimal("0"))

        filas, _ = tablero.tablero_control(2026)
        licitaciones = next(f for f in filas if f["ambito"] == Ambito.LICITACIONES)
        self.assertEqual(licitaciones["proyectado"], Decimal("10000000"))
        self.assertEqual(licitaciones["efectivo"], Decimal("40000000"))
        self.assertEqual(licitaciones["cumplimiento"], Decimal("4"))

    def test_el_semaforo_sigue_el_criterio_escrito_en_la_planilla(self):
        self.assertEqual(tablero.semaforo(Decimal("0.95")), "verde")
        self.assertEqual(tablero.semaforo(Decimal("0.90")), "verde")
        self.assertEqual(tablero.semaforo(Decimal("0.75")), "amarillo")
        self.assertEqual(tablero.semaforo(Decimal("0.70")), "amarillo")
        self.assertEqual(tablero.semaforo(Decimal("0.69")), "rojo")


class ArchivoRealTests(TestCase):
    """Contra el Excel que entregó la OCT, no contra un archivo de juguete."""

    @classmethod
    def setUpTestData(cls):
        ImportadorPlanilla(RUTA_PLANILLA).ejecutar(aplicar=True)

    def test_se_cargan_las_dieciseis_filas_del_registro(self):
        self.assertEqual(Gestion.objects.filter(anio=2026).count(), 16)

    def test_las_gestiones_por_ambito_calzan_con_el_tablero_del_excel(self):
        filas, total = tablero.tablero_control(2026)
        por_ambito = {f["ambito"]: f["gestiones"] for f in filas}
        self.assertEqual(por_ambito[Ambito.PROYECTOS], 1)
        self.assertEqual(por_ambito[Ambito.LICITACIONES], 9)
        self.assertEqual(por_ambito[Ambito.CONVENIOS], 5)
        self.assertEqual(por_ambito[Ambito.DONACIONES], 0)
        self.assertEqual(total["gestiones"], 15)

    def test_las_metas_salen_de_la_hoja_del_tablero(self):
        metas = {m.ambito: m.meta_gestiones for m in MetaAmbito.objects.all()}
        self.assertEqual(metas[Ambito.PROYECTOS], 10)
        self.assertEqual(metas[Ambito.LICITACIONES], 15)
        self.assertEqual(metas[Ambito.CONVENIOS], 8)
        self.assertEqual(metas[Ambito.DONACIONES], 20)

    def test_la_proyeccion_anual_da_el_mismo_total_que_el_excel(self):
        _, total = tablero.proyeccion_por_mes(2026)
        self.assertEqual(total["total"], Decimal("1515412753"))

    def test_los_montos_por_ambito_calzan_con_la_proyeccion_del_excel(self):
        filas, _ = tablero.proyeccion_por_mes(2026)
        por_ambito = {f["ambito"]: f["total"] for f in filas}
        self.assertEqual(por_ambito[Ambito.PROYECTOS], Decimal("559647902"))
        self.assertEqual(por_ambito[Ambito.LICITACIONES], Decimal("175070451"))
        self.assertEqual(por_ambito[Ambito.CONVENIOS], Decimal("713720000"))
        self.assertEqual(por_ambito[Ambito.DONACIONES], Decimal("66974400"))

    def test_el_monto_ofertado_mensual_calza_con_el_avance_del_excel(self):
        bloques = {b["ambito"]: b for b in tablero.avance_mensual(2026)}
        fila = next(f for f in bloques[Ambito.LICITACIONES]["filas"]
                    if f["clave"] == "monto_postulado")
        self.assertEqual(fila["meses"][3], Decimal("45911411"))   # abr
        self.assertEqual(fila["meses"][4], Decimal("16550000"))   # may
        self.assertEqual(fila["meses"][5], Decimal("22152469"))   # jun
        self.assertEqual(fila["total"], Decimal("84613880"))

    def test_volver_a_cargar_el_mismo_archivo_no_duplica_nada(self):
        resultado = ImportadorPlanilla(RUTA_PLANILLA).ejecutar(aplicar=True)
        self.assertEqual(Gestion.objects.filter(anio=2026).count(), 16)
        self.assertFalse(
            [c for c in resultado.relevantes if c.entidad == "Gestión"],
            "la segunda pasada no debería tocar ninguna gestión")


class IdentidadYPodaTests(TestCase):
    """Que el archivo actualizado se refleje sin crear copias."""

    def fila(self, **campos):
        base = {
            "COD": "C-001", "Ámbito": "Convenios", "Tipo de iniciativa": "Institución pública",
            "Nombre de la iniciativa": "Capacitación a servicios",
            "Institución": "DOH", "Fecha de ingreso": date(2026, 3, 1),
            "Monto postulado": 185200000, "Estado": "En evaluación",
            "Responsable": "OCT",
        }
        base.update(campos)
        return base

    def test_el_codigo_manda_sobre_el_nombre(self):
        ImportadorPlanilla(planilla([self.fila()])).ejecutar(aplicar=True)
        ImportadorPlanilla(planilla([
            self.fila(**{"Nombre de la iniciativa": "Capacitación a los servicios DOH"})
        ])).ejecutar(aplicar=True)

        self.assertEqual(Gestion.objects.count(), 1)
        self.assertEqual(
            Gestion.objects.get().nombre, "Capacitación a los servicios DOH")

    def test_sin_codigo_se_reconoce_por_el_nombre_aunque_le_corrijan_la_redaccion(self):
        ImportadorPlanilla(planilla([self.fila(COD="N/A")])).ejecutar(aplicar=True)
        ImportadorPlanilla(planilla([self.fila(
            COD="N/A",
            **{"Nombre de la iniciativa": "Capacitacion a servicios"},
        )])).ejecutar(aplicar=True)

        self.assertEqual(Gestion.objects.count(), 1)

    def test_dos_codigos_distintos_son_dos_gestiones_aunque_se_llamen_igual(self):
        """Pasa de verdad: C-002 y C-003 son el mismo texto, distinto expediente."""
        ImportadorPlanilla(planilla([
            self.fila(COD="C-002"), self.fila(COD="C-003"),
        ])).ejecutar(aplicar=True)
        self.assertEqual(Gestion.objects.count(), 2)

    def test_lo_que_ya_no_viene_en_el_archivo_se_elimina(self):
        ImportadorPlanilla(planilla([
            self.fila(COD="C-001"), self.fila(COD="C-002"),
        ])).ejecutar(aplicar=True)
        ImportadorPlanilla(planilla([self.fila(COD="C-001")])).ejecutar(aplicar=True)

        self.assertEqual(
            list(Gestion.objects.values_list("codigo", flat=True)), ["C-001"])

    def test_lo_cargado_a_mano_nunca_lo_borra_el_importador(self):
        Gestion.objects.create(
            anio=2026, ambito=Ambito.PROYECTOS, nombre="Idea propia",
            origen=Origen.MANUAL, estado=EstadoGestion.EN_IDENTIFICACION)

        ImportadorPlanilla(planilla([self.fila()])).ejecutar(aplicar=True)

        self.assertTrue(Gestion.objects.filter(nombre="Idea propia").exists())

    def test_se_puede_pedir_que_no_pode(self):
        ImportadorPlanilla(planilla([
            self.fila(COD="C-001"), self.fila(COD="C-002"),
        ])).ejecutar(aplicar=True)
        ImportadorPlanilla(
            planilla([self.fila(COD="C-001")]), podar=False).ejecutar(aplicar=True)

        self.assertEqual(Gestion.objects.count(), 2)

    def test_la_vista_previa_no_guarda_nada(self):
        resultado = ImportadorPlanilla(planilla([self.fila()])).ejecutar(aplicar=False)

        self.assertFalse(resultado.aplicado)
        self.assertTrue(resultado.hay_cambios)
        self.assertEqual(Gestion.objects.count(), 0)

    def test_una_fila_con_ambito_desconocido_se_descarta_sin_voltear_la_carga(self):
        resultado = ImportadorPlanilla(planilla([
            self.fila(), self.fila(COD="X-1", **{"Ámbito": "Otra cosa"}),
        ])).ejecutar(aplicar=True)

        self.assertEqual(Gestion.objects.count(), 1)
        self.assertEqual(len(resultado.descartadas), 1)

    def test_un_archivo_sin_la_hoja_de_registro_avisa_en_vez_de_reventar(self):
        libro = Workbook()
        libro.active.title = "Otra hoja"
        ruta = Path(tempfile.mkdtemp()) / "malo.xlsx"
        libro.save(ruta)

        with self.assertRaises(ErrorImportacion) as caso:
            ImportadorPlanilla(ruta).ejecutar()
        self.assertIn("Registro iniciativas", str(caso.exception))

    def test_el_anio_sale_del_titulo_de_la_hoja(self):
        ImportadorPlanilla(planilla(
            [self.fila()], titulo="REGISTRO DE INICIATIVAS 2027")).ejecutar(aplicar=True)
        self.assertEqual(Gestion.objects.get().anio, 2027)


class EdicionEnElSistemaTests(TestCase):
    """El corazón del encargo: el Excel no pisa lo corregido acá sin preguntar."""

    def setUp(self):
        self.usuario = User.objects.create_user("ana", password="x")
        self.ruta = planilla([{
            "COD": "C-001", "Ámbito": "Convenios",
            "Nombre de la iniciativa": "Capacitación DOH",
            "Institución": "DOH", "Fecha de ingreso": date(2026, 3, 1),
            "Monto postulado": 185200000, "Estado": "En evaluación",
            "Responsable": "OCT",
        }])
        ImportadorPlanilla(self.ruta).ejecutar(aplicar=True)

        self.gestion = Gestion.objects.get()
        self.gestion.estado = EstadoGestion.SUSCRITA
        self.gestion.marcar_editada(["estado"], self.usuario)
        self.gestion.save()

    def test_por_defecto_se_conserva_lo_editado(self):
        resultado = ImportadorPlanilla(self.ruta).ejecutar(aplicar=True)

        self.gestion.refresh_from_db()
        self.assertEqual(self.gestion.estado, EstadoGestion.SUSCRITA)
        self.assertEqual(len(resultado.conflictos), 1)
        self.assertFalse(resultado.conflictos[0].resuelto_con_excel)

    def test_el_conflicto_dice_exactamente_que_campo_choca(self):
        resultado = ImportadorPlanilla(self.ruta).ejecutar()

        conflicto = resultado.conflictos[0]
        self.assertEqual([d.campo for d in conflicto.diferencias], ["estado"])
        self.assertEqual(conflicto.diferencias[0].actual, "Suscrita")
        self.assertEqual(conflicto.diferencias[0].propuesto, "En evaluación")
        self.assertEqual(conflicto.editado_por, "ana")

    def test_marcando_la_casilla_el_excel_reemplaza_lo_editado(self):
        previo = ImportadorPlanilla(self.ruta).ejecutar()
        decisiones = {c.clave: True for c in previo.conflictos}

        ImportadorPlanilla(self.ruta, decisiones=decisiones).ejecutar(aplicar=True)

        self.gestion.refresh_from_db()
        self.assertEqual(self.gestion.estado, EstadoGestion.EN_EVALUACION)

    def test_al_aceptar_el_excel_el_campo_deja_de_estar_protegido(self):
        previo = ImportadorPlanilla(self.ruta).ejecutar()
        decisiones = {c.clave: True for c in previo.conflictos}
        ImportadorPlanilla(self.ruta, decisiones=decisiones).ejecutar(aplicar=True)

        self.gestion.refresh_from_db()
        self.assertEqual(self.gestion.campos_editados, [])
        self.assertFalse(self.gestion.editada_en_sistema)

    def test_un_campo_no_editado_se_actualiza_sin_preguntar(self):
        """Proteger un campo no congela la fila entera."""
        libro = planilla([{
            "COD": "C-001", "Ámbito": "Convenios",
            "Nombre de la iniciativa": "Capacitación DOH",
            "Institución": "Dirección de Obras Hidráulicas",
            "Fecha de ingreso": date(2026, 3, 1),
            "Monto postulado": 190000000, "Estado": "En evaluación",
            "Responsable": "OCT",
        }])
        ImportadorPlanilla(libro).ejecutar(aplicar=True)

        self.gestion.refresh_from_db()
        self.assertEqual(self.gestion.monto_postulado, Decimal("190000000"))
        self.assertEqual(self.gestion.institucion, "Dirección de Obras Hidráulicas")
        self.assertEqual(self.gestion.estado, EstadoGestion.SUSCRITA)

    def test_una_gestion_editada_que_desaparece_del_archivo_no_se_borra_sola(self):
        vacia = planilla([{
            "COD": "C-999", "Ámbito": "Convenios",
            "Nombre de la iniciativa": "Otra cosa", "Estado": "En preparación",
        }])
        resultado = ImportadorPlanilla(vacia).ejecutar(aplicar=True)

        self.assertTrue(Gestion.objects.filter(pk=self.gestion.pk).exists())
        eliminaciones = [c for c in resultado.conflictos if c.es_eliminacion]
        self.assertEqual(len(eliminaciones), 1)

    def test_pero_se_puede_pedir_que_se_borre(self):
        vacia = planilla([{
            "COD": "C-999", "Ámbito": "Convenios",
            "Nombre de la iniciativa": "Otra cosa", "Estado": "En preparación",
        }])
        previo = ImportadorPlanilla(vacia).ejecutar()
        decisiones = {c.clave: True for c in previo.conflictos}

        ImportadorPlanilla(vacia, decisiones=decisiones).ejecutar(aplicar=True)

        self.assertFalse(Gestion.objects.filter(pk=self.gestion.pk).exists())

    def test_dos_ediciones_seguidas_protegen_las_dos(self):
        self.gestion.monto_postulado = Decimal("1")
        self.gestion.marcar_editada(["monto_postulado"], self.usuario)
        self.gestion.save()

        self.gestion.refresh_from_db()
        self.assertEqual(self.gestion.campos_editados, ["estado", "monto_postulado"])


# La carga guarda el .xlsx en MEDIA_ROOT mientras espera la confirmación. Sin
# esto, cada corrida de las pruebas dejaba archivos sueltos en la carpeta de
# medios del proyecto.
@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class PantallasTests(TestCase):
    """Que las páginas respondan y que editar deje el rastro que protege."""

    def setUp(self):
        self.usuario = User.objects.create_user("ana", password="clave-larga-123")
        self.client.force_login(self.usuario)
        ImportadorPlanilla(RUTA_PLANILLA).ejecutar(aplicar=True)

    def test_el_tablero_muestra_los_totales(self):
        respuesta = self.client.get(reverse("oct:tablero_maestro"))
        self.assertEqual(respuesta.status_code, 200)
        self.assertContains(respuesta, "Tablero maestro de resultados")
        self.assertEqual(respuesta.context["total"]["gestiones"], 15)

    def test_el_registro_lista_las_gestiones_y_filtra(self):
        respuesta = self.client.get(reverse("oct:tablero_registro"))
        self.assertEqual(len(respuesta.context["gestiones"]), 16)

        respuesta = self.client.get(
            reverse("oct:tablero_registro"), {"ambito": Ambito.CONVENIOS})
        self.assertEqual(len(respuesta.context["gestiones"]), 5)

        # En minúscula y con tilde, contra un nombre que en la planilla está
        # en mayúsculas: "CURSO DE INGLÉS".
        respuesta = self.client.get(reverse("oct:tablero_registro"), {"q": "inglés"})
        self.assertEqual(len(respuesta.context["gestiones"]), 1)

        # Y sin tilde también, que es como se escribe apurado.
        respuesta = self.client.get(reverse("oct:tablero_registro"), {"q": "ingles"})
        self.assertEqual(len(respuesta.context["gestiones"]), 1)

    def test_editar_desde_la_pantalla_marca_el_campo_como_protegido(self):
        gestion = Gestion.objects.filter(codigo="C-001").first()
        datos = {
            "codigo": gestion.codigo, "ambito": gestion.ambito,
            "tipo": gestion.tipo, "nombre": gestion.nombre,
            "institucion": gestion.institucion,
            "fecha_ingreso": gestion.fecha_ingreso.isoformat(),
            "monto_postulado": "200000000",
            "estado": EstadoGestion.SUSCRITA,
            "fecha_resultado": gestion.fecha_resultado.isoformat(),
            "monto_adjudicado": "0", "responsable": gestion.responsable,
            "observaciones": gestion.observaciones,
        }
        respuesta = self.client.post(
            reverse("oct:tablero_gestion_editar", args=[gestion.pk]), datos)
        self.assertEqual(respuesta.status_code, 302)

        gestion.refresh_from_db()
        self.assertEqual(gestion.campos_editados, ["estado", "monto_postulado"])
        self.assertEqual(gestion.editado_por, self.usuario)

    def test_una_gestion_creada_a_mano_no_queda_marcada_como_editada(self):
        """Marcar campos solo tiene sentido si el Excel gobierna la fila."""
        respuesta = self.client.post(reverse("oct:tablero_gestion_nueva"), {
            "codigo": "M-1", "ambito": Ambito.PROYECTOS, "tipo": "",
            "nombre": "Idea nueva", "institucion": "", "fecha_ingreso": "",
            "monto_postulado": "0", "estado": EstadoGestion.EN_IDENTIFICACION,
            "fecha_resultado": "", "monto_adjudicado": "0",
            "responsable": "", "observaciones": "",
        })
        self.assertEqual(respuesta.status_code, 302)

        creada = Gestion.objects.get(codigo="M-1")
        self.assertEqual(creada.origen, Origen.MANUAL)
        self.assertEqual(creada.campos_editados, [])

    def test_la_fecha_de_resultado_no_puede_ser_anterior_a_la_de_ingreso(self):
        respuesta = self.client.post(reverse("oct:tablero_gestion_nueva"), {
            "codigo": "M-2", "ambito": Ambito.PROYECTOS, "tipo": "",
            "nombre": "Al revés", "institucion": "",
            "fecha_ingreso": "2026-06-01", "monto_postulado": "0",
            "estado": EstadoGestion.PRESENTADA,
            "fecha_resultado": "2026-01-01", "monto_adjudicado": "0",
            "responsable": "", "observaciones": "",
        })
        self.assertEqual(respuesta.status_code, 200)
        self.assertFalse(Gestion.objects.filter(codigo="M-2").exists())

    def test_las_pantallas_de_alta_y_baja_se_dibujan(self):
        gestion = Gestion.objects.first()
        for url in [reverse("oct:tablero_gestion_nueva"),
                    reverse("oct:tablero_gestion_editar", args=[gestion.pk]),
                    reverse("oct:tablero_gestion_eliminar", args=[gestion.pk]),
                    reverse("oct:tablero_parametros"),
                    reverse("oct:tablero_importar")]:
            self.assertEqual(self.client.get(url).status_code, 200, url)

    def test_eliminar_saca_la_gestion_y_el_tablero_lo_refleja(self):
        gestion = Gestion.objects.filter(codigo="C-001").first()
        respuesta = self.client.post(
            reverse("oct:tablero_gestion_eliminar", args=[gestion.pk]))

        self.assertEqual(respuesta.status_code, 302)
        self.assertFalse(Gestion.objects.filter(pk=gestion.pk).exists())
        _, total = tablero.tablero_control(2026)
        self.assertEqual(total["gestiones"], 14)

    def test_los_parametros_se_guardan(self):
        datos = {}
        for ambito, _ in Ambito.choices:
            datos[f"meta-{ambito}"] = "12"
            for mes in range(1, 13):
                datos[f"proy-{ambito}-{mes}"] = "1000000"

        respuesta = self.client.post(reverse("oct:tablero_parametros"), datos)
        self.assertEqual(respuesta.status_code, 302)

        self.assertEqual(
            MetaAmbito.objects.get(anio=2026, ambito=Ambito.PROYECTOS).meta_gestiones, 12)
        _, total = tablero.proyeccion_por_mes(2026)
        self.assertEqual(total["total"], Decimal("48000000"))

    def test_la_carga_muestra_la_vista_previa_sin_guardar(self):
        Gestion.objects.all().delete()

        with RUTA_PLANILLA.open("rb") as archivo:
            respuesta = self.client.post(reverse("oct:tablero_importar"), {
                "archivo": archivo, "podar": "on",
            })

        self.assertEqual(respuesta.status_code, 200)
        self.assertTrue(respuesta.context["previsualizacion"])
        self.assertEqual(Gestion.objects.count(), 0)

    def test_y_al_confirmar_se_aplica(self):
        Gestion.objects.all().delete()

        with RUTA_PLANILLA.open("rb") as archivo:
            self.client.post(reverse("oct:tablero_importar"), {
                "archivo": archivo, "podar": "on"})
        respuesta = self.client.post(
            reverse("oct:tablero_importar"), {"confirmar": "1"})

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(Gestion.objects.count(), 16)

    def test_la_pantalla_de_carga_ofrece_las_casillas_de_los_conflictos(self):
        gestion = Gestion.objects.filter(codigo="C-001").first()
        gestion.estado = EstadoGestion.SUSCRITA
        gestion.marcar_editada(["estado"], self.usuario)
        gestion.save()

        with RUTA_PLANILLA.open("rb") as archivo:
            respuesta = self.client.post(reverse("oct:tablero_importar"), {
                "archivo": archivo, "podar": "on"})

        conflictos = respuesta.context["resultado"].conflictos
        self.assertEqual(len(conflictos), 1)
        self.assertContains(respuesta, f'value="{conflictos[0].clave}"')
        self.assertContains(respuesta, "se conserva")

    def test_al_confirmar_marcando_la_casilla_el_excel_manda(self):
        gestion = Gestion.objects.filter(codigo="C-001").first()
        gestion.estado = EstadoGestion.SUSCRITA
        gestion.marcar_editada(["estado"], self.usuario)
        gestion.save()

        with RUTA_PLANILLA.open("rb") as archivo:
            self.client.post(reverse("oct:tablero_importar"), {
                "archivo": archivo, "podar": "on"})
        self.client.post(reverse("oct:tablero_importar"), {
            "confirmar": "1", "usar_excel": [f"g{gestion.pk}"]})

        gestion.refresh_from_db()
        self.assertEqual(gestion.estado, EstadoGestion.EN_EVALUACION)

    def test_y_sin_marcarla_se_conserva(self):
        gestion = Gestion.objects.filter(codigo="C-001").first()
        gestion.estado = EstadoGestion.SUSCRITA
        gestion.marcar_editada(["estado"], self.usuario)
        gestion.save()

        with RUTA_PLANILLA.open("rb") as archivo:
            self.client.post(reverse("oct:tablero_importar"), {
                "archivo": archivo, "podar": "on"})
        self.client.post(reverse("oct:tablero_importar"), {"confirmar": "1"})

        gestion.refresh_from_db()
        self.assertEqual(gestion.estado, EstadoGestion.SUSCRITA)

    def test_las_pantallas_piden_sesion(self):
        self.client.logout()
        for nombre in ["tablero_maestro", "tablero_registro",
                       "tablero_parametros", "tablero_importar",
                       "tablero_exportar_excel", "tablero_informe"]:
            respuesta = self.client.get(reverse(f"oct:{nombre}"))
            self.assertEqual(respuesta.status_code, 302, nombre)


class ExportacionTests(TestCase):
    """El Excel que sale tiene que poder volver a entrar."""

    HOJAS = ["Registro iniciativas", "Proyección financiera",
             "Avance mensual 2026", "Tablero de control", "Instrucciones"]

    def setUp(self):
        self.usuario = User.objects.create_user("ana", password="clave-larga-123")
        self.client.force_login(self.usuario)
        ImportadorPlanilla(RUTA_PLANILLA).ejecutar(aplicar=True)

    def descargar(self):
        respuesta = self.client.get(
            reverse("oct:tablero_exportar_excel"), {"anio": 2026})
        self.assertEqual(respuesta.status_code, 200)
        ruta = Path(tempfile.mkdtemp()) / "exportado.xlsx"
        ruta.write_bytes(respuesta.content)
        return respuesta, ruta

    def test_el_archivo_sale_con_nombre_y_tipo_de_excel(self):
        respuesta, _ = self.descargar()
        self.assertIn("spreadsheetml", respuesta["Content-Type"])
        self.assertIn("Tablero_Resultados_OCT_2026.xlsx",
                      respuesta["Content-Disposition"])

    def test_trae_las_cinco_hojas_y_las_dieciseis_filas(self):
        _, ruta = self.descargar()
        libro = openpyxl.load_workbook(ruta)

        self.assertEqual(libro.sheetnames, self.HOJAS)

        hoja = libro["Registro iniciativas"]
        filas = [f for f in hoja.iter_rows(min_row=4, values_only=True) if f[3]]
        self.assertEqual(len(filas), 16)

        # En el orden del tablero, no en el alfabético de la base.
        self.assertEqual(
            [f[1] for f in filas][:2], ["Proyectos", "Licitaciones"])
        self.assertEqual([f[1] for f in filas][-1], "Donaciones")

    def test_los_totales_del_tablero_llegan_al_archivo(self):
        _, ruta = self.descargar()
        libro = openpyxl.load_workbook(ruta)

        # Fila TOTAL del tablero de control: 4 ámbitos + la fila 4 de inicio.
        control = libro["Tablero de control"]
        total = [c.value for c in control[8]]
        self.assertEqual(total[0], "TOTAL")
        self.assertEqual(total[1], 53)                  # meta anual
        self.assertEqual(total[2], 15)                  # gestiones
        self.assertEqual(total[6], 1515412753.0)        # monto proyectado

    def test_el_avance_mensual_lleva_los_montos_por_mes(self):
        _, ruta = self.descargar()
        hoja = openpyxl.load_workbook(ruta)["Avance mensual 2026"]

        fila = next(
            f for f in hoja.iter_rows(values_only=True)
            if f[0] == "Monto ofertado"
        )
        self.assertEqual(fila[4], 45911411.0)    # abril
        self.assertEqual(fila[13], 84613880.0)   # total del año

    def test_lo_exportado_se_puede_volver_a_importar_sin_cambiar_nada(self):
        """Es la prueba que importa: el archivo sirve de respaldo real."""
        _, ruta = self.descargar()

        resultado = ImportadorPlanilla(ruta).ejecutar(aplicar=True)

        self.assertEqual(Gestion.objects.filter(anio=2026).count(), 16)
        self.assertEqual(
            [c for c in resultado.relevantes if c.entidad == "Gestión"], [],
            "reimportar lo exportado no debería tocar ninguna gestión")
        _, total = tablero.tablero_control(2026)
        self.assertEqual(total["gestiones"], 15)
        self.assertEqual(total["meta"], 53)

    def test_el_archivo_conserva_las_listas_desplegables(self):
        """Se sigue pudiendo completar a mano sin inventar estados."""
        _, ruta = self.descargar()
        hoja = openpyxl.load_workbook(ruta)["Registro iniciativas"]

        listas = {dv.formula1 for dv in hoja.data_validations.dataValidation}
        self.assertTrue(any("Licitaciones" in f for f in listas))
        self.assertTrue(any("No adjudicada" in f for f in listas))

    def test_una_gestion_editada_se_marca_en_el_archivo(self):
        gestion = Gestion.objects.filter(codigo="C-001").first()
        gestion.marcar_editada(["estado"], self.usuario)
        gestion.save()

        _, ruta = self.descargar()
        hoja = openpyxl.load_workbook(ruta)["Registro iniciativas"]
        fila = next(f for f in hoja.iter_rows(min_row=4, values_only=True)
                    if f[0] == "C-001")
        self.assertEqual(fila[14], "estado")

    def test_el_informe_se_dibuja_con_los_totales(self):
        respuesta = self.client.get(reverse("oct:tablero_informe"), {"anio": 2026})

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta.context["total"]["gestiones"], 15)
        self.assertContains(respuesta, "Tablero maestro de resultados")
        self.assertContains(respuesta, "Imprimir / Guardar PDF")

    def test_el_informe_de_un_ambito_recalcula_solo_ese_ambito(self):
        respuesta = self.client.get(reverse("oct:tablero_informe"), {
            "anio": 2026, "ambito": Ambito.CONVENIOS})

        self.assertEqual(respuesta.context["n_gestiones"], 5)
        self.assertEqual(len(respuesta.context["filas"]), 1)
        self.assertEqual(respuesta.context["total"]["gestiones"], 5)
        self.assertEqual(respuesta.context["total"]["meta"], 8)
        self.assertContains(respuesta, "Convenios")

    def test_un_ambito_inventado_no_rompe_el_informe(self):
        respuesta = self.client.get(reverse("oct:tablero_informe"), {
            "anio": 2026, "ambito": "LO-QUE-SEA"})

        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(respuesta.context["filtro"], "")
        self.assertEqual(respuesta.context["n_gestiones"], 16)

    def test_el_informe_puede_abrirse_listo_para_imprimir(self):
        respuesta = self.client.get(reverse("oct:tablero_informe"), {"print": "1"})
        self.assertTrue(respuesta.context["auto_print"])
        self.assertContains(respuesta, "window.print()")


class GraficosDelInformeTests(TestCase):
    """Los gráficos van dibujados en el servidor para que salgan impresos."""

    # Atributos del SVG que llevan un número: si uno queda mal formateado, la
    # figura desaparece sin ningún aviso.
    NUMERICOS = re.compile(r'\s(?:width|height|x|y|x1|x2|y1|y2)="([^"]*)"')

    def setUp(self):
        self.usuario = User.objects.create_user("ana", password="clave-larga-123")
        self.client.force_login(self.usuario)
        ImportadorPlanilla(RUTA_PLANILLA).ejecutar(aplicar=True)

    def test_el_informe_trae_los_tres_graficos_como_svg(self):
        respuesta = self.client.get(reverse("oct:tablero_informe"))
        html = respuesta.content.decode()

        self.assertEqual(html.count('<svg class="graf"'), 3)
        # Dibujados en el servidor: sin canvas y sin depender de un CDN.
        self.assertNotIn("<canvas", html)
        self.assertNotIn("chart.js", html.lower())

    def test_ninguna_coordenada_sale_con_coma_decimal(self):
        """La trampa de es-CL: un 150.25 dibujado como «150,25» rompe el SVG."""
        respuesta = self.client.get(reverse("oct:tablero_informe"))
        html = respuesta.content.decode()
        svgs = re.findall(r"<svg class=\"graf\".*?</svg>", html, re.S)
        self.assertTrue(svgs)

        for svg in svgs:
            for valor in self.NUMERICOS.findall(svg):
                try:
                    float(valor)
                except ValueError:
                    self.fail(f"coordenada no numérica en el SVG: {valor!r}")

    def test_la_barra_completa_es_mas_larga_donde_la_meta_es_mayor(self):
        filas, _ = tablero.tablero_control(2026)
        grafico = graficos.barras_de_avance(filas)

        largos = {
            f["etiqueta"]: float(f["ancho_hecho"]) + float(f["ancho_pendiente"])
            for f in grafico["filas"]
        }
        # Donaciones tiene meta 20 y Convenios 8.
        self.assertGreater(largos["Donaciones"], largos["Convenios"])

    def test_pasarse_de_la_meta_se_ve_y_se_marca(self):
        """Recortar la barra al riel diría que se cumplió justo, y es mentira."""
        filas = [
            {"etiqueta": "Proyectos", "meta": 4, "gestiones": 9,
             "avance_pct": Decimal("225")},
            {"etiqueta": "Convenios", "meta": 9, "gestiones": 4,
             "avance_pct": Decimal("44")},
        ]
        pasado, corto = graficos.barras_de_avance(filas)["filas"]

        self.assertTrue(pasado["supera"])
        self.assertEqual(pasado["texto"], "9 de 4")
        self.assertEqual(float(pasado["ancho_pendiente"]), 0)
        # La marca de la meta queda dentro de la barra, no al final.
        self.assertLess(float(pasado["x_meta"]),
                        float(pasado["x"]) + float(pasado["ancho_hecho"]))

        self.assertFalse(corto["supera"])
        self.assertGreater(float(corto["ancho_pendiente"]), 0)

    def test_un_cero_deja_una_hebra_visible(self):
        filas = [{"etiqueta": "Donaciones", "meta": 20, "gestiones": 0,
                  "avance_pct": Decimal("0")}]
        fila = graficos.barras_de_avance(filas)["filas"][0]
        self.assertGreater(float(fila["ancho_hecho"]), 0)

    def test_el_eje_del_grafico_mensual_queda_en_numeros_redondos(self):
        _, total = tablero.proyeccion_por_mes(2026)
        grafico = graficos.columnas_mensuales(
            tablero.MESES, total["meses"], [Decimal("0")] * 12)

        etiquetas = [r["etiqueta"] for r in grafico["referencias"]]
        self.assertEqual(
            etiquetas, ["$0", "$200M", "$400M", "$600M", "$800M"])

    def test_avisa_cuando_todavia_no_hay_ingreso_efectivo(self):
        respuesta = self.client.get(reverse("oct:tablero_informe"))
        self.assertFalse(respuesta.context["grafico_meses"]["hay_efectivo"])
        self.assertContains(respuesta, "Todavía no hay ingreso efectivo")

    def test_y_deja_de_avisar_cuando_lo_hay(self):
        gestion = Gestion.objects.filter(ambito=Ambito.LICITACIONES).first()
        gestion.estado = EstadoGestion.ADJUDICADA
        gestion.fecha_resultado = date(2026, 6, 12)
        gestion.monto_adjudicado = Decimal("40000000")
        gestion.save()

        respuesta = self.client.get(reverse("oct:tablero_informe"))
        grafico = respuesta.context["grafico_meses"]

        self.assertTrue(grafico["hay_efectivo"])
        junio = grafico["grupos"][5]
        self.assertGreater(float(junio["alto_efectivo"]), 0)
        self.assertNotContains(respuesta, "Todavía no hay ingreso efectivo")

    def test_sin_datos_no_se_dibuja_el_panel_de_estados(self):
        self.assertIsNone(graficos.barras_por_estado([("Presentada", 0)]))
        self.assertIsNone(graficos.barras_por_estado([]))

    def test_un_ambito_solo_tambien_dibuja_sus_graficos(self):
        respuesta = self.client.get(reverse("oct:tablero_informe"), {
            "ambito": Ambito.CONVENIOS})

        self.assertEqual(len(respuesta.context["grafico_avance"]["filas"]), 1)
        self.assertEqual(
            respuesta.context["grafico_avance"]["filas"][0]["texto"], "5 de 8")
