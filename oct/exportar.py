"""Salidas del tablero maestro OCT: Excel e informe imprimible.

Dos vistas, ambas de solo lectura y sin cálculos propios — se apoyan en
``oct/tablero.py``, que es donde viven las reglas:

* ``exportar_tablero_excel`` — un .xlsx con la misma forma que la planilla
  original. **El archivo que sale se puede volver a subir**: las hojas, los
  títulos y los rótulos son los que el importador reconoce, así que sirve de
  respaldo y de punto de partida para la siguiente actualización.
* ``informe_tablero`` — una página optimizada para imprimir o guardar como PDF
  desde el navegador. No se agregó ninguna librería de PDF al servidor, por el
  mismo motivo que en ``proyectos``: el despliegue corre en una Raspberry Pi.

Por eso la exportación a Excel **no acepta filtros**: un archivo con la mitad
de las filas, subido de vuelta con la poda activada, borraría el resto. El
informe sí filtra por ámbito, porque no vuelve a entrar al sistema.
"""

import io
from datetime import date
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from . import graficos, tablero
from .models import Ambito, EstadoGestion, Gestion
from .views import _anio_pedido

# Formatos de celda. El peso chileno va sin decimales.
FMT_CLP = '#,##0'
FMT_PCT = '0.0%'

AZUL = '2563EB'       # el mismo encabezado que usan los otros exports
GRIS = 'E2E8F0'


def _f(valor):
    """Decimal/None -> float, que es lo que openpyxl escribe en una celda."""
    if valor is None:
        return 0.0
    if isinstance(valor, Decimal):
        return float(valor)
    return float(valor)


# =========================
# EXCEL
# =========================

@login_required
def exportar_tablero_excel(request):
    """Descarga el tablero completo de un año como .xlsx."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        return HttpResponse(
            "openpyxl no está instalado. Ejecute: pip install openpyxl",
            status=500,
        )

    anio, _ = _anio_pedido(request)

    # Proyectos, Licitaciones, Convenios, Donaciones — el orden del tablero, no
    # el alfabético que devolvería la base.
    orden = {valor: i for i, valor in enumerate(Ambito.values)}
    gestiones = sorted(
        Gestion.objects.filter(anio=anio),
        key=lambda g: (orden[g.ambito], g.fecha_ingreso or date.max, g.nombre),
    )
    sufijo = str(anio)[-2:]

    blanco = Font(bold=True, color='FFFFFF')
    negrita = Font(bold=True)
    relleno = PatternFill('solid', fgColor=AZUL)
    relleno_gris = PatternFill('solid', fgColor=GRIS)
    derecha = Alignment(horizontal='right')
    centro = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb = Workbook()

    def encabezar(ws, titulo, columnas, fila=3):
        """Título en la fila 1, blanco en la 2 y encabezados en la 3.

        Es la disposición de la planilla original, y es la que el importador
        busca cuando el archivo vuelve a subirse.
        """
        ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=13)
        for col, texto in enumerate(columnas, 1):
            celda = ws.cell(row=fila, column=col, value=texto)
            celda.font = blanco
            celda.fill = relleno
            celda.alignment = centro
        ws.freeze_panes = ws.cell(row=fila + 1, column=1)

    def dinero(ws, fila, col, valor):
        celda = ws.cell(row=fila, column=col, value=_f(valor))
        celda.number_format = FMT_CLP
        celda.alignment = derecha
        return celda

    def anchos(ws, medidas):
        for i, ancho in enumerate(medidas, 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

    # ---------- HOJA 1: REGISTRO ----------
    # Va primero y con el nombre exacto porque es la hoja que el importador
    # exige; las demás son derivadas.
    ws = wb.active
    ws.title = "Registro iniciativas"
    columnas = [
        "COD", "Ámbito", "Tipo de iniciativa", "Nombre de la iniciativa",
        "Institución", "Fecha de ingreso", "Mes de ingreso", "Monto postulado",
        "Estado", "Fecha de resultado", "Mes de resultado", "Monto adjudicado",
        "Responsable", "Observaciones", "Editado en el sistema",
    ]
    encabezar(ws, f"REGISTRO DE INICIATIVAS {anio}", columnas)

    nombre_mes = {n: nombre for n, nombre in tablero.MESES}
    for i, g in enumerate(gestiones, start=4):
        ws.cell(row=i, column=1, value=g.codigo or "N/A")
        ws.cell(row=i, column=2, value=g.get_ambito_display())
        ws.cell(row=i, column=3, value=g.tipo)
        ws.cell(row=i, column=4, value=g.nombre)
        ws.cell(row=i, column=5, value=g.institucion)
        if g.fecha_ingreso:
            celda = ws.cell(row=i, column=6, value=g.fecha_ingreso)
            celda.number_format = 'DD-MM-YYYY'
            ws.cell(row=i, column=7, value=f"{nombre_mes[g.mes_ingreso]}-{sufijo}")
        dinero(ws, i, 8, g.monto_postulado)
        ws.cell(row=i, column=9, value=g.get_estado_display())
        if g.fecha_resultado:
            celda = ws.cell(row=i, column=10, value=g.fecha_resultado)
            celda.number_format = 'DD-MM-YYYY'
            ws.cell(row=i, column=11, value=f"{nombre_mes[g.mes_resultado]}-{sufijo}")
        dinero(ws, i, 12, g.monto_adjudicado)
        ws.cell(row=i, column=13, value=g.responsable)
        ws.cell(row=i, column=14, value=g.observaciones)
        # Columna informativa: el importador ignora las que no conoce, así que
        # el archivo sigue siendo reimportable.
        if g.editada_en_sistema:
            ws.cell(row=i, column=15,
                    value=", ".join(g.campos_editados)).font = Font(italic=True)

    anchos(ws, [18, 14, 24, 46, 28, 15, 13, 16, 16, 16, 14, 16, 26, 34, 22])

    # Las mismas listas desplegables de la planilla, para que el archivo se
    # pueda seguir completando a mano sin inventar estados.
    ultima = max(len(gestiones) + 4, 300)
    lista_ambitos = DataValidation(
        type="list",
        formula1='"' + ",".join(a.label for a in Ambito) + '"',
        allow_blank=True)
    lista_estados = DataValidation(
        type="list",
        formula1='"' + ",".join(e.label for e in EstadoGestion) + '"',
        allow_blank=True)
    ws.add_data_validation(lista_ambitos)
    ws.add_data_validation(lista_estados)
    lista_ambitos.add(f"B4:B{ultima}")
    lista_estados.add(f"I4:I{ultima}")

    # ---------- HOJA 2: PROYECCIÓN ----------
    ws = wb.create_sheet("Proyección financiera")
    meses = [f"{nombre}-{sufijo}" for _, nombre in tablero.MESES]
    encabezar(ws, f"PROYECCIÓN FINANCIERA {anio}", ["Ámbito"] + meses + ["Total anual"])

    filas, total = tablero.proyeccion_por_mes(anio)
    for i, fila in enumerate(filas, start=4):
        ws.cell(row=i, column=1, value=fila["etiqueta"]).font = negrita
        for j, monto in enumerate(fila["meses"], start=2):
            dinero(ws, i, j, monto)
        dinero(ws, i, 14, fila["total"]).font = negrita

    ultima_fila = 4 + len(filas)
    ws.cell(row=ultima_fila, column=1, value="TOTAL INGRESOS PROYECTADOS").font = negrita
    for j, monto in enumerate(total["meses"], start=2):
        dinero(ws, ultima_fila, j, monto).font = negrita
    dinero(ws, ultima_fila, 14, total["total"]).font = negrita
    for col in range(1, 15):
        ws.cell(row=ultima_fila, column=col).fill = relleno_gris

    anchos(ws, [30] + [15] * 12 + [17])

    # ---------- HOJA 3: AVANCE MENSUAL ----------
    ws = wb.create_sheet(f"Avance mensual {anio}")
    encabezar(ws, f"AVANCE MENSUAL DE GESTIÓN {anio}",
              ["Indicador"] + meses + ["Total / Acumulado"])

    i = 4
    for bloque in tablero.avance_mensual(anio, gestiones):
        celda = ws.cell(row=i, column=1, value=bloque["etiqueta"].upper())
        celda.font = negrita
        for col in range(1, 15):
            ws.cell(row=i, column=col).fill = relleno_gris
        i += 1

        for fila in bloque["filas"]:
            ws.cell(row=i, column=1, value=fila["rotulo"])
            for j, valor in enumerate(fila["meses"], start=2):
                if fila["es_tasa"]:
                    celda = ws.cell(row=i, column=j, value=_f(valor))
                    celda.number_format = FMT_PCT
                elif fila["es_monto"]:
                    dinero(ws, i, j, valor)
                else:
                    ws.cell(row=i, column=j, value=int(valor)).alignment = derecha

            if fila["es_tasa"]:
                celda = ws.cell(row=i, column=14, value=_f(fila["total"]))
                celda.number_format = FMT_PCT
            elif fila["es_monto"]:
                dinero(ws, i, 14, fila["total"])
            else:
                ws.cell(row=i, column=14, value=int(fila["total"])).alignment = derecha
            ws.cell(row=i, column=14).font = negrita
            i += 1
        i += 1     # una línea en blanco entre ámbitos, como en la planilla

    anchos(ws, [34] + [13] * 12 + [19])

    # ---------- HOJA 4: TABLERO DE CONTROL ----------
    ws = wb.create_sheet("Tablero de control")
    encabezar(ws, f"TABLERO DE CONTROL — RESULTADOS {anio}", [
        "Ámbito", "Meta anual de gestiones", "Gestiones realizadas",
        "Nivel de avance", "Resultados exitosos", "Tasa de éxito",
        "Monto proyectado", "Monto efectivo", "Cumplimiento financiero",
    ])

    filas, total = tablero.tablero_control(anio, gestiones)

    def escribir_control(i, fila, resaltar=False):
        ws.cell(row=i, column=1, value=fila["etiqueta"]).font = negrita
        ws.cell(row=i, column=2, value=int(fila["meta"])).alignment = derecha
        ws.cell(row=i, column=3, value=int(fila["gestiones"])).alignment = derecha
        for col, clave in ((4, "avance"), (6, "tasa_exito"), (9, "cumplimiento")):
            celda = ws.cell(row=i, column=col, value=_f(fila[clave]))
            celda.number_format = FMT_PCT
            celda.alignment = derecha
        ws.cell(row=i, column=5, value=int(fila["exitosos"])).alignment = derecha
        dinero(ws, i, 7, fila["proyectado"])
        dinero(ws, i, 8, fila["efectivo"])
        if resaltar:
            for col in range(1, 10):
                ws.cell(row=i, column=col).fill = relleno_gris
                ws.cell(row=i, column=col).font = negrita

    for i, fila in enumerate(filas, start=4):
        escribir_control(i, fila)
    escribir_control(4 + len(filas), total, resaltar=True)

    pie = 6 + len(filas)
    ws.cell(row=pie, column=1, value="CRITERIOS DE INTERPRETACIÓN").font = negrita
    for j, (titulo, texto) in enumerate([
        ("Nivel de avance", "Gestiones realizadas / Meta anual de gestiones."),
        ("Tasa de éxito", "Resultados exitosos / Gestiones realizadas."),
        ("Cumplimiento financiero", "Monto efectivo / Monto proyectado."),
        ("Semáforo sugerido", "Verde ≥ 90%; amarillo 70%–89%; rojo < 70%."),
    ], start=1):
        ws.cell(row=pie + j, column=1, value=titulo)
        ws.cell(row=pie + j, column=2, value=texto)

    anchos(ws, [26, 22, 20, 16, 18, 14, 20, 18, 22])

    # ---------- HOJA 5: INSTRUCCIONES ----------
    ws = wb.create_sheet("Instrucciones")
    ws.cell(row=1, column=1, value="INSTRUCCIONES DE USO").font = Font(bold=True, size=13)
    ws.cell(row=3, column=1, value="Paso").font = blanco
    ws.cell(row=3, column=1).fill = relleno
    ws.cell(row=3, column=2, value="Acción").font = blanco
    ws.cell(row=3, column=2).fill = relleno
    pasos = [
        "Ingrese cada iniciativa en la hoja «Registro iniciativas». Una fila por iniciativa.",
        "Complete las fechas reales de ingreso y de resultado: definen en qué mes se cuenta.",
        "Seleccione el ámbito y el estado con las listas desplegables.",
        "Las demás hojas son un reflejo: el sistema las recalcula, no hace falta tocarlas.",
        "Suba este mismo archivo en «Cargar Excel» para actualizar el tablero.",
        "Antes de guardar verá qué cambia; si algo se editó en el sistema, se le preguntará.",
    ]
    for j, texto in enumerate(pasos, start=4):
        ws.cell(row=j, column=1, value=j - 3)
        ws.cell(row=j, column=2, value=texto)
    anchos(ws, [8, 110])

    flujo = io.BytesIO()
    wb.save(flujo)
    flujo.seek(0)

    respuesta = HttpResponse(
        flujo.read(),
        content_type=("application/vnd.openxmlformats-officedocument"
                      ".spreadsheetml.sheet"),
    )
    respuesta["Content-Disposition"] = (
        f'attachment; filename="Tablero_Resultados_OCT_{anio}.xlsx"')
    return respuesta


# =========================
# INFORME IMPRIMIBLE (PDF por el navegador)
# =========================

@login_required
def informe_tablero(request):
    """Informe listo para imprimir o guardar como PDF desde el navegador.

    Acepta ``?ambito=`` para sacar el informe de una sola línea de trabajo, que
    es lo que se pide cuando hay que rendirle a una facultad y no al total.
    """
    anio, _ = _anio_pedido(request)

    filtro = request.GET.get("ambito") or ""
    if filtro not in Ambito.values:
        filtro = ""

    gestiones = list(Gestion.objects.filter(anio=anio).order_by(
        "ambito", "-fecha_ingreso", "nombre"))
    # El informe de un ámbito muestra solo sus números: el tablero se recalcula
    # con el subconjunto, no se recorta la tabla completa.
    del_informe = [g for g in gestiones if not filtro or g.ambito == filtro]

    filas, total = tablero.tablero_control(anio, del_informe)
    bloques = tablero.avance_mensual(anio, del_informe)
    proyeccion, proyeccion_total = tablero.proyeccion_por_mes(anio)

    if filtro:
        filas = [f for f in filas if f["ambito"] == filtro]
        bloques = [b for b in bloques if b["ambito"] == filtro]
        proyeccion = [p for p in proyeccion if p["ambito"] == filtro]
        # Con un solo ámbito, el "total" del tablero es esa misma fila.
        total = filas[0] if filas else total
        proyeccion_total = proyeccion[0] if proyeccion else proyeccion_total

    # Las gestiones se listan agrupadas, que es como se leen en papel.
    grupos = []
    for ambito, etiqueta in Ambito.choices:
        del_ambito = [g for g in del_informe if g.ambito == ambito]
        if del_ambito:
            grupos.append({
                "etiqueta": etiqueta,
                "gestiones": del_ambito,
                "postulado": sum((g.monto_postulado for g in del_ambito), Decimal("0")),
                "adjudicado": sum((g.monto_adjudicado for g in del_ambito), Decimal("0")),
            })

    # --- Gráficos, dibujados como SVG para que salgan también en el papel ---
    conteo_estados = [
        (estado.label, sum(1 for g in del_informe if g.estado == estado.value))
        for estado in EstadoGestion
    ]
    efectivo_por_mes = [
        sum((f["meses"][i] for b in bloques for f in b["filas"]
             if f["clave"] == "monto_adjudicado"), Decimal("0"))
        for i, _ in enumerate(tablero.MESES)
    ]

    return render(request, "oct/tablero/informe.html", {
        "grafico_avance": graficos.barras_de_avance(filas),
        "grafico_estados": graficos.barras_por_estado(conteo_estados),
        "grafico_meses": graficos.columnas_mensuales(
            tablero.MESES, proyeccion_total["meses"], efectivo_por_mes),
        "anio": anio,
        "filtro": filtro,
        "etiqueta_filtro": Ambito(filtro).label if filtro else "",
        "ambitos": Ambito.choices,
        "filas": filas,
        "total": total,
        "bloques": bloques,
        "meses": tablero.MESES,
        "proyeccion": proyeccion,
        "proyeccion_total": proyeccion_total,
        "grupos": grupos,
        "n_gestiones": len(del_informe),
        "generado": timezone.now(),
        "auto_print": request.GET.get("print") == "1",
    })
