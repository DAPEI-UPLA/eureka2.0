"""Carga de «Planilla_Resultados_OCT_2026.xlsx» sin duplicar ni pisar lo editado.

La planilla no trae un identificador estable para cada fila, así que la
identidad se deduce de lo que el archivo ya contiene:

* el **código** (``COD``), cuando dice algo — es lo que usa la OCT para hablar
  de una gestión ("C-002", "ID 628-8-LE26");
* si no hay código (varias filas dicen ``N/A``), el **nombre dentro del
  ámbito**, ignorando tildes, mayúsculas y espacios de más, y tolerando que le
  corrijan la redacción.

Con eso, subir el archivo actualizado dos veces actualiza lo que cambió y deja
lo demás intacto: no hay que agregarle ninguna columna al Excel.

**Solo se leen las tres hojas que se escriben a mano** —el registro, la
proyección financiera y la meta anual del tablero—. El avance mensual y el
resto del tablero de control son fórmulas y el sistema los recalcula
(``oct/tablero.py``), así que da lo mismo si el archivo llega sin recalcular.

**Lo editado en el sistema no se pisa solo.** Cada gestión recuerda qué campos
se corrigieron en pantalla; si el Excel trae otro valor para uno de ellos, la
carga no decide: lo reporta como conflicto, lo conserva por defecto y deja que
la persona elija cuáles reemplazar. Lo mismo con una gestión editada acá que
ya no viene en el archivo: no se borra sin preguntar.

El importador corre siempre dentro de una transacción, de modo que la misma
pasada sirve para previsualizar (``aplicar=False``, se revierte al final) o
para guardar.
"""

import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher

from django.db import transaction

from .models import (
    Ambito,
    EstadoGestion,
    Gestion,
    MetaAmbito,
    Origen,
    ProyeccionMensual,
)

HOJA_REGISTRO = "Registro iniciativas"
HOJA_PROYECCION = "Proyección financiera"
HOJA_TABLERO = "Tablero de control"

# Encabezados de la hoja de registro y el campo del modelo al que van.
COLUMNAS = {
    "cod": "codigo",
    "ambito": "ambito",
    "tipo de iniciativa": "tipo",
    "nombre de la iniciativa": "nombre",
    "institucion": "institucion",
    "fecha de ingreso": "fecha_ingreso",
    "monto postulado": "monto_postulado",
    "estado": "estado",
    "fecha de resultado": "fecha_resultado",
    "monto adjudicado": "monto_adjudicado",
    "responsable": "responsable",
    "observaciones": "observaciones",
}

# Campos que se comparan al reimportar, en el orden en que se muestran.
CAMPOS = [
    "codigo", "ambito", "tipo", "nombre", "institucion", "fecha_ingreso",
    "monto_postulado", "estado", "fecha_resultado", "monto_adjudicado",
    "responsable", "observaciones",
]

ETIQUETAS = {
    "codigo": "Código",
    "ambito": "Ámbito",
    "tipo": "Tipo de iniciativa",
    "nombre": "Nombre",
    "institucion": "Institución",
    "fecha_ingreso": "Fecha de ingreso",
    "monto_postulado": "Monto postulado",
    "estado": "Estado",
    "fecha_resultado": "Fecha de resultado",
    "monto_adjudicado": "Monto adjudicado",
    "responsable": "Responsable",
    "observaciones": "Observaciones",
}

# Valores que en la planilla significan "sin dato" y no sirven de código.
VACIOS = {"", "-", "--", "n/a", "na", "s/i", "sin dato", "por definir", "no aplica"}

# Similitud mínima para reconocer una fila como la misma gestión con el nombre
# corregido, en vez de una gestión nueva.
UMBRAL_RENOMBRE = 0.80

MESES_TEXTO = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dic": 12,
}


# =========================
# NORMALIZACIÓN
# =========================

def texto(valor):
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def clave(valor):
    """Clave de comparación: sin tildes, sin puntuación, sin espacios de más."""
    s = unicodedata.normalize("NFKD", texto(valor).casefold())
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^0-9a-z ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def a_fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    crudo = texto(valor)
    if not crudo:
        return None
    for formato in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(crudo, formato).date()
        except ValueError:
            continue
    return None  # una fecha mal tipeada se descarta, no revienta la carga


def a_monto(valor):
    if valor in (None, ""):
        return Decimal("0")
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor)).quantize(Decimal("1"))
    crudo = texto(valor).replace("$", "").replace(".", "").replace(",", ".")
    try:
        return Decimal(crudo or "0").quantize(Decimal("1"))
    except InvalidOperation:
        return Decimal("0")


def a_entero(valor):
    try:
        return max(0, int(float(valor)))
    except (TypeError, ValueError):
        return 0


def mes_de(valor):
    """Número de mes a partir de un encabezado como ``ene-26`` o una fecha."""
    if isinstance(valor, (datetime, date)):
        return valor.month
    crudo = clave(valor)
    if not crudo:
        return None
    prefijo = crudo.split(" ")[0][:4]
    for largo in (4, 3):
        mes = MESES_TEXTO.get(prefijo[:largo])
        if mes:
            return mes
    return None


AMBITOS = {clave(a.label): a.value for a in Ambito}
ESTADOS = {clave(e.label): e.value for e in EstadoGestion}


# =========================
# RESULTADO DE UNA CARGA
# =========================

CREAR = "creada"
ACTUALIZAR = "actualizada"
RENOMBRAR = "renombrada"
ELIMINAR = "eliminada"
CONSERVAR = "conservada"
IGUAL = "sin cambios"


@dataclass
class Cambio:
    entidad: str
    accion: str
    nombre: str
    detalle: str = ""


@dataclass
class Diferencia:
    campo: str
    etiqueta: str
    actual: str
    propuesto: str


@dataclass
class Conflicto:
    """Un choque entre el archivo y algo que se editó en el sistema.

    ``clave`` es lo que viaja en el formulario para decidir qué hacer:
    marcada = reemplazar con el Excel; sin marcar = conservar lo editado.
    """

    clave: str
    tipo: str            # "campos" o "eliminar"
    nombre: str
    ambito: str
    codigo: str = ""
    diferencias: list = field(default_factory=list)
    resuelto_con_excel: bool = False
    editado_por: str = ""
    fecha_edicion: object = None

    @property
    def es_eliminacion(self):
        return self.tipo == "eliminar"


@dataclass
class Resultado:
    aplicado: bool = False
    anio: int = 0
    cambios: list = field(default_factory=list)
    conflictos: list = field(default_factory=list)
    avisos: list = field(default_factory=list)
    descartadas: list = field(default_factory=list)

    def registrar(self, entidad, accion, nombre, detalle=""):
        self.cambios.append(Cambio(entidad, accion, nombre, detalle))

    @property
    def relevantes(self):
        return [c for c in self.cambios if c.accion != IGUAL]

    @property
    def hay_cambios(self):
        return bool(self.relevantes)

    @property
    def hay_conflictos(self):
        return bool(self.conflictos)

    def resumen_por_entidad(self):
        """Una fila por tipo de registro, con las llaves sin espacios para que
        la plantilla pueda leerlas directo."""
        llaves = {
            CREAR: "nuevas", ACTUALIZAR: "actualizadas", RENOMBRAR: "renombradas",
            ELIMINAR: "eliminadas", CONSERVAR: "conservadas", IGUAL: "iguales",
        }
        entidades = {}
        for c in self.cambios:
            fila = entidades.setdefault(c.entidad, {
                "entidad": c.entidad, "nuevas": 0, "actualizadas": 0,
                "renombradas": 0, "eliminadas": 0, "conservadas": 0, "iguales": 0,
            })
            fila[llaves[c.accion]] += 1
        orden = ["Gestión", "Proyección financiera", "Meta anual"]
        return sorted(
            entidades.values(),
            key=lambda f: orden.index(f["entidad"]) if f["entidad"] in orden else 99,
        )


class ErrorImportacion(Exception):
    pass


class _Rollback(Exception):
    """Revierte la transacción cuando solo se está previsualizando."""


# =========================
# IMPORTADOR
# =========================

class ImportadorPlanilla:
    """Lee la planilla de resultados y la refleja en la base de datos."""

    def __init__(self, ruta, anio=None, decisiones=None, usuario=None, podar=True):
        self.ruta = ruta
        self.anio_forzado = anio
        # {clave del conflicto: True} — True significa "usar el Excel".
        self.decisiones = decisiones or {}
        self.usuario = usuario
        self.podar_faltantes = podar
        self.resultado = Resultado()

    # -- lectura --

    def _leer(self):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise ErrorImportacion("Falta openpyxl (pip install openpyxl).") from exc

        try:
            wb = load_workbook(self.ruta, data_only=True, read_only=True)
        except FileNotFoundError as exc:
            raise ErrorImportacion(f"No se encontró el archivo: {self.ruta}") from exc
        except Exception as exc:
            raise ErrorImportacion(
                "No se pudo leer el archivo. ¿Es un .xlsx válido?") from exc

        hojas = {clave(h): h for h in wb.sheetnames}
        if clave(HOJA_REGISTRO) not in hojas:
            nombres = ", ".join(wb.sheetnames)
            wb.close()
            raise ErrorImportacion(
                f"El archivo no tiene la hoja «{HOJA_REGISTRO}». "
                f"Hojas encontradas: {nombres}.")

        def filas_de(nombre):
            real = hojas.get(clave(nombre))
            return list(wb[real].iter_rows(values_only=True)) if real else []

        datos = {
            "registro": filas_de(HOJA_REGISTRO),
            "proyeccion": filas_de(HOJA_PROYECCION),
            "tablero": filas_de(HOJA_TABLERO),
        }
        wb.close()

        if not datos["registro"]:
            raise ErrorImportacion(f"La hoja «{HOJA_REGISTRO}» está vacía.")
        return datos

    def _detectar_anio(self, datos):
        """El año sale del título de la hoja ("REGISTRO DE INICIATIVAS 2026").

        Si el título no lo dice, se toma el año más repetido entre las fechas
        de ingreso; y si no hay ninguna, el año en curso.
        """
        if self.anio_forzado:
            return int(self.anio_forzado)

        for filas in (datos["registro"], datos["tablero"], datos["proyeccion"]):
            for fila in filas[:3]:
                for celda in fila or ():
                    encontrados = re.findall(r"\b(20\d{2})\b", texto(celda))
                    if encontrados:
                        return int(encontrados[0])

        anios = [
            f.year for f in
            (a_fecha(v) for v in self._columna_fechas(datos["registro"]))
            if f
        ]
        if anios:
            return max(set(anios), key=anios.count)
        return date.today().year

    def _columna_fechas(self, filas):
        cabecera, cuerpo = self._partir(filas)
        i = cabecera.get("fecha_ingreso")
        if i is None:
            return []
        return [f[i] for _, f in cuerpo if i < len(f)]

    def _partir(self, filas):
        """Ubica la fila de encabezados y devuelve ``({campo: columna}, cuerpo)``.

        El cuerpo viene como ``(número de fila en el Excel, fila)`` para que los
        avisos apunten a la línea que la persona ve en pantalla. Se busca el
        encabezado en vez de fijar la fila 3 porque basta que alguien agregue
        una línea al título para que las coordenadas fijas dejen de servir.
        """
        for i, fila in enumerate(filas):
            indice = {}
            for j, celda in enumerate(fila or ()):
                campo = COLUMNAS.get(clave(celda))
                if campo and campo not in indice:
                    indice[campo] = j
            if {"ambito", "nombre", "estado"} <= set(indice):
                cuerpo = [
                    (n, f) for n, f in enumerate(filas[i + 1:], start=i + 2) if f
                ]
                return indice, cuerpo
        raise ErrorImportacion(
            f"No se encontró la fila de encabezados en «{HOJA_REGISTRO}». "
            f"Deben estar al menos las columnas Ámbito, Nombre de la "
            f"iniciativa y Estado.")

    # -- ejecución --

    def ejecutar(self, aplicar=False):
        datos = self._leer()
        self.anio = self._detectar_anio(datos)
        self.resultado.anio = self.anio

        try:
            with transaction.atomic():
                self._procesar_registro(datos["registro"])
                self._procesar_proyeccion(datos["proyeccion"])
                self._procesar_metas(datos["tablero"])
                if not aplicar:
                    raise _Rollback()
        except _Rollback:
            pass
        else:
            self.resultado.aplicado = True

        return self.resultado

    # -- hoja de registro --

    def _fila_a_valores(self, fila, indice, numero):
        """Traduce una fila del Excel a los campos del modelo.

        Devuelve ``None`` si la fila no es una gestión (está vacía, o es la
        fila de totales que alguien haya dejado abajo).
        """
        def celda(campo):
            j = indice.get(campo)
            return fila[j] if j is not None and j < len(fila) else None

        nombre = texto(celda("nombre"))
        ambito_crudo = texto(celda("ambito"))
        if not nombre and not ambito_crudo:
            return None

        ambito = AMBITOS.get(clave(ambito_crudo))
        if not ambito:
            self.resultado.descartadas.append(
                f"Fila {numero}: ámbito «{ambito_crudo or '(vacío)'}» "
                f"desconocido; se omite «{nombre[:60]}».")
            return None
        if not nombre:
            self.resultado.descartadas.append(
                f"Fila {numero}: sin nombre de iniciativa; se omite.")
            return None

        estado_crudo = texto(celda("estado"))
        estado = ESTADOS.get(clave(estado_crudo))
        if not estado:
            estado = EstadoGestion.EN_IDENTIFICACION
            if estado_crudo:
                self.resultado.avisos.append(
                    f"Fila {numero}: estado «{estado_crudo}» no está en la "
                    f"lista; queda como «En identificación».")

        codigo = texto(celda("codigo"))
        if clave(codigo) in VACIOS:
            codigo = ""

        return {
            "codigo": codigo,
            "ambito": ambito,
            "tipo": texto(celda("tipo")),
            "nombre": nombre,
            "institucion": texto(celda("institucion")),
            "fecha_ingreso": a_fecha(celda("fecha_ingreso")),
            "monto_postulado": a_monto(celda("monto_postulado")),
            "estado": estado,
            "fecha_resultado": a_fecha(celda("fecha_resultado")),
            "monto_adjudicado": a_monto(celda("monto_adjudicado")),
            "responsable": texto(celda("responsable")),
            "observaciones": texto(celda("observaciones")),
        }

    def _procesar_registro(self, filas):
        indice, cuerpo = self._partir(filas)

        entrantes = []
        vistas = set()
        for n, fila in cuerpo:
            valores = self._fila_a_valores(fila, indice, n)
            if valores is None:
                continue

            # Con código, la identidad es el código; sin él, el nombre dentro
            # del ámbito. Si la misma clave aparece dos veces en el archivo,
            # se avisa: son dos filas que el sistema no puede distinguir.
            if valores["codigo"]:
                identidad = ("cod", clave(valores["codigo"]))
            else:
                identidad = ("nom", valores["ambito"], clave(valores["nombre"]))
            if identidad in vistas:
                self.resultado.avisos.append(
                    f"Fila {n}: «{valores['nombre'][:50]}» repite el "
                    f"{'código' if valores['codigo'] else 'nombre'} de otra "
                    f"fila; se trata como la misma gestión.")
            vistas.add(identidad)
            entrantes.append((identidad, valores))

        existentes = list(Gestion.objects.filter(anio=self.anio))
        por_codigo = {
            ("cod", clave(g.codigo)): g for g in existentes if g.codigo
        }
        por_nombre = {
            ("nom", g.ambito, clave(g.nombre)): g for g in existentes
        }

        emparejadas = {}
        usadas = set()
        sin_pareja = []

        for identidad, valores in entrantes:
            actual = por_codigo.get(identidad) or por_nombre.get(identidad)
            if actual is not None and actual.pk not in usadas:
                emparejadas[actual.pk] = valores
                usadas.add(actual.pk)
            else:
                sin_pareja.append(valores)

        # Renombres: una fila sin pareja puede ser una gestión que ya está,
        # con la redacción corregida. Sin esto, la poda la borraba y la volvía
        # a crear, perdiendo lo que se hubiera editado en pantalla.
        libres = [g for g in existentes if g.pk not in usadas]
        renombradas = set()
        nuevas = []
        for valores in sin_pareja:
            candidata = self._mejor_parecida(valores, libres)
            if candidata is None:
                nuevas.append(valores)
                continue
            emparejadas[candidata.pk] = valores
            renombradas.add(candidata.pk)
            usadas.add(candidata.pk)
            libres = [g for g in libres if g.pk != candidata.pk]

        for g in existentes:
            if g.pk in emparejadas:
                self._actualizar(g, emparejadas[g.pk], g.pk in renombradas)

        for valores in nuevas:
            gestion = Gestion.objects.create(
                anio=self.anio, origen=Origen.IMPORTADO, **valores)
            self.resultado.registrar(
                "Gestión", CREAR, gestion.nombre,
                f"{gestion.get_ambito_display()} · {gestion.codigo or 'sin código'}")

        if self.podar_faltantes:
            self._podar([g for g in existentes if g.pk not in usadas])

    def _mejor_parecida(self, valores, candidatas):
        """La gestión existente que más se parece, dentro del mismo ámbito.

        Solo se compara contra filas **sin código propio o con el mismo
        código**: si la planilla cambió el código, es otra gestión.
        """
        objetivo = clave(valores["nombre"])
        mejor, puntaje_mejor = None, UMBRAL_RENOMBRE
        for g in candidatas:
            if g.ambito != valores["ambito"]:
                continue
            if g.codigo and valores["codigo"] and clave(g.codigo) != clave(valores["codigo"]):
                continue
            puntaje = SequenceMatcher(None, objetivo, clave(g.nombre)).ratio()
            if puntaje > puntaje_mejor:
                mejor, puntaje_mejor = g, puntaje
        return mejor

    def _actualizar(self, gestion, valores, renombrada):
        protegidos = set(gestion.campos_editados or [])
        distintos = [
            campo for campo in CAMPOS
            if getattr(gestion, campo) != valores[campo]
        ]

        if gestion.origen != Origen.IMPORTADO:
            # Una gestión creada a mano que ahora sí viene en el archivo pasa a
            # estar gobernada por él.
            gestion.origen = Origen.IMPORTADO
            gestion.save(update_fields=["origen"])

        if not distintos:
            self.resultado.registrar("Gestión", IGUAL, gestion.nombre)
            return

        chocados = [c for c in distintos if c in protegidos]
        libres = [c for c in distintos if c not in protegidos]

        usar_excel = True
        if chocados:
            llave = f"g{gestion.pk}"
            usar_excel = bool(self.decisiones.get(llave))
            self.resultado.conflictos.append(Conflicto(
                clave=llave,
                tipo="campos",
                nombre=gestion.nombre,
                ambito=gestion.get_ambito_display(),
                codigo=gestion.codigo,
                diferencias=[
                    Diferencia(
                        campo=c,
                        etiqueta=ETIQUETAS[c],
                        actual=self._mostrar(gestion, c, getattr(gestion, c)),
                        propuesto=self._mostrar(gestion, c, valores[c]),
                    )
                    for c in chocados
                ],
                resuelto_con_excel=usar_excel,
                editado_por=(
                    gestion.editado_por.get_full_name() or gestion.editado_por.username
                    if gestion.editado_por else ""),
                fecha_edicion=gestion.fecha_edicion,
            ))

        aplicar = list(libres)
        if chocados and usar_excel:
            aplicar += chocados

        for campo in aplicar:
            setattr(gestion, campo, valores[campo])

        campos_guardados = list(aplicar)
        if chocados and usar_excel:
            # Ya no hay nada que proteger en esos campos: el Excel manda de
            # nuevo hasta la próxima edición en pantalla.
            gestion.campos_editados = sorted(protegidos - set(chocados))
            campos_guardados.append("campos_editados")

        if campos_guardados:
            gestion.save(update_fields=campos_guardados + ["actualizado"])

        if chocados and not usar_excel:
            detalle = ", ".join(ETIQUETAS[c] for c in chocados)
            self.resultado.registrar(
                "Gestión", CONSERVAR, gestion.nombre,
                f"se mantuvo lo editado en: {detalle}")
        elif renombrada:
            self.resultado.registrar(
                "Gestión", RENOMBRAR, gestion.nombre,
                f"{len(aplicar)} campo{'s' if len(aplicar) != 1 else ''} actualizado"
                f"{'s' if len(aplicar) != 1 else ''}")
        else:
            self.resultado.registrar(
                "Gestión", ACTUALIZAR, gestion.nombre,
                ", ".join(ETIQUETAS[c] for c in aplicar))

    def _mostrar(self, gestion, campo, valor):
        """Cómo se ve un valor en el aviso de conflicto."""
        if valor in (None, ""):
            return "—"
        if campo == "ambito":
            return Ambito(valor).label
        if campo == "estado":
            return EstadoGestion(valor).label
        if isinstance(valor, date):
            return valor.strftime("%d-%m-%Y")
        if isinstance(valor, Decimal):
            return f"${valor:,.0f}".replace(",", ".")
        return texto(valor)

    def _podar(self, sobrantes):
        """Borra lo importado que ya no viene en el archivo.

        Lo cargado a mano nunca se toca: por definición no está en el Excel.
        Lo que se editó en pantalla tampoco se borra solo — se pregunta.
        """
        for gestion in sobrantes:
            if gestion.origen != Origen.IMPORTADO:
                continue

            if gestion.editada_en_sistema:
                llave = f"x{gestion.pk}"
                borrar = bool(self.decisiones.get(llave))
                self.resultado.conflictos.append(Conflicto(
                    clave=llave,
                    tipo="eliminar",
                    nombre=gestion.nombre,
                    ambito=gestion.get_ambito_display(),
                    codigo=gestion.codigo,
                    diferencias=[
                        Diferencia(
                            campo=c,
                            etiqueta=ETIQUETAS.get(c, c),
                            actual=self._mostrar(gestion, c, getattr(gestion, c, "")),
                            propuesto="(ya no viene en el archivo)",
                        )
                        for c in (gestion.campos_editados or [])
                    ],
                    resuelto_con_excel=borrar,
                    editado_por=(
                        gestion.editado_por.get_full_name() or gestion.editado_por.username
                        if gestion.editado_por else ""),
                    fecha_edicion=gestion.fecha_edicion,
                ))
                if not borrar:
                    self.resultado.registrar(
                        "Gestión", CONSERVAR, gestion.nombre,
                        "no viene en el archivo, pero se editó acá")
                    continue

            nombre = gestion.nombre
            gestion.delete()
            self.resultado.registrar(
                "Gestión", ELIMINAR, nombre, "ya no viene en el archivo")

    # -- hoja de proyección financiera --

    def _procesar_proyeccion(self, filas):
        if not filas:
            return

        columnas = None
        for fila in filas:
            posibles = {
                j: mes_de(celda)
                for j, celda in enumerate(fila or ())
                if mes_de(celda)
            }
            if len(set(posibles.values())) >= 12:
                columnas = posibles
                break

        if not columnas:
            self.resultado.avisos.append(
                f"La hoja «{HOJA_PROYECCION}» no tiene una fila de meses "
                f"reconocible; la proyección queda como está.")
            return

        for fila in filas:
            ambito = AMBITOS.get(clave(fila[0] if fila else ""))
            if not ambito:
                continue
            for j, mes in columnas.items():
                monto = a_monto(fila[j] if j < len(fila) else None)
                registro, creado = ProyeccionMensual.objects.get_or_create(
                    anio=self.anio, ambito=ambito, mes=mes,
                    defaults={"monto": monto})
                if creado:
                    if monto:
                        self.resultado.registrar(
                            "Proyección financiera", CREAR,
                            f"{Ambito(ambito).label} · mes {mes:02d}",
                            f"${monto:,.0f}".replace(",", "."))
                elif registro.monto != monto:
                    registro.monto = monto
                    registro.save(update_fields=["monto"])
                    self.resultado.registrar(
                        "Proyección financiera", ACTUALIZAR,
                        f"{Ambito(ambito).label} · mes {mes:02d}",
                        f"${monto:,.0f}".replace(",", "."))
                else:
                    self.resultado.registrar(
                        "Proyección financiera", IGUAL,
                        f"{Ambito(ambito).label} · mes {mes:02d}")

    # -- metas del tablero de control --

    def _procesar_metas(self, filas):
        if not filas:
            return

        columna = None
        for fila in filas:
            for j, celda in enumerate(fila or ()):
                if clave(celda).startswith("meta anual"):
                    columna = j
                    break
            if columna is not None:
                break

        if columna is None:
            self.resultado.avisos.append(
                f"La hoja «{HOJA_TABLERO}» no tiene la columna «Meta anual de "
                f"gestiones»; las metas quedan como están.")
            return

        for fila in filas:
            ambito = AMBITOS.get(clave(fila[0] if fila else ""))
            if not ambito or columna >= len(fila):
                continue
            meta = a_entero(fila[columna])
            registro, creado = MetaAmbito.objects.get_or_create(
                anio=self.anio, ambito=ambito,
                defaults={"meta_gestiones": meta})
            etiqueta = f"{Ambito(ambito).label} {self.anio}"
            if creado:
                self.resultado.registrar(
                    "Meta anual", CREAR, etiqueta, f"{meta} gestiones")
            elif registro.meta_gestiones != meta:
                registro.meta_gestiones = meta
                registro.save(update_fields=["meta_gestiones"])
                self.resultado.registrar(
                    "Meta anual", ACTUALIZAR, etiqueta, f"{meta} gestiones")
            else:
                self.resultado.registrar("Meta anual", IGUAL, etiqueta)
