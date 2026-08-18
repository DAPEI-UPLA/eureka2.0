"""Exportación de proyectos a Excel y a informe imprimible (PDF vía navegador).

Tres salidas, todas reutilizando los @property ya definidos en los modelos
(no hay cálculos nuevos ni migraciones):

  - exportar_proyecto_excel: libro .xlsx de un proyecto con 3 hojas
      (Resumen · Plan de gasto/POA · Egresos).
  - exportar_cartera_excel: libro .xlsx con el resumen de todos los proyectos
      visibles (respeta el permiso jefe vs responsable, igual que la lista).
  - informe_proyecto: página HTML optimizada para imprimir / "Guardar como PDF"
      desde el navegador (sin dependencias de PDF en el servidor). Sus gráficos
      son SVG dibujados acá mismo (`..graficos_informe`), no Chart.js: ver el
      porqué en `core.svg`.

openpyxl ya está en requirements; si faltara, se devuelve un 500 explicativo
(mismo patrón que planificacion.views).
"""

import io
from collections import defaultdict
from datetime import date
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.utils.text import slugify

from .. import graficos_informe as gi
from ..models import Egreso, PlanDeGasto, Proyecto
from .graficos import _cumplimiento_objetivo
from .permisos import es_jefe

# Formato de moneda chilena (sin decimales) para las celdas numéricas.
FMT_CLP = '#,##0'
HEADER_FILL = '2563EB'      # azul (coincide con el resto de exports)
SUBHEADER_FILL = 'E2E8F0'   # gris claro para filas de sección


def _f(valor):
    """Decimal/None -> float para escribir en la celda numérica."""
    if valor is None:
        return 0.0
    if isinstance(valor, Decimal):
        return float(valor)
    return float(valor)


def _egreso_montos(e):
    """(total, pagado, comprometido) de un egreso.

    Alias de `Egreso.montos`, que es donde vive la semántica desde que las
    pantallas y el Excel tuvieron que coincidir.
    """
    return e.montos


def _egreso_detalle(e):
    """Descripción legible del egreso según su tipo."""
    if e.tipo == Egreso.TIPO_HONORARIO:
        base = e.nombre_completo_persona or 'Honorario'
        return f'{base} — {e.profesion}' if e.profesion else base
    if e.tipo == Egreso.TIPO_COMPRA:
        ge = str(e.gasto_elegible) if e.gasto_elegible_id else e.get_subtipo_compra_display()
        return ge or 'Compra'
    return e.descripcion or e.get_tipo_display()


# =========================
# EXCEL — POR PROYECTO
# =========================

@login_required
def exportar_proyecto_excel(request, pk):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return HttpResponse(
            "openpyxl no está instalado. Ejecuta: pip install openpyxl",
            status=500,
        )

    proyecto = get_object_or_404(
        Proyecto.objects.prefetch_related('objetivos__resultados__actividades'),
        pk=pk,
    )

    bold_white = Font(bold=True, color='FFFFFF')
    bold = Font(bold=True)
    header_fill = PatternFill('solid', fgColor=HEADER_FILL)
    right = Alignment(horizontal='right')

    def escribir_encabezado(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = bold_white
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def dinero(ws, row, col, valor):
        c = ws.cell(row=row, column=col, value=_f(valor))
        c.number_format = FMT_CLP
        c.alignment = right
        return c

    wb = Workbook()

    # ---------- HOJA 1: RESUMEN ----------
    ws = wb.active
    ws.title = 'Resumen'
    ws['A1'] = proyecto.nombre
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    filas = [
        ('Código', proyecto.codigo or '—'),
        ('Tipo', proyecto.get_tipo_display()),
        ('Estado', proyecto.get_estado_display()),
        ('Prioridad', proyecto.get_prioridad_display()),
        ('Responsable', str(proyecto.responsable) if proyecto.responsable else '—'),
        ('Fecha inicio', proyecto.fecha_inicio.strftime('%d-%m-%Y') if proyecto.fecha_inicio else '—'),
        ('Fecha fin', proyecto.fecha_fin.strftime('%d-%m-%Y') if proyecto.fecha_fin else '—'),
        ('Cumplimiento físico', f'{_f(proyecto.cumplimiento):.1f}%'),
        ('', ''),
        ('PRESUPUESTO', ''),
    ]
    r = 3
    for etq, val in filas:
        ws.cell(row=r, column=1, value=etq).font = bold
        ws.cell(row=r, column=2, value=val)
        r += 1

    montos = [
        ('Presupuesto total', proyecto.presupuesto_total),
        ('  Corriente', proyecto.presupuesto_corriente),
        ('  Capital', proyecto.presupuesto_capital),
        ('Asignado a objetivos', proyecto.presupuesto_asignado),
        ('Disponible sin asignar', proyecto.presupuesto_disponible_real),
        ('', None),
        ('Gastos comprometidos', proyecto.gastos_comprometidos),
        ('Gastos pagados', proyecto.gastos_pagados),
        ('Gastos totales', proyecto.gastos_total),
        ('Disponible para gastar', proyecto.disponible_para_gastar),
        ('% ejecutado', None),
    ]
    for etq, val in montos:
        if not etq:
            r += 1
            continue
        ws.cell(row=r, column=1, value=etq).font = bold if not etq.startswith('  ') else Font()
        if etq == '% ejecutado':
            ws.cell(row=r, column=2, value=f'{_f(proyecto.porcentaje_gastado):.1f}%').alignment = right
        elif val is not None:
            dinero(ws, r, 2, val)
        r += 1

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 22

    # ---------- HOJA 2: PLAN DE GASTO (POA) ----------
    ws2 = wb.create_sheet('Plan de gasto')
    headers = [
        'Año', 'Objetivo', 'Resultado', 'Actividad', 'Transferencia',
        'Tipo de gasto', 'Gasto', 'Gasto elegible', 'Unidad',
        'Planificado', 'Comprometido', 'Ejecutado', 'Saldo',
    ]
    escribir_encabezado(ws2, headers)

    objetivos = list(proyecto.objetivos.filter(eliminado=False).order_by('orden', 'id'))
    obj_index = {o.id: i for i, o in enumerate(objetivos, start=1)}

    planes = (
        PlanDeGasto.objects
        .filter(
            resultado__objetivo__proyecto=proyecto,
            resultado__eliminado=False,
            resultado__objetivo__eliminado=False,
        )
        .select_related(
            'resultado__objetivo',
            'gasto_elegible__gasto__tipo_gasto__transferencia',
            'unidad_responsable',
        )
        .order_by(
            'anio',
            'resultado__objetivo__orden', 'resultado__objetivo_id',
            'resultado__orden', 'resultado_id',
            'actividad__orden', 'actividad_id',
        )
    )

    r = 2
    tot_plan = tot_comp = tot_ejec = tot_saldo = Decimal('0')
    for p in planes:
        act = p.actividad
        res = act.resultado
        obj = res.objetivo
        ge = p.gasto_elegible
        ws2.cell(row=r, column=1, value=p.anio)
        ws2.cell(row=r, column=2, value=f'OE{obj_index.get(obj.id, "?")}')
        ws2.cell(row=r, column=3, value=(res.descripcion or f'R{res.id}')[:60])
        ws2.cell(row=r, column=4, value=act.nombre[:60])
        ws2.cell(row=r, column=5, value=str(ge.gasto.tipo_gasto.transferencia))
        ws2.cell(row=r, column=6, value=str(ge.gasto.tipo_gasto.nombre))
        ws2.cell(row=r, column=7, value=str(ge.gasto.nombre))
        ws2.cell(row=r, column=8, value=str(ge.nombre))
        ws2.cell(row=r, column=9, value=str(p.unidad_responsable) if p.unidad_responsable_id else '—')
        # Planificado es lo que dice el POA; comprometido y ejecutado son los
        # gastos realmente cargados al plan.
        dinero(ws2, r, 10, p.monto)
        dinero(ws2, r, 11, p.comprometido)
        dinero(ws2, r, 12, p.ejecutado)
        dinero(ws2, r, 13, p.saldo)
        tot_plan += p.monto
        tot_comp += p.comprometido
        tot_ejec += p.ejecutado
        tot_saldo += p.saldo
        r += 1

    if r > 2:
        ws2.cell(row=r, column=9, value='TOTAL').font = bold
        for col, val in ((10, tot_plan), (11, tot_comp), (12, tot_ejec), (13, tot_saldo)):
            c = dinero(ws2, r, col, val)
            c.font = bold
    else:
        ws2.cell(row=2, column=1, value='Sin planes de gasto registrados.')

    for col, w in zip('ABCDEFGHIJKLM', (7, 10, 34, 34, 22, 20, 20, 26, 16, 15, 15, 15, 15)):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = 'A2'

    # ---------- HOJA 3: EGRESOS ----------
    ws3 = wb.create_sheet('Egresos')
    headers = [
        'Fecha', 'Tipo', 'Estado', 'Detalle', 'Plan de gasto', 'Transferencia',
        'Bolsa', 'SC', 'OC', 'Factura', 'Total', 'Pagado', 'Comprometido',
    ]
    escribir_encabezado(ws3, headers)

    egresos = (
        proyecto.egresos.filter(eliminado=False)
        .select_related(
            'plan_de_gasto__gasto_elegible__gasto__tipo_gasto__transferencia',
            'gasto_elegible__gasto__tipo_gasto__transferencia',
        )
        .order_by('-fecha', '-creado_en')
    )
    r = 2
    tot_total = tot_pag = tot_comp = Decimal('0')
    for e in egresos:
        total, pagado, comp = _egreso_montos(e)
        if e.plan_de_gasto_id:
            transf = str(e.plan_de_gasto.transferencia)
            plan = e.plan_de_gasto.nombre_corto
        else:
            transf = e.get_tipo_display()
            plan = '—'
        ws3.cell(row=r, column=1, value=e.fecha.strftime('%d-%m-%Y') if e.fecha else '—')
        ws3.cell(row=r, column=2, value=e.get_tipo_display())
        ws3.cell(row=r, column=3, value=e.get_estado_display())
        ws3.cell(row=r, column=4, value=_egreso_detalle(e)[:70])
        ws3.cell(row=r, column=5, value=plan)
        ws3.cell(row=r, column=6, value=transf)
        ws3.cell(row=r, column=7, value='Capital' if e.es_capital else 'Corriente')
        ws3.cell(row=r, column=8, value=e.solicitud_compra or '—')
        ws3.cell(row=r, column=9, value=e.orden_compra or '—')
        ws3.cell(row=r, column=10, value=e.factura or '—')
        dinero(ws3, r, 11, total)
        dinero(ws3, r, 12, pagado)
        dinero(ws3, r, 13, comp)
        tot_total += total
        tot_pag += pagado
        tot_comp += comp
        r += 1

    if r > 2:
        ws3.cell(row=r, column=10, value='TOTAL').font = bold
        for col, val in ((11, tot_total), (12, tot_pag), (13, tot_comp)):
            c = dinero(ws3, r, col, val)
            c.font = bold
    else:
        ws3.cell(row=2, column=1, value='Sin egresos registrados.')

    for col, w in zip('ABCDEFGHIJKLM', (12, 12, 14, 40, 30, 22, 12, 14, 14, 14, 15, 15, 15)):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = 'A2'

    # ---------- RESPUESTA ----------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = slugify(proyecto.codigo or proyecto.nombre)[:40] or f'proyecto_{proyecto.pk}'
    fecha = date.today().strftime('%Y%m%d')
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="proyecto_{nombre}_{fecha}.xlsx"'
    return resp


# =========================
# EXCEL — CARTERA
# =========================

@login_required
def exportar_cartera_excel(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return HttpResponse(
            "openpyxl no está instalado. Ejecuta: pip install openpyxl",
            status=500,
        )

    user = request.user
    jefe = es_jefe(user)
    qs = Proyecto.objects.with_resumen().prefetch_related(
        'objetivos__resultados__actividades'
    ).order_by('-fecha_creacion')
    if not jefe:
        qs = qs.filter(responsable=user)
    proyectos = list(qs)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Cartera'

    headers = [
        'Proyecto', 'Código', 'Tipo', 'Estado', 'Responsable',
        'Presupuesto total', 'Comprometido', 'Pagado', 'Total gastado',
        'Disponible', '% ejecutado', 'Cumplimiento', 'Atrasado',
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=HEADER_FILL)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    right = Alignment(horizontal='right')

    def dinero(row, col, valor):
        c = ws.cell(row=row, column=col, value=_f(valor))
        c.number_format = FMT_CLP
        c.alignment = right

    r = 2
    tot_pres = tot_comp = tot_pag = tot_gasto = tot_disp = Decimal('0')
    for p in proyectos:
        ws.cell(row=r, column=1, value=p.nombre)
        ws.cell(row=r, column=2, value=p.codigo or '—')
        ws.cell(row=r, column=3, value=p.get_tipo_display())
        ws.cell(row=r, column=4, value=p.get_estado_display())
        ws.cell(row=r, column=5, value=str(p.responsable) if p.responsable else '—')
        dinero(r, 6, p.presupuesto_total)
        dinero(r, 7, p.gastos_comprometidos)
        dinero(r, 8, p.gastos_pagados)
        dinero(r, 9, p.gastos_total)
        dinero(r, 10, p.disponible_para_gastar)
        ws.cell(row=r, column=11, value=f'{_f(p.porcentaje_gastado):.1f}%').alignment = right
        ws.cell(row=r, column=12, value=f'{_f(p.cumplimiento):.1f}%').alignment = right
        ws.cell(row=r, column=13, value='Sí' if p.atrasado else 'No')
        tot_pres += p.presupuesto_total or Decimal('0')
        tot_comp += p.gastos_comprometidos
        tot_pag += p.gastos_pagados
        tot_gasto += p.gastos_total
        tot_disp += p.disponible_para_gastar
        r += 1

    if proyectos:
        ws.cell(row=r, column=5, value='TOTAL').font = Font(bold=True)
        for col, val in ((6, tot_pres), (7, tot_comp), (8, tot_pag), (9, tot_gasto), (10, tot_disp)):
            dinero(r, col, val)
            ws.cell(row=r, column=col).font = Font(bold=True)
    else:
        ws.cell(row=2, column=1, value='No hay proyectos visibles.')

    for col, w in zip('ABCDEFGHIJKLM', (40, 14, 16, 16, 22, 17, 15, 15, 15, 15, 12, 13, 10)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = date.today().strftime('%Y%m%d')
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="cartera_proyectos_{fecha}.xlsx"'
    return resp


# =========================
# INFORME IMPRIMIBLE (PDF vía navegador)
# =========================

@login_required
def informe_proyecto(request, pk):
    proyecto = get_object_or_404(
        Proyecto.objects.prefetch_related('objetivos__resultados__actividades'),
        pk=pk,
    )

    # Gastos por transferencia (mismo criterio que graficos_proyecto).
    trans_totales = defaultdict(Decimal)
    for egreso in proyecto.egresos.filter(eliminado=False).select_related(
        'plan_de_gasto__gasto_elegible__gasto__tipo_gasto__transferencia'
    ):
        if egreso.plan_de_gasto_id:
            nombre = str(egreso.plan_de_gasto.transferencia)
        else:
            nombre = egreso.get_tipo_display()
        total, _pag, _comp = _egreso_montos(egreso)
        trans_totales[nombre] += total
    transferencias = [
        {'nombre': n, 'monto': trans_totales[n]}
        for n in sorted(trans_totales, key=lambda t: -trans_totales[t])
    ]

    objetivos = []
    for i, obj in enumerate(proyecto.objetivos.filter(eliminado=False).order_by('orden', 'id'), start=1):
        # Lo gastado del objetivo se arma sumando sus resultados, que es donde
        # los egresos quedan colgados; el objetivo no lo sabe por sí mismo.
        presupuesto = obj.presupuesto_asignado
        gastado = sum(
            (r.gastos_total for r in obj.resultados.filter(eliminado=False)),
            Decimal('0'),
        )
        objetivos.append({
            'etiqueta': f'OE{i}',
            'descripcion': obj.descripcion,
            'presupuesto': presupuesto,
            'gastado': gastado,
            'cumplimiento': _cumplimiento_objetivo(obj),
            'pct_gastado': gi.porcentaje(gastado, presupuesto),
        })

    return render(request, 'proyectos/informe_proyecto.html', {
        'proyecto': proyecto,
        'transferencias': transferencias,
        'objetivos': objetivos,
        'generado': timezone.now(),
        'auto_print': request.GET.get('print') == '1',
        # Gráficos del informe: SVG servido, no Chart.js (ver core.svg).
        'grafico_ejecucion': gi.barra_de_ejecucion(
            proyecto.gastos_pagados,
            proyecto.gastos_comprometidos,
            proyecto.presupuesto_total,
        ),
        'grafico_bolsas': gi.barras_de_bolsas([
            ('Corriente', proyecto.gastos_corriente, proyecto.presupuesto_corriente, gi.COLOR_CORRIENTE),
            ('Capital', proyecto.gastos_capital, proyecto.presupuesto_capital, gi.COLOR_CAPITAL),
        ]),
        'grafico_transferencias': gi.barras_de_montos(
            [(t['nombre'], t['monto']) for t in transferencias],
            ancho=gi.ANCHO_TOTAL, col_etiqueta=168, largo_rotulo=30,
        ),
        'grafico_objetivos': gi.barras_de_objetivos([
            {
                'etiqueta': f"{o['etiqueta']}: {gi.recortar(o['descripcion'] or 'Sin descripción', 24)}",
                'titulo': o['descripcion'] or '',
                'fisico': o['cumplimiento'],
                'financiero': o['pct_gastado'],
            }
            for o in objetivos
        ]),
    })
