from calendar import Calendar, monthrange
from collections import Counter
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import StringIO
from pathlib import Path
from uuid import uuid4

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.management import call_command
from django.core.management.base import CommandError
from django.db.models import Count, Q, Sum
from django.http import Http404, HttpResponseBadRequest
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_POST

from . import flujo
from .calendario import BLOQUES
from .forms import (
    ActividadForm,
    CostoActividadForm,
    CostoDirectoForm,
    CostoTransversalForm,
    GastoExtraFormSet,
    LineaFinancieraForm,
    MetaAnualForm,
    PropuestaForm,
    SesionClaseForm,
    SubirTableroForm,
    SupuestosFinancierosForm,
)
from .graficos import (
    CERTEZA,
    COLOR_RIESGO,
    ETIQUETA_CERTEZA,
    MESES_CORTOS,
    SERIES,
    TINTA,
)
from .importador import ErrorImportacion, ImportadorTablero
from .importador import clave as clave_nombre
from .models import (
    ALERTA_CRITICOS,
    CLASES_ALERTA,
    CLASES_RIESGO,
    GRUPO_ENCARGADO,
    GRUPO_PROFESIONAL,
    Actividad,
    Contacto,
    CostoDirecto,
    CostoTransversal,
    DiaActividad,
    Estado,
    Etapa,
    Feriado,
    Institucion,
    ItemChecklist,
    LineaFinanciera,
    MetaAnual,
    Origen,
    Propuesta,
    Relator,
    ReservaZoom,
    SalaZoom,
    SesionClase,
    SupuestosFinancieros,
    equipo_otec,
    rol_otec,
)

CERO = Decimal("0")

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre",
    12: "Diciembre",
}


def _actividades_base():
    """Queryset con todo lo que necesitan las properties calculadas.

    El prefetch de ``items__plantilla`` es lo que evita una consulta por cada
    ítem al calcular avance, alerta y riesgo.
    """
    return (
        Actividad.objects
        .select_related("propuesta", "propuesta__institucion", "relator", "costos")
        .prefetch_related("items__plantilla", "gastos_extra", "sesiones")
    )


def _rango_activo(actividad):
    """(inicio, término) de una actividad, tolerando que falte el término."""
    inicio = actividad.fecha_inicio
    if not inicio:
        return None
    return inicio, (actividad.fecha_termino or inicio)


def _solapan(a, b):
    ra, rb = _rango_activo(a), _rango_activo(b)
    if not ra or not rb:
        return False
    return ra[0] <= rb[1] and rb[0] <= ra[1]


def _max_simultaneas(actividades):
    """Máximo de actividades que una persona lleva a la vez.

    Barrido de eventos: +1 al empezar cada actividad, −1 al día siguiente de
    terminarla; el pico del acumulado es la carga simultánea máxima.
    """
    eventos = []
    for a in actividades:
        rango = _rango_activo(a)
        if rango:
            eventos.append((rango[0], 1))
            eventos.append((rango[1] + timedelta(days=1), -1))
    if not eventos:
        return 0
    eventos.sort()
    actual = pico = 0
    for _fecha, delta in eventos:
        actual += delta
        pico = max(pico, actual)
    return pico


@login_required
def indicadores(request):
    """Seguimiento de indicadores de gestión y carga de trabajo del equipo."""
    actividades = list(_actividades_base())
    ganadas = [
        a for a in actividades
        if a.propuesta.estado_comercial == Propuesta.EstadoComercial.GANADA
    ]

    anio = max((p.anio for p in Propuesta.objects.all()), default=date.today().year)
    meta = MetaAnual.objects.filter(anio=anio).first()
    adjudicado = sum(a.monto_adjudicado for a in actividades)

    # --- Indicadores ---
    avances = [a.avance_checklist for a in actividades if a.avance_checklist is not None]
    margenes = [a.margen_estimado for a in actividades if a.margen_estimado is not None]

    decretadas = [
        p for p in Propuesta.objects.all()
        if p.tiempo_decreto_dias is not None
    ]
    dias_decreto = [p.tiempo_decreto_dias for p in decretadas]

    # Plazo de cobro: se usa el pago efectivo cuando existe y el estimado si no,
    # que es lo que hace la planilla. Contar solo los efectivos daría 0 días,
    # porque las pocas líneas ya cobradas se facturaron y pagaron el mismo día.
    cobros, cobros_reales = [], 0
    for l in LineaFinanciera.objects.exclude(fecha_facturacion=None):
        pago = l.fecha_pago_efectiva or l.fecha_pago_estimada
        if not pago:
            continue
        cobros.append((pago - l.fecha_facturacion).days)
        cobros_reales += int(l.fecha_pago_efectiva is not None)

    resumen_flujo = None
    if LineaFinanciera.objects.exists():
        resumen_flujo = flujo.resumen(anio)

    def indicador(nombre, valor, detalle="", pct=None, clase=""):
        return {
            "nombre": nombre, "valor": valor, "detalle": detalle,
            "pct": pct, "clase": clase,
        }

    ejecutadas = [
        a for a in actividades
        if a.estado_ejecucion in (
            Actividad.EstadoEjecucion.EJECUTADA, Actividad.EstadoEjecucion.EN_EJECUCION
        )
    ]

    lista_indicadores = [
        indicador(
            "Avance de la meta financiera",
            f"{adjudicado / meta.monto * 100:.1f}%" if meta and meta.monto else "—",
            f"${adjudicado:,.0f} de ${meta.monto:,.0f}".replace(",", ".") if meta else "sin meta cargada",
            pct=float(adjudicado / meta.monto * 100) if meta and meta.monto else None,
        ),
        indicador(
            "Ingresos asegurados",
            f"{resumen_flujo['pct_asegurados']:.1f}%" if resumen_flujo and resumen_flujo["pct_asegurados"] else "—",
            "cobrado más contratado en firme" if resumen_flujo else "requiere el flujo de caja",
            pct=float(resumen_flujo["pct_asegurados"]) if resumen_flujo and resumen_flujo["pct_asegurados"] else None,
        ),
        indicador(
            "Conversión comercial",
            f"{len(ganadas)} de {len(actividades)}",
            "actividades en propuestas ganadas",
            pct=len(ganadas) * 100 / len(actividades) if actividades else None,
        ),
        indicador(
            "Avance del checklist",
            f"{round(sum(avances) * 100 / len(avances))}%" if avances else "—",
            f"promedio de {len(avances)} actividades",
            pct=sum(avances) * 100 / len(avances) if avances else None,
        ),
        indicador(
            "Personas por capacitar",
            f"{sum(a.n_participantes for a in ganadas)}",
            f"{sum(a.horas for a in ganadas)} horas comprometidas",
        ),
        indicador(
            "Actividades en ejecución",
            f"{len(ejecutadas)}",
            f"de {len(ganadas)} ganadas",
        ),
        indicador(
            "Tiempo de decretación",
            f"{round(sum(dias_decreto) / len(dias_decreto))} días" if dias_decreto else "—",
            f"promedio de {len(dias_decreto)} propuestas decretadas"
            if dias_decreto else "aún no hay decretos con fecha",
        ),
        indicador(
            "Plazo de cobro",
            f"{round(sum(cobros) / len(cobros))} días" if cobros else "—",
            f"de la factura al pago en {len(cobros)} líneas "
            f"({cobros_reales} ya cobradas, el resto con fecha estimada)"
            if cobros else "requiere el flujo de caja",
        ),
        indicador(
            "Margen estimado",
            f"{round(sum(margenes) / len(margenes) * 100)}%" if margenes else "—",
            f"promedio de {len(margenes)} actividades con valor ofertado",
        ),
    ]

    # --- Carga laboral del equipo OTEC ---
    # Se reparte entre quienes están en los grupos «Encargado OTEC» y
    # «Profesional OTEC». Los relatores no entran: ellos dictan el curso, no lo
    # gestionan.
    por_actividad = {}
    for a in actividades:
        por_actividad[a.pk] = list(a.responsables.all())

    equipo = []
    for persona in equipo_otec():
        suyas = [a for a in actividades if persona in por_actividad[a.pk]]
        fechas = [f for a in suyas for f in _rango_activo(a) or ()]
        equipo.append({
            "persona": persona,
            "nombre": persona.get_full_name() or persona.username,
            "rol": rol_otec(persona),
            "actividades": suyas,
            "n": len(suyas),
            "horas": sum(a.horas for a in suyas),
            "participantes": sum(a.n_participantes for a in suyas),
            "criticos": sum(1 for a in suyas if a.pendientes_criticos),
            "simultaneas": _max_simultaneas(suyas),
            "desde": min(fechas) if fechas else None,
            "hasta": max(fechas) if fechas else None,
        })
    # Encargados primero, y dentro de cada rol por carga.
    equipo.sort(key=lambda x: (x["rol"] != "Encargado", -x["n"], x["nombre"]))

    sin_responsable = [a for a in actividades if not por_actividad[a.pk]]

    # Nombres que la planilla menciona y que todavía no tienen usuario: es lo
    # que falta para que la carga laboral quede completa.
    # Se reconoce a la persona por nombre completo, nombre de pila o usuario:
    # la planilla la nombra de cualquiera de esas formas, igual que en el
    # importador.
    conocidos = {
        clave_nombre(etiqueta)
        for p in equipo
        for etiqueta in (p["nombre"], p["persona"].first_name, p["persona"].username)
        if etiqueta
    }
    nombres_planilla = Counter()
    for a in actividades:
        for n in a.responsable_seguimiento.split("/"):
            n = n.strip()
            if n and clave_nombre(n) not in conocidos:
                nombres_planilla[n] += 1
    sin_usuario = nombres_planilla.most_common()

    # --- Carga por mes ---
    meses_rango = sorted({
        (f.year, f.month)
        for a in actividades
        for f in (_rango_activo(a) or ())
    })
    carga_mensual = []
    for anio_m, mes_m in meses_rango:
        primero = date(anio_m, mes_m, 1)
        ultimo = date(anio_m + (mes_m == 12), (mes_m % 12) + 1, 1) - timedelta(days=1)
        activas = [
            a for a in actividades
            if (r := _rango_activo(a)) and r[0] <= ultimo and primero <= r[1]
        ]
        inician = [a for a in actividades if a.fecha_inicio and
                   (a.fecha_inicio.year, a.fecha_inicio.month) == (anio_m, mes_m)]
        personas = {
            p.pk for a in activas for p in por_actividad[a.pk]
        }
        carga_mensual.append({
            "etiqueta": f"{MESES_CORTOS[mes_m]} {anio_m}",
            "activas": len(activas),
            "horas": sum(a.horas for a in inician),
            "personas": len(personas),
        })

    g_carga = {
        "labels": [m["etiqueta"] for m in carga_mensual],
        "activas": [m["activas"] for m in carga_mensual],
        "personas": [m["personas"] for m in carga_mensual],
    }
    g_horas = {
        "labels": [m["etiqueta"] for m in carga_mensual],
        "data": [m["horas"] for m in carga_mensual],
    }

    return render(request, "otec/indicadores.html", {
        "anio": anio,
        "indicadores": lista_indicadores,
        "equipo": equipo,
        "sin_responsable": sin_responsable,
        "sin_usuario": sin_usuario,
        "hay_equipo": bool(equipo),
        "carga_mensual": carga_mensual,
        "g_carga": g_carga,
        "g_horas": g_horas,
        "tinta": TINTA,
        "grupo_encargado": GRUPO_ENCARGADO,
        "grupo_profesional": GRUPO_PROFESIONAL,
    })


@login_required
def otec_home(request):
    """Entrada de OTEC: dos secciones para elegir."""
    actividades = list(_actividades_base())
    ganadas = [
        a for a in actividades
        if a.propuesta.estado_comercial == Propuesta.EstadoComercial.GANADA
    ]
    avances = [a.avance_checklist for a in actividades if a.avance_checklist is not None]

    secciones = [
        {
            "icono": "📋",
            "titulo": "Gestión de capacitación",
            "texto": (
                "Propuestas, cursos y su checklist, calendario de clases y salas, "
                "gráficos de cartera y flujo de caja."
            ),
            "url": reverse("otec:panel_gestion"),
            "dato": (
                f"{len(actividades)} cursos en {Propuesta.objects.count()} propuestas · "
                f"{len(ganadas)} ganados"
            ),
        },
        {
            "icono": "🎯",
            "titulo": "Seguimiento de indicadores y carga laboral",
            "texto": (
                "Indicadores de gestión del año y reparto del trabajo entre "
                "relatores y responsables, con alerta de cruces."
            ),
            "url": reverse("otec:indicadores"),
            "dato": (
                f"avance de checklist {round(sum(avances) * 100 / len(avances))}% · "
                f"{equipo_otec().count()} personas en el equipo"
                if avances else "sin datos aún"
            ),
        },
    ]
    return render(request, "otec/home.html", {"secciones": secciones})


@login_required
def panel_gestion(request):
    """Tablero de la gestión de capacitación."""
    actividades = list(_actividades_base())
    ganadas = [a for a in actividades if a.propuesta.estado_comercial == Propuesta.EstadoComercial.GANADA]

    totales = Actividad.objects.aggregate(
        valor_ofertado=Sum("valor_ofertado"),
        adjudicado=Sum("monto_adjudicado"),
        facturado=Sum("monto_facturado"),
        pagado=Sum("monto_pagado"),
    )
    for clave, valor in totales.items():
        totales[clave] = valor or CERO

    # Los costos ya no son columnas de la actividad sino su desglose, así que
    # se suman en Python sobre lo que ya viene traído.
    totales["costo_relatoria"] = sum((a.costo_relatoria for a in actividades), CERO)
    totales["otros_gastos"] = sum((a.otros_gastos for a in actividades), CERO)
    excedente = totales["valor_ofertado"] - totales["costo_relatoria"] - totales["otros_gastos"]

    # Riesgo y alerta se calculan en Python: son propiedades derivadas del
    # checklist, no columnas de la base.
    riesgos = [
        {"etiqueta": etiqueta, "n": n, "clase": CLASES_RIESGO.get(etiqueta, "")}
        for etiqueta, n in Counter(a.riesgo_ejecutivo for a in actividades).most_common()
    ]
    alertas = [
        {"etiqueta": etiqueta, "n": n, "clase": CLASES_ALERTA.get(etiqueta, "")}
        for etiqueta, n in Counter(a.alerta_checklist for a in actividades).most_common()
    ]

    avances = [a.avance_checklist for a in actividades if a.avance_checklist is not None]
    proximas = sorted(
        (a for a in actividades if a.proxima_fecha_critica),
        key=lambda a: a.proxima_fecha_critica,
    )[:6]

    return render(request, "otec/panel.html", {
        "total_actividades": len(actividades),
        "total_ganadas": len(ganadas),
        "total_propuestas": Propuesta.objects.count(),
        "personas_a_capacitar": sum(a.n_participantes for a in ganadas),
        "horas_totales": sum(a.horas for a in ganadas),
        "totales": totales,
        "excedente": excedente,
        "pendiente_facturar": totales["adjudicado"] - totales["facturado"],
        "avance_promedio": round(sum(avances) * 100 / len(avances)) if avances else None,
        "riesgos": riesgos,
        "alertas": alertas,
        "proximas": proximas,
    })


@login_required
def lista_actividades(request):
    q = request.GET.get("q", "").strip()
    estado_ejecucion = request.GET.get("estado_ejecucion", "")
    estado_comercial = request.GET.get("estado_comercial", "")
    institucion_id = request.GET.get("institucion", "")
    riesgo = request.GET.get("riesgo", "")

    qs = _actividades_base()

    if q:
        qs = qs.filter(
            Q(nombre__icontains=q)
            | Q(propuesta__codigo__icontains=q)
            | Q(propuesta__institucion__nombre__icontains=q)
            | Q(relator__nombre__icontains=q)
        )
    if estado_ejecucion:
        qs = qs.filter(estado_ejecucion=estado_ejecucion)
    if estado_comercial:
        qs = qs.filter(propuesta__estado_comercial=estado_comercial)
    if institucion_id.isdigit():
        qs = qs.filter(propuesta__institucion_id=int(institucion_id))

    actividades = list(qs)

    # El riesgo es calculado, así que se filtra después de materializar.
    if riesgo:
        actividades = [a for a in actividades if a.riesgo_ejecutivo == riesgo]

    riesgos_disponibles = sorted({a.riesgo_ejecutivo for a in _actividades_base()})

    contexto = {
        "actividades": actividades,
        "instituciones": Institucion.objects.annotate(
            n=Count("propuestas__actividades")
        ).filter(n__gt=0),
        "estados_ejecucion": Actividad.EstadoEjecucion.choices,
        "estados_comerciales": Propuesta.EstadoComercial.choices,
        "riesgos_disponibles": riesgos_disponibles,
        "filtros": {
            "q": q,
            "estado_ejecucion": estado_ejecucion,
            "estado_comercial": estado_comercial,
            "institucion": institucion_id,
            "riesgo": riesgo,
        },
        "hay_filtros": any([q, estado_ejecucion, estado_comercial, institucion_id, riesgo]),
    }
    return render(request, "otec/lista_actividades.html", contexto)


def _contexto_checklist(actividad):
    """Ítems agrupados por etapa, en el orden del catálogo."""
    etapas = []
    for resumen in actividad.avance_por_etapa():
        etapas.append({
            **resumen,
            "items": [i for i in actividad.items.all() if i.plantilla.etapa == resumen["etapa"]],
        })

    # Las etapas cuyos ítems son todos "No aplica" no entran en el avance,
    # pero igual deben poder verse y editarse.
    con_avance = {e["etapa"] for e in etapas}
    for etapa, label in Etapa.choices:
        if etapa in con_avance:
            continue
        items = [i for i in actividad.items.all() if i.plantilla.etapa == etapa]
        if items:
            etapas.append({
                "etapa": etapa, "label": label, "completados": 0,
                "total": 0, "pct": None, "items": items,
            })

    return {"actividad": actividad, "etapas": etapas, "estados": Estado.choices}


@login_required
def detalle_actividad(request, pk):
    actividad = get_object_or_404(_actividades_base(), pk=pk)
    contexto = {
        "propuesta": actividad.propuesta,
        **_contexto_checklist(actividad),
    }
    return render(request, "otec/detalle_actividad.html", contexto)


@login_required
@require_POST
def actualizar_item(request, pk):
    """Cambia el estado de un ítem del checklist (HTMX).

    Devuelve el checklist completo, no solo la fila, para que las barras de
    avance por etapa y el total queden al día en el mismo swap.
    """
    item = get_object_or_404(ItemChecklist.objects.select_related("actividad"), pk=pk)

    nuevo = request.POST.get("estado", "")
    if nuevo not in Estado.values:
        return HttpResponseBadRequest("Estado no válido.")

    item.estado = nuevo
    # Queda marcado para que una recarga del Excel no pise la edición manual.
    item.editado_en_sistema = True
    item.save(update_fields=["estado", "editado_en_sistema", "actualizado_en"])

    actividad = get_object_or_404(_actividades_base(), pk=item.actividad_id)
    return render(request, "otec/partials/checklist.html", _contexto_checklist(actividad))


@login_required
def graficos_otec(request):
    """Tablero de gráficos de la cartera.

    Cada bloque de datos se arma acá y viaja por ``json_script``; la plantilla
    solo dibuja. Los colores salen de ``otec.graficos`` para que la asignación
    categórica siga siempre el mismo orden validado.
    """
    actividades = list(_actividades_base())
    ganadas = [
        a for a in actividades
        if a.propuesta.estado_comercial == Propuesta.EstadoComercial.GANADA
    ]

    adjudicado = sum(a.monto_adjudicado for a in actividades)
    ofertado = sum(a.valor_ofertado for a in actividades)
    costos = sum((a.costo_total for a in actividades), CERO)

    anio = max((p.anio for p in Propuesta.objects.all()), default=date.today().year)
    meta = MetaAnual.objects.filter(anio=anio).first()

    # --- Adjudicado por institución (magnitud por identidad) ---
    por_institucion = Counter()
    for a in actividades:
        if a.monto_adjudicado:
            por_institucion[a.propuesta.institucion.nombre] += float(a.monto_adjudicado)
    instituciones = por_institucion.most_common()
    # Más de ocho identidades no se colorean: el resto se agrupa.
    if len(instituciones) > len(SERIES):
        cabeza = instituciones[:len(SERIES) - 1]
        resto = sum(v for _, v in instituciones[len(SERIES) - 1:])
        instituciones = cabeza + [("Otras instituciones", resto)]

    grafico_instituciones = {
        "labels": [n for n, _ in instituciones],
        "data": [v for _, v in instituciones],
        "colors": [SERIES[i] for i in range(len(instituciones))],
    }

    # --- Estado comercial de las actividades ---
    estados = Counter(a.propuesta.get_estado_comercial_display() for a in actividades)
    grafico_estados = {
        "labels": list(estados.keys()),
        "data": list(estados.values()),
        "colors": [SERIES[i] for i in range(len(estados))],
    }

    # --- Riesgo ejecutivo (paleta de estados, no categórica) ---
    riesgos = Counter(a.riesgo_ejecutivo for a in actividades)
    grafico_riesgo = {
        "labels": list(riesgos.keys()),
        "data": list(riesgos.values()),
        "colors": [COLOR_RIESGO.get(k, TINTA["apagada"]) for k in riesgos],
    }

    # --- Avance del checklist por etapa, sobre toda la cartera ---
    etapas = []
    for etapa, label in Etapa.choices:
        evaluables = completados = 0
        for a in actividades:
            for i in a.items.all():
                if i.plantilla.etapa != etapa or i.estado == Estado.NO_APLICA:
                    continue
                evaluables += 1
                completados += int(i.estado == Estado.SI)
        if evaluables:
            etapas.append({
                "label": label,
                "pct": round(completados * 100 / evaluables),
                "completados": completados,
                "total": evaluables,
            })
    grafico_etapas = {
        "labels": [e["label"] for e in etapas],
        "data": [e["pct"] for e in etapas],
        "detalle": [f"{e['completados']} de {e['total']}" for e in etapas],
    }

    # --- Calendarización: días de ejecución por mes ---
    dias = DiaActividad.objects.order_by("fecha")
    por_mes = Counter((d.fecha.year, d.fecha.month) for d in dias)
    meses = sorted(por_mes)
    grafico_meses = {
        "labels": [f"{MESES_CORTOS[m]} {a}" for a, m in meses],
        "data": [por_mes[k] for k in meses],
    }

    # --- Horas de sala Zoom por mes y sala (dos series) ---
    reservas = ReservaZoom.objects.select_related("sala").filter(hora_inicio__isnull=False)
    salas = list(SalaZoom.objects.filter(activa=True))
    horas = {s.pk: Counter() for s in salas}
    for r in reservas:
        if r.sala_id in horas:
            horas[r.sala_id][(r.fecha.year, r.fecha.month)] += r.duracion_horas or 0
    meses_zoom = sorted({k for c in horas.values() for k in c})
    grafico_zoom = {
        "labels": [f"{MESES_CORTOS[m]} {a}" for a, m in meses_zoom],
        "series": [
            {
                "label": s.nombre,
                "data": [round(horas[s.pk][k], 1) for k in meses_zoom],
                "color": SERIES[i],
            }
            for i, s in enumerate(salas)
        ],
    }

    # --- Excedente estimado por actividad (las de mayor aporte) ---
    con_excedente = sorted(
        (a for a in actividades if a.valor_ofertado),
        key=lambda a: a.excedente_estimado,
        reverse=True,
    )[:10]
    grafico_excedente = {
        "labels": [a.nombre[:44] for a in con_excedente],
        "data": [float(a.excedente_estimado) for a in con_excedente],
        "margen": [a.margen_estimado_pct for a in con_excedente],
    }

    return render(request, "otec/graficos.html", {
        "anio": anio,
        "meta": meta,
        "adjudicado": adjudicado,
        "ofertado": ofertado,
        "excedente": ofertado - costos,
        "brecha": (meta.monto - adjudicado) if meta else None,
        "avance_meta": (
            round(adjudicado * 100 / meta.monto, 1) if meta and meta.monto else None
        ),
        "total_actividades": len(actividades),
        "total_ganadas": len(ganadas),
        "personas": sum(a.n_participantes for a in ganadas),
        "etapas": etapas,
        "instituciones": instituciones,
        "g_instituciones": grafico_instituciones,
        "g_estados": grafico_estados,
        "g_riesgo": grafico_riesgo,
        "g_etapas": grafico_etapas,
        "g_meses": grafico_meses,
        "g_zoom": grafico_zoom,
        "g_excedente": grafico_excedente,
        "tinta": TINTA,
    })


@login_required
def flujo_caja(request):
    """Gráficos del flujo de caja, recalculados desde las líneas y los costos."""
    anio = request.GET.get("anio", "")
    anios = sorted(
        {a for a in LineaFinanciera.objects
            .exclude(fecha_pago_estimada=None)
            .values_list("fecha_pago_estimada__year", flat=True)}
    )
    if not anios:
        # Puede haber datos cargados y ninguna fecha de pago. Sin fecha no
        # entran en ningún mes, pero igual tienen que poder abrirse para
        # ponérsela: la pantalla vacía dejaría el dato inalcanzable.
        if not LineaFinanciera.objects.exists() and not CostoTransversal.objects.exists():
            return render(request, "otec/flujo.html", {"sin_datos": True})
        anios = [date.today().year]

    anio = int(anio) if anio.isdigit() and int(anio) in anios else anios[-1]
    r = flujo.resumen(anio)
    filas = r["filas"]

    # --- Ingresos por mes, apilados por nivel de certeza (escala ordenada) ---
    g_ingresos = {
        "labels": [MESES_CORTOS[f["mes"]] for f in filas],
        "series": [
            {
                "label": ETIQUETA_CERTEZA[c.value],
                "data": [float(f["ingresos"][c.value]) for f in filas],
                "color": CERTEZA[c.value],
            }
            for c in LineaFinanciera.Certeza
        ],
    }

    # --- Saldo de caja proyectado ---
    proyectados = [f for f in filas if f["proyectado"]]
    g_saldo = {
        "labels": [MESES_CORTOS[f["mes"]] for f in proyectados],
        "data": [float(f["saldo_final"]) for f in proyectados],
        "minimo": float(r["supuestos"].saldo_minimo or 0),
        "neto": [float(f["neto"]) for f in proyectados],
    }

    # --- Estructura de costos ---
    categorias = Counter()
    for linea in LineaFinanciera.objects.select_related("costo"):
        costo = getattr(linea, "costo", None)
        if not costo:
            continue
        for item in costo.por_categoria():
            categorias[item["label"]] += float(item["monto"])
    for c in CostoTransversal.objects.filter(incluir_en_flujo=True):
        categorias[f"Transversal · {c.tipo or 'otros'}"] += float(c.monto)
    costos_ordenados = categorias.most_common()
    g_costos = {
        "labels": [k for k, _ in costos_ordenados],
        "data": [v for _, v in costos_ordenados],
    }

    # --- Escenarios contra la meta ---
    meta = MetaAnual.objects.filter(anio=anio).first()
    acumulado, escenarios = Decimal("0"), []
    for c in LineaFinanciera.Certeza:
        acumulado += r["por_certeza"][c.value]
        escenarios.append({
            "label": ETIQUETA_CERTEZA[c.value],
            "acumulado": float(acumulado),
            "pct": float(acumulado / meta.monto * 100) if meta and meta.monto else None,
            "color": CERTEZA[c.value],
        })

    # --- Resultado por línea, las de mayor aporte ---
    top = sorted(r["resultados"], key=lambda x: x["resultado"], reverse=True)[:12]
    g_resultado = {
        "labels": [x["linea"].codigo for x in top],
        "descripcion": [x["linea"].descripcion[:60] for x in top],
        "data": [float(x["resultado"]) for x in top],
        "margen": [round(float(x["margen"]) * 100) if x["margen"] is not None else None for x in top],
        "proyeccion": [x["linea"].es_proyeccion for x in top],
    }

    # --- Detalle editable: cada línea con lo que aporta al año elegido ---
    # Se listan todas, no solo las del año: una línea sin fecha de pago no
    # entra en ningún mes, y es justamente la que hay que poder corregir.
    detalle_lineas = [
        {**x, "en_el_anio": _cae_en(x["linea"].fecha_ingreso, anio)}
        for x in sorted(r["resultados"], key=lambda x: x["linea"].codigo)
    ]
    transversales = [
        {"costo": c, "en_el_anio": _cae_en(c.fecha_pago, anio)}
        for c in CostoTransversal.objects.all()
    ]

    return render(request, "otec/flujo.html", {
        "anio": anio,
        "anios": anios,
        "r": r,
        "supuestos": r["supuestos"],
        "filas": filas,
        "meta": meta,
        "escenarios": escenarios,
        "costos_ordenados": costos_ordenados,
        "detalle_lineas": detalle_lineas,
        "transversales": transversales,
        "g_ingresos": g_ingresos,
        "g_saldo": g_saldo,
        "g_costos": g_costos,
        "g_resultado": g_resultado,
        "tinta": TINTA,
        "n_proyeccion": LineaFinanciera.objects.filter(
            certeza__in=[LineaFinanciera.Certeza.PROBABLE, LineaFinanciera.Certeza.PROYECTADO]
        ).count(),
        "n_lineas": LineaFinanciera.objects.count(),
    })


def _cae_en(fecha, anio):
    return bool(fecha) and fecha.year == anio


def _volver_al_flujo(request):
    """URL del flujo conservando el año desde el que se abrió la edición."""
    url = reverse("otec:flujo_caja")
    anio = request.GET.get("anio", "")
    return f"{url}?anio={anio}" if anio.isdigit() else url


@login_required
def editar_supuestos(request, anio):
    """Los parámetros del año. Si no existen, se crean al guardar."""
    supuestos = (
        SupuestosFinancieros.objects.filter(anio=anio).first()
        or SupuestosFinancieros(anio=anio)
    )
    form = SupuestosFinancierosForm(request.POST or None, instance=supuestos)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f"Supuestos de {anio} actualizados.")
        return redirect(_volver_al_flujo(request))

    return render(request, "otec/flujo_form.html", {
        "titulo": f"Supuestos financieros {anio}",
        "descripcion": (
            "Con estos parámetros se recalcula todo el flujo: cambiar uno mueve "
            "el saldo proyectado de los doce meses."
        ),
        "bloques": [{"titulo": "Parámetros del año", "form": form}],
        "volver": _volver_al_flujo(request),
    })


@login_required
def editar_meta(request, anio):
    """La meta anual de ingresos. Si no existe, se crea al guardar."""
    meta = MetaAnual.objects.filter(anio=anio).first() or MetaAnual(anio=anio)
    form = MetaAnualForm(request.POST or None, instance=meta)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f"Meta de {anio} actualizada.")
        return redirect(_volver_al_flujo(request))

    return render(request, "otec/flujo_form.html", {
        "titulo": f"Meta de ingresos {anio}",
        "descripcion": "Es la referencia contra la que se miden los escenarios de ingreso.",
        "bloques": [{"titulo": f"Meta {anio}", "form": form}],
        "volver": _volver_al_flujo(request),
    })


@login_required
def editar_linea(request, pk):
    """La línea de ingreso y sus costos directos, que van juntos.

    El resultado de una línea es ingreso menos costos, así que editarlos por
    separado obligaría a dar dos veces la misma vuelta.
    """
    linea = get_object_or_404(
        LineaFinanciera.objects.select_related("institucion", "costo"), pk=pk
    )
    costo = getattr(linea, "costo", None) or CostoDirecto(linea=linea)

    form = LineaFinancieraForm(request.POST or None, instance=linea)
    costos = CostoDirectoForm(request.POST or None, instance=costo, prefix="costos")

    if request.method == "POST" and form.is_valid() and costos.is_valid():
        form.save()
        guardado = costos.save(commit=False)
        guardado.linea = linea
        guardado.save()
        messages.success(request, f"Línea «{linea.codigo}» actualizada.")
        return redirect(_volver_al_flujo(request))

    return render(request, "otec/flujo_form.html", {
        "titulo": f"Editar {linea.codigo}",
        "descripcion": linea.descripcion,
        "bloques": [
            {"titulo": "Ingreso", "form": form},
            {
                "titulo": "Costos directos",
                "form": costos,
                "nota": (
                    "Se descuentan del ingreso para calcular el resultado de la "
                    "línea y salen de la caja en su propia fecha de pago."
                ),
            },
        ],
        "volver": _volver_al_flujo(request),
    })


@login_required
def editar_costo_transversal(request, pk):
    """Un costo del área que no cuelga de ningún curso."""
    costo = get_object_or_404(CostoTransversal, pk=pk)
    form = CostoTransversalForm(request.POST or None, instance=costo)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f"Costo «{costo.codigo}» actualizado.")
        return redirect(_volver_al_flujo(request))

    return render(request, "otec/flujo_form.html", {
        "titulo": f"Editar {costo.codigo}",
        "descripcion": costo.descripcion,
        "bloques": [{"titulo": "Costo transversal", "form": form}],
        "volver": _volver_al_flujo(request),
    })


def _dias_derivados(actividad, feriados):
    """Días hábiles entre el inicio y el término de una actividad.

    Sirve para las actividades que no tienen días marcados: las creadas en el
    sistema y las que en la planilla venían con la fila vacía. Es una
    estimación a partir de las fechas, y la vista la dibuja distinto para que
    no se confunda con lo que sí está calendarizado.
    """
    inicio = actividad.fecha_inicio
    if not inicio:
        return []
    fin = actividad.fecha_termino or inicio
    if fin < inicio:
        return []

    dias, actual = [], inicio
    while actual <= fin:
        if actual.weekday() < 5 and actual not in feriados:
            dias.append(actual)
        actual += timedelta(days=1)
    return dias


def _cursos_del_gantt():
    """Cada curso con los días que ocupa, vengan marcados o estimados.

    Unifica las dos fuentes en una sola lista para que el resto de la vista no
    tenga que recorrerlas por separado: la distinción solo vuelve a importar al
    dibujar la celda, que se pinta distinto según de dónde salga el día.
    """
    feriados = set(Feriado.objects.values_list("fecha", flat=True))

    calendarizadas = {}
    for d in (
        DiaActividad.objects
        .select_related("actividad", "actividad__propuesta__institucion")
        .order_by("fecha")
    ):
        calendarizadas.setdefault(d.actividad, {})[d.fecha] = d

    cursos = [
        {"actividad": a, "mapa": mapa, "marcados": set(mapa), "estimada": False}
        for a, mapa in calendarizadas.items()
    ]

    # Las que no tienen días marcados pero sí fechas: se estima su tramo.
    for actividad in (
        Actividad.objects
        .select_related("propuesta__institucion")
        .exclude(pk__in=[a.pk for a in calendarizadas])
        .exclude(fecha_inicio=None)
    ):
        estimados = _dias_derivados(actividad, feriados)
        if estimados:
            cursos.append({
                "actividad": actividad, "mapa": {},
                "marcados": set(estimados), "estimada": True,
            })

    cursos.sort(key=lambda c: (min(c["marcados"]), c["actividad"].nombre))
    return cursos, feriados


@login_required
def carta_gantt(request):
    """Grilla actividad × día con los días de ejecución y de cierre."""
    cursos, feriados_todos = _cursos_del_gantt()
    if not cursos:
        return render(request, "otec/carta_gantt.html", {"sin_datos": True})

    # Filtro por curso: aísla una actividad y encoge la grilla a sus días, que
    # es lo que hace útil mirar uno solo.
    curso_param = request.GET.get("curso", "")
    curso_activo = None
    if curso_param.isdigit():
        curso_activo = next(
            (c for c in cursos if c["actividad"].pk == int(curso_param)), None
        )
    visibles = [curso_activo] if curso_activo else cursos

    fechas = sorted({f for c in visibles for f in c["marcados"]})

    # Filtro por mes: mostrar 124 columnas de una vez es ilegible. Los meses
    # ofrecidos son los del tramo visible, así que nunca se llega a uno vacío.
    meses = sorted({(f.year, f.month) for f in fechas})
    mes_activo = None
    try:
        elegido = tuple(int(p) for p in request.GET.get("mes", "").split("-"))
    except ValueError:
        elegido = None
    if elegido in meses:
        mes_activo = elegido
        fechas = [f for f in fechas if (f.year, f.month) == mes_activo]

    en_rango = set(fechas)
    feriados = feriados_todos & en_rango

    filas = []
    for curso in visibles:
        if not curso["marcados"] & en_rango:
            continue  # este curso no tiene nada en el tramo elegido
        filas.append({
            "actividad": curso["actividad"],
            "estimada": curso["estimada"],
            "total": len(curso["marcados"]),
            "celdas": [
                {
                    "fecha": f,
                    "dia": curso["mapa"].get(f),
                    "derivado": curso["estimada"] and f in curso["marcados"],
                    "feriado": f in feriados,
                }
                for f in fechas
            ],
        })

    # Fila "actividades en paralelo": en la planilla era una fórmula. Cuenta
    # sobre todos los cursos y no solo sobre los visibles: al aislar uno, lo
    # que interesa es justamente cuántos más corren esos mismos días.
    conteos = [sum(1 for c in cursos if f in c["marcados"]) for f in fechas]
    max_paralelo = max(conteos, default=0)
    paralelo = [
        {
            "fecha": f,
            "n": n,
            "alto": round(n * 100 / max_paralelo) if max_paralelo else 0,
            "feriado": f in feriados,
        }
        for f, n in zip(fechas, conteos)
    ]

    # Cabecera agrupada por mes, para no repetir el nombre en cada columna.
    cabecera = []
    for f in fechas:
        etiqueta = MESES_ES[f.month]
        if cabecera and cabecera[-1]["mes"] == etiqueta:
            cabecera[-1]["ancho"] += 1
        else:
            cabecera.append({"mes": etiqueta, "ancho": 1})

    return render(request, "otec/carta_gantt.html", {
        "filas": filas,
        "fechas": fechas,
        "cabecera": cabecera,
        "paralelo": paralelo,
        "max_paralelo": max_paralelo,
        # El desplegable va alfabético aunque la grilla vaya cronológica: en la
        # lista se busca por nombre, en la grilla se lee por fecha.
        "cursos": sorted(
            (
                {
                    "actividad": c["actividad"],
                    "activo": c is curso_activo,
                }
                for c in cursos
            ),
            key=lambda c: c["actividad"].nombre,
        ),
        "curso_activo": curso_activo["actividad"] if curso_activo else None,
        "meses": [
            {"valor": f"{a}-{m}", "etiqueta": f"{MESES_ES[m]} {a}", "activo": (a, m) == mes_activo}
            for a, m in meses
        ],
        "hay_filtro": bool(curso_activo or mes_activo),
        "n_estimadas": sum(1 for f in filas if f["estimada"]),
    })


def _mes_pedido(crudo):
    """(año, mes) de un «2026-08», o None si no viene o viene mal."""
    try:
        anio, mes = (int(p) for p in (crudo or "").split("-"))
    except ValueError:
        return None
    return (anio, mes) if 1 <= mes <= 12 else None


def _fecha_pedida(crudo):
    try:
        return datetime.strptime(crudo or "", "%Y-%m-%d").date()
    except ValueError:
        return None


@login_required
def calendario_curso(request, pk):
    """Calendario del curso: acá se colocan las clases, una por una.

    Las fechas las pone el relator y casi nunca caen todas el mismo día de la
    semana, así que no hay regla que declarar: se hace clic en el día y se
    carga la clase con su hora.
    """
    actividad = get_object_or_404(
        Actividad.objects.select_related("propuesta__institucion"), pk=pk
    )
    volver = reverse("otec:detalle_actividad", args=[actividad.pk])

    editando = None
    if request.GET.get("sesion", "").isdigit():
        editando = actividad.sesiones.filter(pk=request.GET["sesion"]).first()

    if request.method == "POST":
        instancia = None
        if request.POST.get("sesion", "").isdigit():
            instancia = actividad.sesiones.filter(pk=request.POST["sesion"]).first()
        form = SesionClaseForm(request.POST, instance=instancia, actividad=actividad)
        if form.is_valid():
            sesion = form.save()
            messages.success(
                request,
                f"Clase del {sesion.fecha:%d-%m-%Y} a las "
                f"{sesion.hora_inicio:%H:%M} guardada.",
            )
            return redirect(
                f"{reverse('otec:calendario_curso', args=[actividad.pk])}"
                f"?mes={sesion.fecha.year}-{sesion.fecha.month}"
            )
        mes_activo = _mes_pedido(request.POST.get("mes"))
    else:
        form = SesionClaseForm(
            instance=editando,
            initial=None if editando else _sugerencia(actividad, request),
            actividad=actividad,
        )
        mes_activo = _mes_pedido(request.GET.get("mes"))

    # Mes por defecto: el de la clase que se está editando, si no el del inicio
    # del curso, si no el de la primera clase cargada.
    sesiones = list(actividad.sesiones.select_related("sala"))
    if not mes_activo:
        referencia = (
            (editando.fecha if editando else None)
            or actividad.fecha_inicio
            or (sesiones[0].fecha if sesiones else None)
            or date.today()
        )
        mes_activo = (referencia.year, referencia.month)
    anio, mes = mes_activo

    por_fecha = {}
    for sesion in sesiones:
        por_fecha.setdefault(sesion.fecha, []).append(sesion)

    feriados = dict(Feriado.objects.values_list("fecha", "nombre"))
    hoy = date.today()
    semanas = [
        [
            {
                "fecha": dia,
                "del_mes": dia.month == mes,
                "sesiones": por_fecha.get(dia, []),
                "feriado": feriados.get(dia),
                "hoy": dia == hoy,
                "en_curso": _dentro_del_curso(actividad, dia),
            }
            for dia in semana
        ]
        for semana in Calendar(firstweekday=0).monthdatescalendar(anio, mes)
    ]

    # Choques de sala con otros cursos. Al guardar se bloquean, pero los datos
    # que vinieron del Tablero pueden traerlos de antes.
    n_conflictos = 0
    for sesion in sesiones:
        sesion.choques_de_sala = sesion.choques() if sesion.sala_id else []
        n_conflictos += bool(sesion.choques_de_sala)

    anterior = date(anio, mes, 1) - timedelta(days=1)
    siguiente = date(anio, mes, monthrange(anio, mes)[1]) + timedelta(days=1)

    return render(request, "otec/calendario_curso.html", {
        "actividad": actividad,
        "propuesta": actividad.propuesta,
        "form": form,
        "editando": editando,
        "semanas": semanas,
        "sesiones": sesiones,
        "n_conflictos": n_conflictos,
        "mes_nombre": MESES_ES[mes],
        "anio": anio,
        "mes_actual": f"{anio}-{mes}",
        "mes_anterior": f"{anterior.year}-{anterior.month}",
        "mes_siguiente": f"{siguiente.year}-{siguiente.month}",
        "fecha_elegida": _fecha_pedida(request.GET.get("fecha")),
        "volver": volver,
    })


def _sugerencia(actividad, request):
    """Valores con que llega el formulario al hacer clic en un día.

    La duración y la sala se copian de la última clase cargada: cambian las
    fechas, no la forma del curso, así que repetirlas a mano sería trabajo
    inventado.
    """
    inicial = {}
    fecha = _fecha_pedida(request.GET.get("fecha"))
    if fecha:
        inicial["fecha"] = fecha

    ultima = actividad.sesiones.order_by("fecha", "hora_inicio").last()
    if ultima:
        inicial["hora_inicio"] = ultima.hora_inicio
        inicial["duracion_horas"] = ultima.duracion_horas
        inicial["sala"] = ultima.sala_id
        inicial["grupo"] = ultima.grupo
    return inicial


def _dentro_del_curso(actividad, dia):
    """¿El día cae entre el inicio y el término declarados del curso?"""
    if not actividad.fecha_inicio:
        return False
    return actividad.fecha_inicio <= dia <= (actividad.fecha_termino or actividad.fecha_inicio)


@login_required
@require_POST
def eliminar_sesion(request, pk):
    """Quita una clase del calendario del curso."""
    sesion = get_object_or_404(SesionClase.objects.select_related("actividad"), pk=pk)
    actividad_pk = sesion.actividad_id
    fecha = sesion.fecha
    sesion.delete()
    messages.success(request, f"Clase del {fecha:%d-%m-%Y} eliminada.")
    return redirect(
        f"{reverse('otec:calendario_curso', args=[actividad_pk])}"
        f"?mes={fecha.year}-{fecha.month}"
    )


@login_required
def calendario_zoom(request):
    """Grilla bloque × día de cada sala, para la semana elegida."""
    hoy = date.today()

    sesiones_con_sala = list(
        SesionClase.objects
        .select_related("sala", "actividad", "actividad__propuesta__institucion")
        .exclude(sala=None)
    )

    # Las fechas con uso salen de las reservas del Tablero y de las clases
    # cargadas en el calendario de cada curso.
    fechas_con_uso = sorted(
        set(ReservaZoom.objects.values_list("fecha", flat=True).distinct())
        | {s.fecha for s in sesiones_con_sala}
    )
    if not fechas_con_uso:
        return render(request, "otec/calendario_zoom.html", {"sin_datos": True})

    # Semana por defecto: la primera con reservas a partir de hoy, o la última.
    referencia = next((f for f in fechas_con_uso if f >= hoy), fechas_con_uso[-1])
    if request.GET.get("semana"):
        try:
            referencia = datetime.strptime(request.GET["semana"], "%Y-%m-%d").date()
        except ValueError:
            pass

    lunes = referencia - timedelta(days=referencia.weekday())
    dias_semana = [lunes + timedelta(days=i) for i in range(5)]  # L a V

    feriados = dict(
        Feriado.objects.filter(fecha__in=dias_semana).values_list("fecha", "nombre")
    )

    reservas = list(
        ReservaZoom.objects
        .filter(fecha__in=dias_semana)
        .select_related("sala", "actividad", "actividad__propuesta__institucion")
    )

    for r in reservas:
        r.derivada = False

    # Las clases cargadas en el calendario del curso se dibujan junto a las
    # reservas del Tablero.
    #
    # Donde el Tablero ya trajo la sesión, esa manda — y la comparación es por
    # (curso, fecha, hora), no por sala: varios cursos cambian de sala a mitad
    # de camino, y comparar por sala duplicaría la clase en la sala equivocada.
    ya_en_tablero = {
        (r.actividad_id, r.fecha, r.hora_inicio) for r in reservas if r.actividad_id
    }
    ocupacion = {(r.sala_id, r.fecha, r.hora_inicio): r for r in reservas}

    choques = []
    for sesion in sesiones_con_sala:
        if sesion.fecha not in dias_semana:
            continue
        if (sesion.actividad_id, sesion.fecha, sesion.hora_inicio) in ya_en_tablero:
            continue

        llave = (sesion.sala_id, sesion.fecha, sesion.hora_inicio)
        previa = ocupacion.get(llave)
        if previa is not None:
            if previa.actividad_id != sesion.actividad_id:
                choques.append({
                    "fecha": sesion.fecha,
                    "sala": sesion.sala,
                    "hora": sesion.hora_inicio,
                    "actividades": [previa.actividad, sesion.actividad],
                })
            continue

        derivada = ReservaZoom(
            sala=sesion.sala,
            fecha=sesion.fecha,
            hora_inicio=sesion.hora_inicio,
            hora_fin=sesion.hora_fin,
            actividad=sesion.actividad,
            etiqueta=sesion.actividad.nombre,
        )
        derivada.derivada = True
        reservas.append(derivada)
        ocupacion[llave] = derivada

    # Color estable por actividad, para reconocerla de un vistazo en la grilla.
    actividades_semana = sorted(
        {r.actividad_id for r in reservas if r.actividad_id},
        key=lambda x: (x is None, x),
    )
    tono = {pk: i % 8 for i, pk in enumerate(actividades_semana)}

    salas = []
    for sala in SalaZoom.objects.filter(activa=True):
        de_la_sala = [r for r in reservas if r.sala_id == sala.pk]
        sin_bloque = [r for r in de_la_sala if r.hora_inicio is None]

        cuerpo = []
        for inicio, fin in BLOQUES:
            celdas = []
            for dia in dias_semana:
                reserva = next(
                    (
                        r for r in de_la_sala
                        if r.fecha == dia
                        and r.hora_inicio is not None
                        and r.hora_inicio <= inicio < r.hora_fin
                    ),
                    None,
                )
                celdas.append({
                    "dia": dia,
                    "reserva": reserva,
                    "inicia": reserva is not None and reserva.hora_inicio == inicio,
                    "tono": tono.get(reserva.actividad_id) if reserva else None,
                    "derivada": bool(reserva and getattr(reserva, "derivada", False)),
                    "feriado": dia in feriados,
                })
            cuerpo.append({"inicio": inicio, "fin": fin, "celdas": celdas})

        salas.append({
            "sala": sala,
            "cuerpo": cuerpo,
            "sin_bloque": sin_bloque,
            "ocupados": sum(1 for r in de_la_sala if r.hora_inicio),
            "horas": sum((r.duracion_horas or 0) for r in de_la_sala),
        })

    # Días con horas asincrónicas en la semana.
    asincronicas = (
        DiaActividad.objects
        .filter(fecha__in=dias_semana, horas_asincronicas__isnull=False)
        .select_related("actividad")
        .order_by("fecha")
    )

    return render(request, "otec/calendario_zoom.html", {
        "salas": salas,
        "dias_semana": dias_semana,
        "feriados": feriados,
        "asincronicas": asincronicas,
        "choques": choques,
        "n_derivadas": sum(1 for r in reservas if getattr(r, "derivada", False)),
        "semana_anterior": (lunes - timedelta(days=7)).isoformat(),
        "semana_siguiente": (lunes + timedelta(days=7)).isoformat(),
        "lunes": lunes,
        "viernes": dias_semana[-1],
    })


SUBCARPETA_CARGAS = "otec/cargas"


def _ruta_carga(nombre):
    """Ruta absoluta de un archivo subido, validando que no escape la carpeta."""
    base = Path(settings.MEDIA_ROOT) / SUBCARPETA_CARGAS
    destino = (base / nombre).resolve()
    if base.resolve() not in destino.parents:
        raise Http404("Archivo no válido.")
    return destino


@login_required
def importar_tablero(request):
    """Carga del Excel con vista previa antes de guardar.

    El archivo no necesita ninguna columna extra: la identidad se deduce del
    código de propuesta, del nombre del curso dentro de ella y del nombre de la
    institución, así que subir el mismo archivo dos veces no duplica nada.
    """
    form = SubirTableroForm()
    contexto = {"form": form}

    if request.method != "POST":
        return render(request, "otec/importar.html", contexto)

    # --- Paso 2: confirmar una previsualización ya hecha ---
    if request.POST.get("confirmar"):
        nombre = request.session.get("otec_carga")
        if not nombre:
            messages.error(request, "La previsualización expiró. Suba el archivo otra vez.")
            return redirect("otec:importar")

        ruta = _ruta_carga(nombre)

        if request.session.get("otec_carga_tipo") == "flujo":
            salida = _ejecutar_flujo(ruta, aplicar=True)
            request.session.pop("otec_carga", None)
            request.session.pop("otec_carga_tipo", None)
            ruta.unlink(missing_ok=True)
            messages.success(request, "Flujo de caja actualizado.")
            return render(request, "otec/importar.html", {
                "form": SubirTableroForm(),
                "salida_flujo": salida,
                "aplicado": True,
            })

        opciones = request.session.get("otec_carga_opciones", {})
        try:
            resultado = ImportadorTablero(ruta, **opciones).ejecutar(aplicar=True)
        except ErrorImportacion as exc:
            messages.error(request, str(exc))
            return redirect("otec:importar")

        calendario = _importar_calendario_si_corresponde(ruta)

        request.session.pop("otec_carga", None)
        request.session.pop("otec_carga_opciones", None)
        ruta.unlink(missing_ok=True)

        if resultado.hay_cambios:
            messages.success(request, f"Se aplicaron {len(resultado.relevantes)} cambios.")
        else:
            messages.info(request, "El archivo ya estaba reflejado: no hubo cambios.")

        return render(request, "otec/importar.html", {
            "form": SubirTableroForm(),
            "resultado": resultado,
            "calendario": calendario,
            "aplicado": True,
        })

    # --- Paso 1: subir y previsualizar ---
    form = SubirTableroForm(request.POST, request.FILES)
    if not form.is_valid():
        return render(request, "otec/importar.html", {"form": form})

    carpeta = Path(settings.MEDIA_ROOT) / SUBCARPETA_CARGAS
    carpeta.mkdir(parents=True, exist_ok=True)

    nombre = f"{uuid4().hex}.xlsx"
    ruta = carpeta / nombre
    with ruta.open("wb") as destino:
        for trozo in form.cleaned_data["archivo"].chunks():
            destino.write(trozo)

    # El archivo dice qué es por sus hojas: no hay que elegir el tipo a mano.
    if _es_archivo_de_flujo(ruta):
        salida = _ejecutar_flujo(ruta, aplicar=False)
        request.session["otec_carga"] = nombre
        request.session["otec_carga_tipo"] = "flujo"
        return render(request, "otec/importar.html", {
            "form": form,
            "salida_flujo": salida,
            "previsualizacion": True,
            "es_flujo": True,
            "archivo_original": form.cleaned_data["archivo"].name,
        })

    request.session["otec_carga_tipo"] = "tablero"
    opciones = {
        "separar_conflictos": form.cleaned_data["separar_conflictos"],
        "sobrescribir_ediciones": form.cleaned_data["sobrescribir_ediciones"],
    }
    try:
        resultado = ImportadorTablero(ruta, **opciones).ejecutar(aplicar=False)
    except ErrorImportacion as exc:
        ruta.unlink(missing_ok=True)
        messages.error(request, str(exc))
        return render(request, "otec/importar.html", {"form": form})

    request.session["otec_carga"] = nombre
    request.session["otec_carga_opciones"] = opciones

    return render(request, "otec/importar.html", {
        "form": form,
        "resultado": resultado,
        "previsualizacion": True,
        "archivo_original": form.cleaned_data["archivo"].name,
    })


HOJAS_FLUJO = {"Actividades", "Costos directos", "Parámetros"}


def _hojas_de(ruta):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(ruta, data_only=True, read_only=True)
        hojas = set(wb.sheetnames)
        wb.close()
        return hojas
    except Exception:
        return set()


def _es_archivo_de_flujo(ruta):
    hojas = _hojas_de(ruta)
    return HOJAS_FLUJO <= hojas and "Registro Actividades" not in hojas


def _ejecutar_flujo(ruta, aplicar):
    salida = StringIO()
    try:
        call_command(
            "otec_importar_flujo", archivo=str(ruta),
            dry_run=not aplicar, stdout=salida, stderr=salida,
        )
    except CommandError as exc:
        return f"No se pudo importar el flujo de caja: {exc}"
    return salida.getvalue().strip()


def _importar_calendario_si_corresponde(ruta):
    """Importa Gantt y Zoom si el archivo trae esas hojas."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(ruta, data_only=True, read_only=True)
        hojas = set(wb.sheetnames)
        wb.close()
    except Exception:
        return None

    if not {"Carta Gantt", "Zoom"} <= hojas:
        return None

    salida = StringIO()
    try:
        call_command("otec_importar_calendario", archivo=str(ruta), stdout=salida)
    except CommandError as exc:
        return f"No se pudo importar el calendario: {exc}"
    return salida.getvalue().strip()


# =========================
# ALTA, EDICIÓN Y BAJA
# =========================

@login_required
def crear_propuesta(request):
    form = PropuestaForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        propuesta = form.save(commit=False)
        propuesta.origen = Origen.MANUAL
        propuesta.save()
        messages.success(request, f"Propuesta «{propuesta.codigo}» creada.")
        return redirect("otec:detalle_propuesta", pk=propuesta.pk)

    return render(request, "otec/propuesta_form.html", {
        "form": form,
        "titulo": "Nueva propuesta",
        "volver": reverse("otec:lista_propuestas"),
    })


@login_required
def editar_propuesta(request, pk):
    propuesta = get_object_or_404(Propuesta, pk=pk)
    form = PropuestaForm(request.POST or None, instance=propuesta)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, f"Propuesta «{propuesta.codigo}» actualizada.")
        return redirect("otec:detalle_propuesta", pk=propuesta.pk)

    return render(request, "otec/propuesta_form.html", {
        "form": form,
        "propuesta": propuesta,
        "titulo": f"Editar {propuesta.codigo}",
        "volver": reverse("otec:detalle_propuesta", args=[propuesta.pk]),
    })


@login_required
def eliminar_propuesta(request, pk):
    propuesta = get_object_or_404(
        Propuesta.objects.select_related("institucion"), pk=pk
    )
    actividades = list(propuesta.actividades.all())

    if request.method == "POST":
        codigo = propuesta.codigo
        propuesta.delete()
        messages.success(
            request,
            f"Propuesta «{codigo}» eliminada junto con {len(actividades)} "
            f"actividad{'es' if len(actividades) != 1 else ''}.",
        )
        return redirect("otec:lista_propuestas")

    return render(request, "otec/confirmar_eliminar.html", {
        "objeto": propuesta,
        "titulo": f"Eliminar la propuesta {propuesta.codigo}",
        "descripcion": propuesta.institucion.nombre,
        "arrastra": [
            f"{len(actividades)} actividad{'es' if len(actividades) != 1 else ''}"
            f" con su checklist y sus días de calendario",
        ] if actividades else [],
        "detalle": [a.nombre for a in actividades],
        "nota": (
            "Las líneas del flujo de caja que apuntaban a estas actividades no "
            "se borran: quedan sin curso asociado."
            if LineaFinanciera.objects.filter(actividad__in=actividades).exists()
            else ""
        ),
        "aviso_importacion": (
            "Esta propuesta vino del Excel: si vuelve a subir el archivo, se "
            "volverá a crear."
            if propuesta.origen == Origen.IMPORTADO else ""
        ),
        "volver": reverse("otec:detalle_propuesta", args=[propuesta.pk]),
    })


def _guardar_costos(actividad, costos_form, extras):
    """Guarda el desglose y lo marca solo si alguien lo tocó de verdad.

    La marca es la que impide que la próxima importación vuelque encima los dos
    totales de la planilla y borre el detalle abierto acá. Por eso no se pone
    al guardar cualquier cambio de la actividad: quien corrigió el nombre del
    curso no quiso con eso desconectar los costos del Excel.
    """
    extras.instance = actividad
    tocado = costos_form.has_changed() or extras.has_changed()

    costos = costos_form.save(commit=False)
    if costos.pk or tocado:
        costos.actividad = actividad
        if tocado:
            costos.editado_en_sistema = True
        costos.save()

    extras.save()


@login_required
def crear_actividad(request, propuesta_pk=None):
    propuesta = (
        get_object_or_404(Propuesta, pk=propuesta_pk) if propuesta_pk else None
    )
    form = ActividadForm(request.POST or None, propuesta=propuesta)
    costos = CostoActividadForm(request.POST or None)
    extras = GastoExtraFormSet(request.POST or None, prefix="extras")

    formularios = [form, costos, extras]
    # La lista se evalúa entera a propósito: con un generador, el primer
    # formulario inválido cortaría la validación y los demás volverían a la
    # pantalla sin sus errores marcados.
    if request.method == "POST" and all([f.is_valid() for f in formularios]):
        actividad = form.save(commit=False)
        if propuesta is not None:
            actividad.propuesta = propuesta
        actividad.origen = Origen.MANUAL
        actividad.save()
        form.save_m2m()
        _guardar_costos(actividad, costos, extras)
        creados = actividad.sincronizar_checklist()
        messages.success(
            request,
            f"Actividad «{actividad.nombre}» creada con {creados} ítems de checklist.",
        )
        return redirect("otec:detalle_actividad", pk=actividad.pk)

    return render(request, "otec/actividad_form.html", {
        "form": form,
        "costos": costos,
        "extras": extras,
        "propuesta": propuesta,
        # Al crear desde una propuesta el campo va oculto; al editar se puede
        # mover el curso a otra propuesta, así que se muestra.
        "propuesta_fija": propuesta is not None,
        "titulo": "Nueva actividad",
        "volver": (
            reverse("otec:detalle_propuesta", args=[propuesta.pk]) if propuesta
            else reverse("otec:lista_actividades")
        ),
    })


@login_required
def editar_actividad(request, pk):
    actividad = get_object_or_404(
        Actividad.objects.select_related("propuesta", "costos"), pk=pk
    )
    form = ActividadForm(request.POST or None, instance=actividad)
    costos = CostoActividadForm(
        request.POST or None, instance=getattr(actividad, "costos", None)
    )
    extras = GastoExtraFormSet(
        request.POST or None, instance=actividad, prefix="extras"
    )

    formularios = [form, costos, extras]
    # La lista se evalúa entera a propósito: con un generador, el primer
    # formulario inválido cortaría la validación y los demás volverían a la
    # pantalla sin sus errores marcados.
    if request.method == "POST" and all([f.is_valid() for f in formularios]):
        form.save()
        _guardar_costos(actividad, costos, extras)
        actividad.sincronizar_checklist()
        messages.success(request, f"Actividad «{actividad.nombre}» actualizada.")
        return redirect("otec:detalle_actividad", pk=actividad.pk)

    return render(request, "otec/actividad_form.html", {
        "form": form,
        "costos": costos,
        "extras": extras,
        "actividad": actividad,
        "propuesta": actividad.propuesta,
        "propuesta_fija": False,
        "titulo": f"Editar {actividad.nombre[:40]}",
        "volver": reverse("otec:detalle_actividad", args=[actividad.pk]),
    })


@login_required
def eliminar_actividad(request, pk):
    actividad = get_object_or_404(
        Actividad.objects.select_related("propuesta"), pk=pk
    )
    propuesta_pk = actividad.propuesta_id

    if request.method == "POST":
        nombre = actividad.nombre
        actividad.delete()
        messages.success(request, f"Actividad «{nombre}» eliminada.")
        return redirect("otec:detalle_propuesta", pk=propuesta_pk)

    n_items = actividad.items.count()
    n_dias = actividad.dias.count()
    n_reservas = actividad.reservas_zoom.count()
    arrastra = []
    if n_items:
        arrastra.append(f"{n_items} ítems de checklist")
    if n_dias:
        arrastra.append(f"{n_dias} días de la carta Gantt")
    if n_reservas:
        arrastra.append(f"{n_reservas} reservas de sala Zoom")

    return render(request, "otec/confirmar_eliminar.html", {
        "objeto": actividad,
        "titulo": "Eliminar la actividad",
        "descripcion": f"{actividad.propuesta.codigo} · {actividad.nombre}",
        "arrastra": arrastra,
        "detalle": [],
        "nota": (
            "Las líneas del flujo de caja que apuntaban a esta actividad no se "
            "borran: quedan sin curso asociado."
            if actividad.lineas_financieras.exists() else ""
        ),
        "aviso_importacion": (
            "Esta actividad vino del Excel: si vuelve a subir el archivo, se "
            "volverá a crear."
            if actividad.origen == Origen.IMPORTADO else ""
        ),
        "volver": reverse("otec:detalle_actividad", args=[actividad.pk]),
    })


@login_required
def contactos_de_institucion(request):
    """Repuebla el desplegable de contactos al cambiar de institución (HTMX)."""
    institucion = request.GET.get("institucion", "")
    contactos = (
        Contacto.objects.filter(institucion_id=institucion)
        if institucion.isdigit() else Contacto.objects.none()
    )
    return render(request, "otec/partials/contactos_options.html", {
        "contactos": contactos,
        "seleccionado": request.GET.get("contacto", ""),
    })


@login_required
def lista_propuestas(request):
    propuestas = (
        Propuesta.objects
        .select_related("institucion", "contacto")
        .prefetch_related("actividades")
        .annotate(
            n_actividades=Count("actividades"),
            total_ofertado=Sum("actividades__valor_ofertado"),
            total_adjudicado=Sum("actividades__monto_adjudicado"),
        )
    )
    estado = request.GET.get("estado_comercial", "")
    if estado:
        propuestas = propuestas.filter(estado_comercial=estado)

    return render(request, "otec/lista_propuestas.html", {
        "propuestas": propuestas,
        "estados_comerciales": Propuesta.EstadoComercial.choices,
        "filtros": {"estado_comercial": estado},
    })


@login_required
def detalle_propuesta(request, pk):
    propuesta = get_object_or_404(
        Propuesta.objects.select_related("institucion", "contacto"), pk=pk
    )
    actividades = list(
        _actividades_base().filter(propuesta=propuesta)
    )
    return render(request, "otec/detalle_propuesta.html", {
        "propuesta": propuesta,
        "actividades": actividades,
    })
