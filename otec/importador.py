"""Importación del Tablero Maestro OTEC sin duplicar lo que ya está cargado.

La planilla no trae identificadores propios, así que la identidad se deduce de
lo que ya contiene:

* la **propuesta** por su ``ID Propuesta``;
* la **actividad** por (propuesta, nombre), tolerando cambios de redacción;
* la **institución** y el **relator/a** por nombre normalizado (sin tildes,
  mayúsculas ni espacios de más).

Con eso, volver a subir el mismo archivo actualiza lo que cambió y deja intacto
lo demás — no hay que agregarle ninguna columna al Excel.

El importador corre siempre dentro de una transacción y registra cada cambio,
de modo que la misma pasada sirve para previsualizar (``aplicar=False``, se
revierte al final) o para guardar.
"""

import re
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher

from django.db import transaction

from .checklist import COLUMNAS_EXCEL, sincronizar_catalogo
from .models import (
    Actividad,
    Contacto,
    CostoActividad,
    Estado,
    Institucion,
    ItemChecklist,
    MetaAnual,
    Origen,
    PlantillaItem,
    Propuesta,
    Relator,
    equipo_otec,
)

HOJA = "Registro Actividades"

# Columnas que describen el expediente y, por lo tanto, deberían tener el mismo
# valor en todas las filas de una misma propuesta. Si no lo tienen, el ID está
# cubriendo más de un expediente y el importador lo reporta.
COLUMNAS_PROPUESTA = [
    "Institución Cliente",
    "Contacto",
    "Canal",
    "Tipo Institución",
    "Estado Comercial",
    "Fecha envío propuesta",
    "Estado Convenio/ Tramitación",
    "Estado Decretación",
    "Memo Decretación",
    "Fecha Memo",
    "CR",
    "N° Decreto",
    "Fecha Resolución",
]

# Valores de la planilla que significan "sin dato" en un campo de texto.
# "no aplica" entra acá: si no, el importador creaba un relator llamado
# "No aplica" que después aparecía en la carga de trabajo como si fuera una
# persona. Que no aplique ya lo dice ``tipo_relator``.
VACIOS = {
    "", "-", "n/a", "na", "por definir", "no definido",
    "sin información", "no aplica",
}

# Similitud mínima para considerar que una fila es el mismo curso con el nombre
# corregido, en vez de un curso nuevo.
UMBRAL_RENOMBRE = 0.80

# Columnas de monto que en la planilla son fórmulas. Si el archivo se guarda con
# una herramienta que no recalcula (o que descarta el valor cacheado), llegan
# vacías y escribirlas pondría todo en cero. Cuando una de estas columnas viene
# vacía en TODAS las filas, se ignora y se avisa, en vez de borrar los montos.
COLUMNAS_MONTO = {
    "Valor Ofertado": "valor_ofertado",
    "Costo Relatoría": "costo_relatoria",
    "Otros gastos": "otros_gastos",
    "Monto Adjudicado": "monto_adjudicado",
    "Monto Facturado": "monto_facturado",
    "Monto Pagado": "monto_pagado",
}


# =========================
# NORMALIZACIÓN
# =========================

def texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def normaliza(valor):
    return texto(valor).casefold()


def clave(valor):
    """Clave de comparación: sin tildes, sin puntuación, sin espacios de más.

    Hace que "Instituto Nacional de Deportes" y "INSTITUTO NACIONAL DE
    DEPORTES " se reconozcan como la misma institución.
    """
    s = unicodedata.normalize("NFKD", texto(valor).casefold())
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^0-9a-z ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def a_fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    crudo = texto(valor)
    if not crudo:
        return None
    for formato in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(crudo, formato).date()
        except ValueError:
            continue
    return None  # fechas mal tipeadas (p. ej. "22-07-206") se descartan


def a_decimal(valor):
    if valor is None or valor == "":
        return Decimal("0")
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    crudo = texto(valor).replace("$", "").replace(".", "").replace(",", ".").strip()
    try:
        return Decimal(crudo or "0")
    except InvalidOperation:
        return Decimal("0")


def a_entero(valor):
    if valor in (None, ""):
        return None
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return None


def mapea(valor, tabla, defecto):
    """Busca ``valor`` en ``tabla`` (claves normalizadas); si no está, ``defecto``."""
    return tabla.get(normaliza(valor), defecto)


# =========================
# TABLAS DE EQUIVALENCIA
# =========================

TIPOS_INSTITUCION = {
    "pública": Institucion.Tipo.PUBLICA,
    "publica": Institucion.Tipo.PUBLICA,
    "privada": Institucion.Tipo.PRIVADA,
    "interna": Institucion.Tipo.INTERNA,
    "personas naturales": Institucion.Tipo.PERSONA,
}

CANALES = {
    "convenio / trato directo": Propuesta.Canal.CONVENIO,
    "convenio/trato directo": Propuesta.Canal.CONVENIO,
    "licitación": Propuesta.Canal.LICITACION,
    "licitacion": Propuesta.Canal.LICITACION,
    "venta directa/ matrícula": Propuesta.Canal.VENTA_DIRECTA,
    "venta directa / matrícula": Propuesta.Canal.VENTA_DIRECTA,
    "gestión interna/ osl": Propuesta.Canal.INTERNA,
    "gestión interna": Propuesta.Canal.INTERNA,
}

PRIORIDADES = {
    "alta": Actividad.Prioridad.ALTA,
    "media": Actividad.Prioridad.MEDIA,
    "baja": Actividad.Prioridad.BAJA,
}

ESTADOS_COMERCIALES = {
    "ganada": Propuesta.EstadoComercial.GANADA,
    "en revisión": Propuesta.EstadoComercial.EN_REVISION,
    "en revision": Propuesta.EstadoComercial.EN_REVISION,
    "en preparación": Propuesta.EstadoComercial.EN_PREPARACION,
    "en preparacion": Propuesta.EstadoComercial.EN_PREPARACION,
    "enviada": Propuesta.EstadoComercial.ENVIADA,
    "perdida": Propuesta.EstadoComercial.PERDIDA,
    "desistida": Propuesta.EstadoComercial.DESISTIDA,
}

ESTADOS_CONVENIO = {
    "formalizado": Propuesta.EstadoConvenio.FORMALIZADO,
    "en revisión": Propuesta.EstadoConvenio.EN_REVISION,
    "en revision": Propuesta.EstadoConvenio.EN_REVISION,
    "no iniciado": Propuesta.EstadoConvenio.NO_INICIADO,
}

ESTADOS_DECRETACION = {
    "listo": Propuesta.EstadoDecretacion.LISTO,
    "pendiente recepción convenio firmado": Propuesta.EstadoDecretacion.PENDIENTE_CONVENIO,
    "pendiente decretación interna": Propuesta.EstadoDecretacion.PENDIENTE_INTERNA,
    "no iniciado": Propuesta.EstadoDecretacion.NO_INICIADO,
}

MODALIDADES = {
    "e-learning": Actividad.Modalidad.ELEARNING,
    "e-learning asincrónico": Actividad.Modalidad.ELEARNING_ASINC,
    "e-learning asincrónica": Actividad.Modalidad.ELEARNING_ASINC,
    "b-learning": Actividad.Modalidad.BLEARNING,
    "presencial": Actividad.Modalidad.PRESENCIAL,
}

TIPOS_RELATOR = {
    "interno": Actividad.TipoRelator.INTERNO,
    "externo": Actividad.TipoRelator.EXTERNO,
    "no definido": Actividad.TipoRelator.NO_DEFINIDO,
    "no aplica": Actividad.TipoRelator.NO_APLICA,
}

ESTADOS_EJECUCION = {
    "programada": Actividad.EstadoEjecucion.PROGRAMADA,
    "no programada": Actividad.EstadoEjecucion.NO_PROGRAMADA,
    "en ejecución": Actividad.EstadoEjecucion.EN_EJECUCION,
    "en ejecución / seguimiento": Actividad.EstadoEjecucion.EN_EJECUCION,
    "ejecutada": Actividad.EstadoEjecucion.EJECUTADA,
    "suspendida": Actividad.EstadoEjecucion.SUSPENDIDA,
}

ESTADOS_ITEM = {
    "sí": Estado.SI,
    "si": Estado.SI,
    "listo": Estado.SI,
    "ok": Estado.SI,
    "no": Estado.NO,
    "no aplica": Estado.NO_APLICA,
}


def estado_item(valor):
    """Traduce el texto libre de la planilla a (estado, detalle)."""
    crudo = texto(valor)
    normalizado = crudo.casefold()
    if not crudo:
        return Estado.PENDIENTE, ""
    if normalizado in ESTADOS_ITEM:
        return ESTADOS_ITEM[normalizado], ""
    if normalizado.startswith("pendiente"):
        # "Pendiente actualización", "Pendiente recepción..." conservan el matiz.
        return Estado.PENDIENTE, "" if normalizado == "pendiente" else crudo
    # "En revisión", "En validación operativa", "En ajuste", "Solicitado", "En curso"...
    return Estado.EN_PROCESO, crudo


# =========================
# RESULTADO
# =========================

CREAR = "crear"
ACTUALIZAR = "actualizar"
RENOMBRAR = "renombrar"
ELIMINAR = "eliminar"
CONSERVAR = "conservado"
IGUAL = "sin cambios"


@dataclass
class Cambio:
    entidad: str
    accion: str
    nombre: str
    detalle: str = ""


@dataclass
class Resultado:
    aplicado: bool = False
    cambios: list = field(default_factory=list)
    avisos: list = field(default_factory=list)
    descartadas: list = field(default_factory=list)

    def registrar(self, entidad, accion, nombre, detalle=""):
        self.cambios.append(Cambio(entidad, accion, nombre, detalle))

    def contar(self, accion=None, entidad=None):
        return sum(
            1 for c in self.cambios
            if (accion is None or c.accion == accion)
            and (entidad is None or c.entidad == entidad)
        )

    @property
    def relevantes(self):
        """Cambios que vale la pena mostrar (todo menos lo que quedó igual)."""
        return [c for c in self.cambios if c.accion != IGUAL]

    @property
    def hay_cambios(self):
        return bool(self.relevantes)

    def resumen_por_entidad(self):
        """Una fila por tipo de registro. Las llaves evitan espacios para que
        la plantilla pueda leerlas directamente."""
        llaves = {
            CREAR: "nuevos",
            ACTUALIZAR: "actualizados",
            RENOMBRAR: "renombrados",
            ELIMINAR: "eliminados",
            CONSERVAR: "conservados",
            IGUAL: "iguales",
        }
        entidades = {}
        for c in self.cambios:
            fila = entidades.setdefault(c.entidad, {
                "entidad": c.entidad, "nuevos": 0, "actualizados": 0,
                "renombrados": 0, "eliminados": 0, "conservados": 0, "iguales": 0,
            })
            fila[llaves[c.accion]] += 1
        orden = [
            "Institución", "Contacto", "Relator/a", "Propuesta",
            "Actividad", "Responsables", "Checklist", "Meta anual",
        ]
        return sorted(
            entidades.values(),
            key=lambda f: orden.index(f["entidad"]) if f["entidad"] in orden else 99,
        )


class ErrorImportacion(Exception):
    pass


class _Rollback(Exception):
    """Revierte la transacción cuando solo se está previsualizando."""


# =========================
# IMPORTADOR
# =========================

def guardar(modelo, buscar, valores):
    """Crea o actualiza sin tocar lo que ya está igual.

    Devuelve (objeto, acción, campos_modificados).
    """
    obj = modelo.objects.filter(**buscar).first()
    if obj is None:
        obj = modelo.objects.create(**buscar, **valores)
        return obj, CREAR, []

    modificados = []
    for campo, nuevo in valores.items():
        actual = getattr(obj, campo)
        if hasattr(actual, "pk") or hasattr(nuevo, "pk"):
            iguales = getattr(actual, "pk", None) == getattr(nuevo, "pk", None)
        else:
            iguales = actual == nuevo
        if not iguales:
            modificados.append(campo)
            setattr(obj, campo, nuevo)

    if modificados:
        obj.save(update_fields=modificados)
        return obj, ACTUALIZAR, modificados
    return obj, IGUAL, []


class ImportadorTablero:
    """Lee la hoja ``Registro Actividades`` y la refleja en la base de datos."""

    def __init__(self, ruta, separar_conflictos=False, sobrescribir_ediciones=False):
        self.ruta = ruta
        self.separar_conflictos = separar_conflictos
        # Por defecto la planilla NO pisa lo que se editó desde la aplicación:
        # quien marca un ítem en pantalla suele tener información más fresca.
        self.sobrescribir_ediciones = sobrescribir_ediciones
        self.resultado = Resultado()

    # -- utilidades de lectura --

    def _leer(self):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise ErrorImportacion("Falta openpyxl (pip install openpyxl).") from exc

        try:
            wb = load_workbook(self.ruta, data_only=True, read_only=True)
        except FileNotFoundError as exc:
            raise ErrorImportacion(f"No se encontró el archivo: {self.ruta}") from exc
        except Exception as exc:
            raise ErrorImportacion(
                "No se pudo leer el archivo. ¿Es un .xlsx válido?"
            ) from exc

        if HOJA not in wb.sheetnames:
            wb.close()
            raise ErrorImportacion(
                f"El archivo no tiene la hoja '{HOJA}'. "
                f"Hojas encontradas: {', '.join(wb.sheetnames)}."
            )

        filas = list(wb[HOJA].iter_rows(values_only=True))
        self.filas_dashboard = (
            list(wb["Dashboard"].iter_rows(values_only=True))
            if "Dashboard" in wb.sheetnames else []
        )
        wb.close()
        if not filas:
            raise ErrorImportacion(f"La hoja '{HOJA}' está vacía.")
        return filas

    def _meta_anual(self):
        """Lee la meta financiera del Dashboard, si la hoja viene en el archivo.

        Se busca por la etiqueta y se toma el número de la celda de abajo, para
        no depender de una coordenada fija.
        """
        for i, fila in enumerate(self.filas_dashboard):
            if not fila:
                continue
            for j, celda in enumerate(fila):
                etiqueta = texto(celda)
                if not etiqueta.lower().startswith("meta financiera"):
                    continue
                anio = next(
                    (int(t) for t in re.findall(r"\b(20\d{2})\b", etiqueta)), None
                )
                siguiente = self.filas_dashboard[i + 1] if i + 1 < len(self.filas_dashboard) else None
                valor = siguiente[j] if siguiente and j < len(siguiente) else None
                if anio and isinstance(valor, (int, float)) and valor > 0:
                    meta, accion, _ = guardar(
                        MetaAnual, {"anio": anio}, {"monto": a_decimal(valor)}
                    )
                    if accion != IGUAL:
                        self.resultado.registrar(
                            "Meta anual", accion, f"{anio}: ${meta.monto:,.0f}"
                        )
                    return meta
        return None

    # -- ejecución --

    def ejecutar(self, aplicar=False):
        filas = self._leer()

        encabezados = [texto(h) for h in filas[0]]
        self.indice = {h: i for i, h in enumerate(encabezados) if h}

        faltantes = [c for c in COLUMNAS_EXCEL if c not in self.indice]
        if faltantes:
            self.resultado.avisos.append(
                f"{len(faltantes)} columnas de checklist no están en el archivo "
                f"y se omiten: {', '.join(faltantes[:6])}"
                + ("…" if len(faltantes) > 6 else "")
            )

        try:
            with transaction.atomic():
                self._procesar(filas[1:])
                if not aplicar:
                    raise _Rollback()
        except _Rollback:
            pass
        else:
            self.resultado.aplicado = True

        return self.resultado

    def _valor(self, fila, columna):
        i = self.indice.get(columna)
        if i is None or i >= len(fila):
            return None
        return fila[i]

    def _procesar(self, filas):
        sincronizar_catalogo()
        self.plantillas = {
            (p.etapa, p.nombre): p for p in PlantillaItem.objects.filter(activo=True)
        }
        # Cachés por clave normalizada, para no crear duplicados por tildes.
        self.cache_instituciones = {clave(i.nombre): i for i in Institucion.objects.all()}
        self.cache_relatores = {clave(r.nombre): r for r in Relator.objects.all()}

        # El equipo OTEC se indexa por nombre completo, nombre de pila y
        # usuario, porque la planilla los nombra de cualquiera de esas formas.
        self.usuarios_otec = {}
        self.nombres_sin_usuario = Counter()
        self.costos_respetados = 0
        for usuario in equipo_otec():
            for etiqueta in (
                usuario.get_full_name(), usuario.first_name, usuario.username
            ):
                if etiqueta:
                    self.usuarios_otec.setdefault(clave(etiqueta), usuario)

        self.campos_ignorados = self._detectar_formulas_sin_valor(filas)

        grupos = self._agrupar(filas)
        trabajo = self._resolver_conflictos(grupos)

        propuestas_tocadas = set()
        actividades_vistas = set()

        for codigo, entradas in trabajo:
            propuesta = self._propuesta(codigo, entradas[0][1])
            propuestas_tocadas.add(propuesta.pk)
            reclamadas = set()
            for _numero, fila in entradas:
                actividad = self._actividad(propuesta, fila, reclamadas)
                actividades_vistas.add(actividad.pk)
                reclamadas.add(actividad.pk)

        self._podar(propuestas_tocadas, actividades_vistas)
        self._meta_anual()

        if self.costos_respetados:
            self.resultado.avisos.append(
                f"{self.costos_respetados} actividad"
                f"{'es' if self.costos_respetados != 1 else ''} tienen los costos "
                f"desglosados en el sistema y se dejaron como estaban: la planilla "
                f"solo trae dos totales y volcarlos encima borraría el detalle. "
                f"Marque «sobrescribir ediciones» si quiere que mande el Excel."
            )

        if self.nombres_sin_usuario:
            detalle = ", ".join(
                f"{n} ({c})" for n, c in self.nombres_sin_usuario.most_common()
            )
            self.resultado.avisos.append(
                f"Nombres del campo de seguimiento sin usuario en los grupos de "
                f"OTEC: {detalle}. Créelos y agréguelos al grupo para que "
                f"aparezcan en la carga laboral."
            )

    def _detectar_formulas_sin_valor(self, filas):
        """Campos de monto que llegan vacíos en todo el archivo.

        Es la firma de un .xlsx guardado sin recalcular: la fórmula sigue ahí
        pero sin valor. Aplicarlos dejaría todos los montos en cero.
        """
        con_datos = [
            f for f in filas
            if f and texto(self._valor(f, "ID Propuesta"))
            and texto(self._valor(f, "Actividad/Curso"))
        ]
        if not con_datos:
            return set()

        ignorados = set()
        for columna, campo in COLUMNAS_MONTO.items():
            if columna not in self.indice:
                continue
            if all(self._valor(f, columna) in (None, "") for f in con_datos):
                ignorados.add(campo)
                self.resultado.avisos.append(
                    f"La columna «{columna}» llegó vacía en todas las filas. "
                    f"Suele pasar cuando el archivo se guardó sin recalcular las "
                    f"fórmulas: se ignora y se conservan los montos ya cargados."
                )
        return ignorados

    def _agrupar(self, filas):
        grupos = {}
        for numero, fila in enumerate(filas, start=2):
            codigo = texto(self._valor(fila, "ID Propuesta"))
            if not codigo:
                continue
            if not texto(self._valor(fila, "Institución Cliente")) or not texto(
                self._valor(fila, "Actividad/Curso")
            ):
                self.resultado.descartadas.append(
                    f"Fila {numero}: ID «{codigo}» sin institución o sin nombre de curso."
                )
                continue
            grupos.setdefault(codigo, []).append((numero, fila))
        return grupos

    def _resolver_conflictos(self, grupos):
        def huella(fila):
            return tuple(texto(self._valor(fila, c)) for c in COLUMNAS_PROPUESTA)

        trabajo = []
        for codigo, entradas in grupos.items():
            if len({huella(f) for _, f in entradas}) == 1:
                trabajo.append((codigo, entradas))
                continue

            campos = [
                c for c in COLUMNAS_PROPUESTA
                if len({texto(self._valor(f, c)) for _, f in entradas}) > 1
            ]
            if self.separar_conflictos:
                por_huella = {}
                for numero, fila in entradas:
                    por_huella.setdefault(huella(fila), []).append((numero, fila))
                self.resultado.avisos.append(
                    f"«{codigo}» agrupa {len(por_huella)} expedientes distintos "
                    f"(difieren en {', '.join(campos)}). Se separan con sufijo."
                )
                for n, (_, sub) in enumerate(sorted(por_huella.items()), start=1):
                    trabajo.append((codigo if n == 1 else f"{codigo}-{n}", sub))
            else:
                self.resultado.avisos.append(
                    f"«{codigo}»: sus filas difieren en {', '.join(campos)}. "
                    f"Se usa la primera fila; marque «separar expedientes» para dividirla."
                )
                trabajo.append((codigo, entradas))
        return trabajo

    # -- entidades --

    def _institucion(self, fila):
        nombre = texto(self._valor(fila, "Institución Cliente"))
        k = clave(nombre)
        existente = self.cache_instituciones.get(k)
        if existente:
            return existente

        institucion = Institucion.objects.create(
            nombre=nombre,
            tipo=mapea(
                self._valor(fila, "Tipo Institución"),
                TIPOS_INSTITUCION,
                Institucion.Tipo.PUBLICA,
            ),
        )
        self.cache_instituciones[k] = institucion
        self.resultado.registrar("Institución", CREAR, nombre)
        return institucion

    def _contacto(self, institucion, fila):
        nombre = texto(self._valor(fila, "Contacto"))
        if not nombre or nombre.casefold() in VACIOS:
            return None
        existente = next(
            (c for c in institucion.contactos.all() if clave(c.nombre) == clave(nombre)),
            None,
        )
        if existente:
            return existente
        contacto = Contacto.objects.create(institucion=institucion, nombre=nombre)
        self.resultado.registrar("Contacto", CREAR, nombre, institucion.nombre)
        return contacto

    def _relator(self, fila, tipo_relator):
        nombre = texto(self._valor(fila, "Relator Asignado"))
        if not nombre or nombre.casefold() in VACIOS:
            return None
        k = clave(nombre)
        existente = self.cache_relatores.get(k)
        if existente:
            return existente
        relator = Relator.objects.create(
            nombre=nombre,
            tipo=(
                Relator.Tipo.INTERNO
                if tipo_relator == Actividad.TipoRelator.INTERNO
                else Relator.Tipo.EXTERNO
            ),
        )
        self.cache_relatores[k] = relator
        self.resultado.registrar("Relator/a", CREAR, nombre)
        return relator

    def _propuesta(self, codigo, fila):
        institucion = self._institucion(fila)
        contacto = self._contacto(institucion, fila)

        crudo_convenio = texto(self._valor(fila, "Estado Convenio/ Tramitación"))
        estado_convenio = mapea(
            crudo_convenio,
            ESTADOS_CONVENIO,
            Propuesta.EstadoConvenio.EN_TRAMITE if crudo_convenio
            else Propuesta.EstadoConvenio.NO_INICIADO,
        )
        observacion = (
            crudo_convenio if normaliza(crudo_convenio) not in ESTADOS_CONVENIO else ""
        )

        propuesta, accion, campos = guardar(
            Propuesta,
            {"codigo": codigo},
            {
                "origen": Origen.IMPORTADO,
                "institucion": institucion,
                "contacto": contacto,
                "canal": mapea(self._valor(fila, "Canal"), CANALES, Propuesta.Canal.CONVENIO),
                "estado_comercial": mapea(
                    self._valor(fila, "Estado Comercial"),
                    ESTADOS_COMERCIALES,
                    Propuesta.EstadoComercial.EN_PREPARACION,
                ),
                "fecha_envio": a_fecha(self._valor(fila, "Fecha envío propuesta")),
                "anio": a_entero(self._valor(fila, "Año")) or date.today().year,
                "estado_convenio": estado_convenio,
                "observacion_convenio": observacion[:500],
                "estado_decretacion": mapea(
                    self._valor(fila, "Estado Decretación"),
                    ESTADOS_DECRETACION,
                    Propuesta.EstadoDecretacion.NO_INICIADO,
                ),
                "memo_decretacion": texto(self._valor(fila, "Memo Decretación"))[:50],
                "fecha_memo": a_fecha(self._valor(fila, "Fecha Memo")),
                "cr": texto(self._valor(fila, "CR"))[:30],
                "n_decreto": texto(self._valor(fila, "N° Decreto"))[:50],
                "fecha_resolucion": a_fecha(self._valor(fila, "Fecha Resolución")),
            },
        )
        self.resultado.registrar("Propuesta", accion, codigo, ", ".join(campos))
        return propuesta

    def _buscar_actividad(self, propuesta, nombre, reclamadas):
        """Ubica la actividad aunque le hayan corregido la redacción.

        Sin esto, cambiar "Portugués Básico." por "Portugués Básico" en el Excel
        crearía un curso nuevo y borraría el anterior con todo su checklist.
        """
        existentes = [a for a in propuesta.actividades.all() if a.pk not in reclamadas]

        k = clave(nombre)
        exacta = next((a for a in existentes if clave(a.nombre) == k), None)
        if exacta:
            return exacta, False

        mejor, puntaje = None, 0
        for a in existentes:
            s = SequenceMatcher(None, k, clave(a.nombre)).ratio()
            if s > puntaje:
                mejor, puntaje = a, s
        if mejor is not None and puntaje >= UMBRAL_RENOMBRE:
            return mejor, True
        return None, False

    def _actividad(self, propuesta, fila, reclamadas):
        nombre = texto(self._valor(fila, "Actividad/Curso"))
        tipo_relator = mapea(
            self._valor(fila, "Tipo Relator"),
            TIPOS_RELATOR,
            Actividad.TipoRelator.NO_DEFINIDO,
        )
        relator = self._relator(fila, tipo_relator)

        valores = {
            "nombre": nombre,
            # Si la actividad figura en el archivo, pasa a estar gobernada por
            # él, aunque se hubiera creado a mano.
            "origen": Origen.IMPORTADO,
            "modalidad": mapea(
                self._valor(fila, "Modalidad"), MODALIDADES, Actividad.Modalidad.ELEARNING
            ),
            "prioridad": mapea(
                self._valor(fila, "Prioridad"), PRIORIDADES, Actividad.Prioridad.MEDIA
            ),
            "n_participantes": a_entero(self._valor(fila, "N° Participantes")) or 0,
            "horas": a_entero(self._valor(fila, "Horas")) or 0,
            "tipo_relator": tipo_relator,
            "relator": relator,
            "fecha_confirmacion_relator": a_fecha(self._valor(fila, "Fecha confirmación relator/a")),
            "fecha_inicio": a_fecha(self._valor(fila, "Fecha Inicio curso")),
            "fecha_termino": a_fecha(self._valor(fila, "Fecha Término curso")),
            "proxima_fecha_critica": a_fecha(self._valor(fila, "Próxima fecha crítica")),
            "estado_ejecucion": mapea(
                self._valor(fila, "Estado Ejecución"),
                ESTADOS_EJECUCION,
                Actividad.EstadoEjecucion.NO_PROGRAMADA,
            ),
            "n_participantes_ejecucion": a_entero(self._valor(fila, "N° Participantes Ejecución")),
            "n_becados": a_entero(self._valor(fila, "N° Becados")),
            "n_aprobados": a_entero(self._valor(fila, "N° Aprobados")),
            "valor_ofertado": a_decimal(self._valor(fila, "Valor Ofertado")),
            "monto_adjudicado": a_decimal(self._valor(fila, "Monto Adjudicado")),
            "n_factura": texto(self._valor(fila, "N° Factura"))[:50],
            "fecha_factura": a_fecha(self._valor(fila, "Fecha Factura")),
            "monto_facturado": a_decimal(self._valor(fila, "Monto Facturado")),
            "fecha_pago": a_fecha(self._valor(fila, "Fecha Pago")),
            "monto_pagado": a_decimal(self._valor(fila, "Monto Pagado")),
            "observaciones": texto(self._valor(fila, "Observaciones / próximo paso")),
            "responsable_seguimiento": texto(self._valor(fila, "Responsable seguimiento"))[:255],
            "actualizado_al": a_fecha(self._valor(fila, "Actualizado al")),
        }

        for campo in self.campos_ignorados:
            valores.pop(campo, None)

        existente, es_renombre = self._buscar_actividad(propuesta, nombre, reclamadas)

        if existente is None:
            actividad = Actividad.objects.create(propuesta=propuesta, **valores)
            self.resultado.registrar("Actividad", CREAR, nombre, propuesta.codigo)
        else:
            nombre_anterior = existente.nombre
            actividad, accion, campos = guardar(
                Actividad, {"pk": existente.pk}, valores
            )
            if es_renombre and nombre_anterior != nombre:
                self.resultado.registrar(
                    "Actividad", RENOMBRAR, nombre,
                    f"antes «{nombre_anterior}» — se conserva su checklist",
                )
            else:
                self.resultado.registrar("Actividad", accion, nombre, ", ".join(campos))

        self._costos(actividad, fila)
        self._responsables(actividad)
        self._checklist(actividad, fila)
        return actividad

    def _costos(self, actividad, fila):
        """Vuelca los dos totales de la planilla en el desglose de costos.

        La planilla solo distingue relatoría y "otros gastos", así que llegan a
        esas dos categorías. Si alguien detalló los costos en el sistema, el
        desglose queda intacto: volcarle encima dos totales borraría el detalle
        que se abrió a mano. Es la misma regla del checklist.
        """
        campos = {}
        if "costo_relatoria" not in self.campos_ignorados:
            campos["relatoria"] = a_decimal(self._valor(fila, "Costo Relatoría"))
        if "otros_gastos" not in self.campos_ignorados:
            campos["otros"] = a_decimal(self._valor(fila, "Otros gastos"))
        if not campos:
            return

        costos = getattr(actividad, "costos", None)
        if costos is not None and costos.editado_en_sistema and not self.sobrescribir_ediciones:
            self.costos_respetados += 1
            return

        es_nuevo = costos is None
        if es_nuevo:
            costos = CostoActividad(actividad=actividad)

        cambios = [c for c, valor in campos.items() if getattr(costos, c) != valor]
        if not cambios:
            return

        for campo in cambios:
            setattr(costos, campo, campos[campo])
        costos.editado_en_sistema = False
        costos.save()
        self.resultado.registrar(
            "Costos", CREAR if es_nuevo else ACTUALIZAR,
            actividad.nombre[:50], ", ".join(cambios),
        )

    def _responsables(self, actividad):
        """Enlaza los nombres del campo de seguimiento con usuarios del equipo.

        La planilla escribe texto libre ("Pablo / Alessandra") y a veces nombra
        unidades en vez de personas ("Jurídica"). Se enlaza lo que calce con un
        usuario de los grupos de OTEC y se avisa del resto.

        Solo actúa cuando la actividad no tiene responsables asignados: si
        alguien los eligió en el sistema, la planilla no los pisa.
        """
        if actividad.responsables.exists():
            return

        encontrados = []
        for nombre in actividad.responsable_seguimiento.split("/"):
            nombre = nombre.strip()
            if not nombre:
                continue
            usuario = self.usuarios_otec.get(clave(nombre))
            if usuario:
                encontrados.append(usuario)
            else:
                self.nombres_sin_usuario[nombre] += 1

        if encontrados:
            actividad.responsables.set(encontrados)
            self.resultado.registrar(
                "Responsables", ACTUALIZAR, actividad.nombre[:50],
                ", ".join(u.get_full_name() or u.username for u in encontrados),
            )

    def _checklist(self, actividad, fila):
        existentes = {i.plantilla_id: i for i in actividad.items.all()}
        for columna, llave in COLUMNAS_EXCEL.items():
            plantilla = self.plantillas.get(llave)
            if plantilla is None or columna not in self.indice:
                continue
            estado, detalle = estado_item(self._valor(fila, columna))
            item = existentes.get(plantilla.pk)
            if item is None:
                ItemChecklist.objects.create(
                    actividad=actividad, plantilla=plantilla,
                    estado=estado, detalle=detalle[:255],
                )
                self.resultado.registrar("Checklist", CREAR, plantilla.nombre, actividad.nombre)
                continue

            sin_cambios = item.estado == estado and item.detalle == detalle[:255]
            if sin_cambios:
                self.resultado.registrar("Checklist", IGUAL, plantilla.nombre, actividad.nombre)
                continue

            if item.editado_en_sistema and not self.sobrescribir_ediciones:
                self.resultado.registrar(
                    "Checklist", CONSERVAR, plantilla.nombre,
                    f"{actividad.nombre[:40]} — se editó en el sistema "
                    f"({item.get_estado_display()}); el Excel dice "
                    f"«{dict(Estado.choices).get(estado, estado)}»",
                )
                continue

            item.estado = estado
            item.detalle = detalle[:255]
            item.editado_en_sistema = False
            item.save(update_fields=["estado", "detalle", "editado_en_sistema"])
            self.resultado.registrar(
                "Checklist", ACTUALIZAR, plantilla.nombre, actividad.nombre
            )

    def _podar(self, propuestas_tocadas, actividades_vistas):
        # Solo se poda lo que vino del Excel: lo creado en el sistema no está
        # en el archivo por definición y borrarlo sería destruir trabajo.
        sobrantes = (
            Actividad.objects
            .filter(propuesta_id__in=propuestas_tocadas, origen=Origen.IMPORTADO)
            .exclude(pk__in=actividades_vistas)
            .select_related("propuesta")
        )
        for actividad in sobrantes:
            self.resultado.registrar(
                "Actividad", ELIMINAR, actividad.nombre,
                f"ya no aparece en {actividad.propuesta.codigo}",
            )
        sobrantes.delete()
