"""Importa el archivo de flujo de caja de OTEC.

Carga las líneas de ingreso, sus costos directos, los costos transversales y
los parámetros del modelo. **No importa los resultados ya calculados** de la
planilla: el sistema los recalcula (ver ``otec/flujo.py``).

Depende de que las actividades estén cargadas para poder enlazar cada línea con
su curso, pero funciona igual sin ellas — las líneas quedan sin enlazar.

    python manage.py otec_importar_flujo [--dry-run]
"""

import re
from datetime import date
from decimal import Decimal
from difflib import SequenceMatcher

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from otec.importador import a_decimal, a_entero, a_fecha, clave, mapea, texto
from otec.models import (
    Actividad,
    CostoDirecto,
    CostoTransversal,
    Institucion,
    LineaFinanciera,
    SupuestosFinancieros,
)

HOJAS = {
    "actividades": "Actividades",
    "costos": "Costos directos",
    "transversales": "Costos transversales",
    "parametros": "Parámetros",
}

# Las líneas de cartera proyectada no corresponden a ningún curso del registro.
RE_SUPUESTO = re.compile(r"^supuesto comercial", re.IGNORECASE)
UMBRAL_ENLACE = 0.62

CERTEZAS = {
    "efectivo": LineaFinanciera.Certeza.EFECTIVO,
    "confirmado": LineaFinanciera.Certeza.CONFIRMADO,
    "probable": LineaFinanciera.Certeza.PROBABLE,
    "proyectado": LineaFinanciera.Certeza.PROYECTADO,
}

ESTADOS = {
    "ejecutada": LineaFinanciera.EstadoLinea.EJECUTADA,
    "contratada": LineaFinanciera.EstadoLinea.CONTRATADA,
    "adjudicada": LineaFinanciera.EstadoLinea.ADJUDICADA,
    "en negociación": LineaFinanciera.EstadoLinea.EN_NEGOCIACION,
    "en negociacion": LineaFinanciera.EstadoLinea.EN_NEGOCIACION,
    "proyectada": LineaFinanciera.EstadoLinea.PROYECTADA,
}

TIPOS_INSTITUCION = {
    "personas naturales": Institucion.Tipo.PERSONA,
    "empresas": Institucion.Tipo.PRIVADA,
    "osl upla": Institucion.Tipo.INTERNA,
}

# Etiqueta de la hoja Parámetros -> (campo del modelo, conversión)
PARAMETROS = {
    "saldo inicial de caja": ("saldo_inicial", a_decimal),
    "fecha de corte": ("fecha_corte", a_fecha),
    "valor hora de relatoria": ("valor_hora_relatoria", a_decimal),
    "distribucion upla por actividad": ("pct_upla", a_decimal),
    "distribucion otec por actividad": ("pct_otec", a_decimal),
    "plazo de pago de costos directos": ("plazo_pago_costos_dias", a_entero),
    "saldo minimo de caja": ("saldo_minimo", a_decimal),
}

COLUMNAS_COSTO = [
    ("relatoria", 1), ("materiales", 2), ("plataformas", 3), ("certificaciones", 4),
    ("traslados", 5), ("alimentacion", 6), ("arriendo", 7), ("otros", 8),
]


def tipo_institucion(nombre):
    k = clave(nombre)
    for fragmento, tipo in TIPOS_INSTITUCION.items():
        if fragmento in k:
            return tipo
    return Institucion.Tipo.PUBLICA


class Command(BaseCommand):
    help = "Importa el flujo de caja de OTEC (líneas, costos y parámetros)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--archivo",
            default=r"C:\Users\claud\Desktop\OTEC\Flujo de caja OTEC UPLA 2026 Ajustado.xlsx",
        )
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--anio", type=int, default=2026)

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise CommandError("Falta openpyxl.") from exc

        try:
            wb = load_workbook(options["archivo"], data_only=True, read_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f"No se encontró el archivo: {options['archivo']}") from exc

        faltan = [h for h in HOJAS.values() if h not in wb.sheetnames]
        if faltan:
            wb.close()
            raise CommandError(f"Al archivo le faltan las hojas: {', '.join(faltan)}.")

        datos = {k: list(wb[h].iter_rows(values_only=True)) for k, h in HOJAS.items()}
        wb.close()

        self.avisos = []
        try:
            with transaction.atomic():
                r = self._importar(datos, options["anio"])
                if options["dry_run"]:
                    raise _Rollback()
        except _Rollback:
            self.stdout.write(self.style.WARNING("DRY-RUN: no se guardó nada."))

        for aviso in self.avisos:
            self.stdout.write(self.style.WARNING(aviso))

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(
            f"Líneas: {r['lineas']} ({r['enlazadas']} enlazadas a un curso, "
            f"{r['proyeccion']} de cartera proyectada) · "
            f"Costos directos: {r['costos']} · "
            f"Costos transversales: {r['transversales']} · "
            f"Instituciones nuevas: {r['instituciones']}"
        ))

    # ------------------------------------------------------------------

    def _importar(self, datos, anio):
        resumen = {
            "lineas": 0, "enlazadas": 0, "proyeccion": 0,
            "costos": 0, "transversales": 0, "instituciones": 0,
        }

        supuestos = self._parametros(datos["parametros"], anio)

        # El archivo es la fuente completa de estas tablas: se reemplazan enteras
        # para que borrar una línea en la planilla también la borre acá.
        CostoTransversal.objects.all().delete()
        LineaFinanciera.objects.all().delete()

        actividades = list(Actividad.objects.all())
        cache_inst = {clave(i.nombre): i for i in Institucion.objects.all()}

        for fila in datos["actividades"][2:]:
            if not fila or not fila[0]:
                continue
            codigo = texto(fila[0])
            nombre_inst = texto(fila[1]) or "Sin institución"
            descripcion = texto(fila[2])

            k = clave(nombre_inst)
            institucion = cache_inst.get(k)
            if institucion is None:
                institucion = Institucion.objects.create(
                    nombre=nombre_inst, tipo=tipo_institucion(nombre_inst)
                )
                cache_inst[k] = institucion
                resumen["instituciones"] += 1

            origen = texto(fila[16])
            es_supuesto = bool(RE_SUPUESTO.match(origen))
            autoaprendizaje = "autoaprendizaje" in clave(descripcion)

            actividad = None
            if not es_supuesto:
                actividad = self._enlazar(descripcion, actividades)

            certeza = mapea(fila[4], CERTEZAS, None)
            if certeza is None:
                certeza = LineaFinanciera.Certeza.PROYECTADO
                if texto(fila[4]):
                    self.avisos.append(
                        f"{codigo}: nivel de certeza «{texto(fila[4])}» desconocido, "
                        f"se toma como proyectado."
                    )

            linea = LineaFinanciera.objects.create(
                codigo=codigo,
                institucion=institucion,
                actividad=actividad,
                descripcion=descripcion[:300],
                estado=mapea(fila[3], ESTADOS, LineaFinanciera.EstadoLinea.PROYECTADA),
                certeza=certeza,
                autoaprendizaje=autoaprendizaje,
                participantes=a_entero(fila[5]) or 0,
                horas=a_entero(fila[6]) or 0,
                fecha_inicio=a_fecha(fila[7]),
                fecha_termino=a_fecha(fila[8]),
                fecha_facturacion=a_fecha(fila[9]),
                fecha_pago_estimada=a_fecha(fila[10]),
                fecha_pago_efectiva=a_fecha(fila[11]),
                valor_ofertado=a_decimal(fila[12]),
                monto_contratado=a_decimal(fila[13]),
                monto_facturado=a_decimal(fila[14]),
                monto_pagado=a_decimal(fila[15]),
                origen=origen[:200],
                observacion=texto(fila[17])[:500],
            )
            resumen["lineas"] += 1
            resumen["enlazadas"] += int(actividad is not None)
            resumen["proyeccion"] += int(linea.es_proyeccion)

        # --- Costos directos ---
        por_codigo = {l.codigo: l for l in LineaFinanciera.objects.all()}
        for fila in datos["costos"][2:]:
            if not fila or not fila[0]:
                continue
            linea = por_codigo.get(texto(fila[0]))
            if linea is None:
                self.avisos.append(
                    f"Costos de «{texto(fila[0])}» sin línea de ingreso equivalente: se omiten."
                )
                continue
            CostoDirecto.objects.create(
                linea=linea,
                **{campo: a_decimal(fila[i]) for campo, i in COLUMNAS_COSTO},
                fecha_pago_estimada=a_fecha(fila[10]),
                fecha_pago_efectiva=a_fecha(fila[11]),
                estado=texto(fila[12])[:40],
                observacion=texto(fila[13])[:500],
            )
            resumen["costos"] += 1

        # --- Costos transversales ---
        for fila in datos["transversales"][2:]:
            if not fila or not fila[0]:
                continue
            CostoTransversal.objects.create(
                codigo=texto(fila[0])[:40],
                tipo=texto(fila[1])[:60],
                descripcion=texto(fila[2])[:300],
                area=texto(fila[3])[:120],
                monto=a_decimal(fila[4]),
                fecha_pago=a_fecha(fila[5]),
                criterio=texto(fila[6])[:120],
                incluir_en_flujo=clave(fila[7]) in ("si", "sí", "true", "1") or not texto(fila[7]),
                observacion=texto(fila[8])[:500],
                fuente_financiamiento=texto(fila[9])[:120],
            )
            resumen["transversales"] += 1

        supuestos.save()
        return resumen

    def _enlazar(self, descripcion, actividades):
        """Une la línea con su curso por parecido de nombre.

        A diferencia de la Gantt no es uno a uno: el flujo parte un mismo curso
        de autoaprendizaje en ocho líneas mensuales de facturación.
        """
        mejor, puntaje = None, 0
        objetivo = clave(descripcion)
        for a in actividades:
            s = SequenceMatcher(None, objetivo, clave(a.nombre)).ratio()
            if s > puntaje:
                mejor, puntaje = a, s
        return mejor if puntaje >= UMBRAL_ENLACE else None

    def _parametros(self, filas, anio):
        supuestos, _ = SupuestosFinancieros.objects.get_or_create(anio=anio)
        vistos = []
        for fila in filas[2:]:
            if not fila or not fila[0]:
                continue
            etiqueta = clave(fila[0])
            for buscado, (campo, conversion) in PARAMETROS.items():
                if etiqueta.startswith(buscado):
                    setattr(supuestos, campo, conversion(fila[1]))
                    vistos.append(campo)
                    break
        faltantes = set(c for c, _ in PARAMETROS.values()) - set(vistos)
        if faltantes:
            self.avisos.append(
                f"Parámetros no encontrados en la hoja (quedan en su valor actual): "
                f"{', '.join(sorted(faltantes))}"
            )
        return supuestos


class _Rollback(Exception):
    """Aborta la transacción cuando se corre con --dry-run."""
