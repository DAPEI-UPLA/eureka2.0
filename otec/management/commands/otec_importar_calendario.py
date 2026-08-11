"""Importa las hojas ``Carta Gantt`` y ``Zoom`` del Tablero Maestro OTEC.

Depende de que las actividades ya estén cargadas con
``otec_importar_tablero``: ambas hojas identifican los cursos por nombre
abreviado, no por ID.

    python manage.py otec_importar_calendario [--dry-run]
"""

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from otec.calendario import (
    BLOQUES_POR_ETIQUETA,
    CODIGOS_ZOOM,
    FILA_ASINCRONICAS,
    FILA_OTRAS,
    SALAS,
)
from otec.models import Actividad, DiaActividad, Feriado, ReservaZoom, SalaZoom

HOJA_GANTT = "Carta Gantt"
HOJA_ZOOM = "Zoom"

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

# Filas de la Gantt que no son actividades: la leyenda ("E = Ejecución") y el
# conteo de actividades en paralelo, que acá se calcula.
RE_NO_ACTIVIDAD = re.compile(
    r"^(?:[ec]\s*=|cantidad de actividades|calendarizaci)", re.IGNORECASE
)

# Similitud mínima de nombre para aceptar un emparejamiento. Los nombres de la
# Gantt son versiones abreviadas ("PMG Atención de Usuarios - SERPAT (N)" contra
# "PMG Atención de usuarios: Comunicación escrita..."), así que el umbral es
# bajo; lo que evita cruces es que la asignación sea uno a uno.
UMBRAL_NOMBRE = 0.35


def texto(valor):
    return "" if valor is None else str(valor).strip()


def normaliza(valor):
    """Minúsculas sin puntuación, para comparar nombres."""
    s = re.sub(r"[^0-9a-záéíóúüñ ]", " ", texto(valor).lower())
    return re.sub(r"\s+", " ", s).strip()


class Command(BaseCommand):
    help = "Importa la carta Gantt y el calendario de salas Zoom."

    def add_arguments(self, parser):
        parser.add_argument(
            "--archivo",
            default=r"C:\Users\claud\Desktop\OTEC\Tablero Maestro OTEC UPLA 2026 Compartida.xlsx",
        )
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument(
            "--anio",
            type=int,
            default=2026,
            help="Año de las columnas de la Gantt (solo traen mes y día).",
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise CommandError("Falta openpyxl (pip install openpyxl).") from exc

        if not Actividad.objects.exists():
            raise CommandError(
                "No hay actividades cargadas. Corra antes otec_importar_tablero."
            )

        try:
            wb = load_workbook(options["archivo"], data_only=True, read_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f"No se encontró el archivo: {options['archivo']}") from exc

        for hoja in (HOJA_GANTT, HOJA_ZOOM):
            if hoja not in wb.sheetnames:
                raise CommandError(f"El archivo no tiene la hoja {hoja!r}.")

        gantt = list(wb[HOJA_GANTT].iter_rows(values_only=True))
        zoom = list(wb[HOJA_ZOOM].iter_rows(values_only=True))
        wb.close()

        resumen = {"mensajes": []}
        try:
            with transaction.atomic():
                filas_actividad = self._importar_gantt(gantt, options["anio"], resumen)
                self._importar_zoom(zoom, filas_actividad, resumen)
                if options["dry_run"]:
                    raise _Rollback()
        except _Rollback:
            self.stdout.write(self.style.WARNING("DRY-RUN: no se guardó nada."))

        for linea in resumen["mensajes"]:
            self.stdout.write(linea)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(
            f"Gantt: {resumen['dias']} días en {resumen['filas_gantt']} actividades · "
            f"Zoom: {resumen['reservas']} reservas en {resumen['salas']} salas · "
            f"Feriados: {resumen['feriados']} · "
            f"Horas asincrónicas: {resumen['asincronicas']} días"
        ))

    # ------------------------------------------------------------------

    def _fechas_columnas(self, gantt, anio):
        """Columna -> fecha. El mes va en la fila 1 (combinada) y el día en la 2."""
        fechas = {}
        mes = None
        fila_mes, fila_dia = gantt[0], gantt[1]
        for j in range(max(len(fila_mes), len(fila_dia))):
            valor_mes = texto(fila_mes[j]) if j < len(fila_mes) else ""
            if valor_mes.upper() in MESES:
                mes = MESES[valor_mes.upper()]
            dia = fila_dia[j] if j < len(fila_dia) else None
            if mes and isinstance(dia, int) and 1 <= dia <= 31:
                fechas[j] = date(anio, mes, dia)
        return fechas

    def _emparejar_actividades(self, etiquetas, resumen):
        """Asigna cada fila de la Gantt a una actividad distinta.

        Se resuelve de forma golosa por mejor puntaje global: dos filas hablan
        de "indicadores" y un emparejamiento fila por fila las confunde.
        """
        actividades = list(Actividad.objects.all())

        pares = []
        for indice, etiqueta in etiquetas.items():
            base = normaliza(texto(etiqueta).split(" - ")[0])
            for actividad in actividades:
                puntaje = SequenceMatcher(None, base, normaliza(actividad.nombre)).ratio()
                pares.append((puntaje, indice, actividad))

        pares.sort(key=lambda p: -p[0])
        asignadas, usadas, filas_listas, puntajes = {}, set(), set(), {}
        for puntaje, indice, actividad in pares:
            if indice in filas_listas or actividad.pk in usadas or puntaje < UMBRAL_NOMBRE:
                continue
            asignadas[indice] = actividad
            puntajes[indice] = puntaje
            filas_listas.add(indice)
            usadas.add(actividad.pk)

        # El emparejamiento se hace por nombre, así que conviene poder revisarlo.
        resumen["mensajes"].append("Gantt · emparejamiento de filas:")
        for indice in sorted(asignadas):
            marca = " " if puntajes[indice] >= 0.6 else "?"
            resumen["mensajes"].append(
                f"  {marca} {puntajes[indice]:.2f}  {texto(etiquetas[indice])[:44]:<46}"
                f" -> {asignadas[indice].nombre[:44]}"
            )

        sin_pareja = [e for i, e in etiquetas.items() if i not in asignadas]
        if sin_pareja:
            resumen["mensajes"].append(self.style.WARNING(
                "  Sin actividad equivalente: "
                + "; ".join(texto(e)[:45] for e in sin_pareja)
            ))
        return asignadas

    def _importar_gantt(self, gantt, anio, resumen):
        fechas = self._fechas_columnas(gantt, anio)

        etiquetas = {}
        for i in range(3, len(gantt)):
            fila = gantt[i]
            etiqueta = texto(fila[0]) if fila else ""
            if not etiqueta or RE_NO_ACTIVIDAD.match(etiqueta):
                continue
            etiquetas[i] = etiqueta

        asignadas = self._emparejar_actividades(etiquetas, resumen)

        # Se reemplazan los días de las actividades que esta hoja describe, para
        # que borrar una marca en la planilla también la borre acá.
        DiaActividad.objects.filter(
            actividad__in=asignadas.values()
        ).delete()

        nuevos = []
        for indice, actividad in asignadas.items():
            fila = gantt[indice]
            for columna, fecha in fechas.items():
                valor = texto(fila[columna]) if columna < len(fila) else ""
                if not valor:
                    continue
                tipo = (
                    DiaActividad.Tipo.CIERRE
                    if valor.upper().startswith("C")
                    else DiaActividad.Tipo.EJECUCION
                )
                nuevos.append(DiaActividad(actividad=actividad, fecha=fecha, tipo=tipo))

        DiaActividad.objects.bulk_create(nuevos)
        resumen["dias"] = len(nuevos)
        resumen["filas_gantt"] = len(asignadas)
        return asignadas

    # ------------------------------------------------------------------

    def _resolver_codigo(self, etiqueta, cache):
        """Código de la planilla -> (actividad|None, soporte).

        ``FORM (S.AP)`` -> actividad Formulación..., soporte "S.AP".
        """
        crudo = texto(etiqueta)
        soporte = ""
        parentesis = re.search(r"\(([^)]*)\)\s*$", crudo)
        if parentesis:
            soporte = parentesis.group(1).strip()
            crudo = crudo[: parentesis.start()].strip()
        # Sufijos numéricos ("FISCYSUP 2", "FORM 1,5") indican horas, no curso.
        codigo = re.sub(r"[\s.]*[\d,]+$", "", crudo).strip(" .").upper()

        if codigo in cache:
            return cache[codigo], soporte

        fragmento = CODIGOS_ZOOM.get(codigo)
        actividad = None
        if fragmento:
            actividad = next(
                (a for a in Actividad.objects.all() if fragmento in a.nombre.lower()),
                None,
            )
        cache[codigo] = actividad
        return actividad, soporte

    def _horas_de_etiqueta(self, etiqueta):
        """Extrae las horas de "FORM 1,5" -> Decimal("1.5")."""
        m = re.search(r"([\d]+(?:[,.][\d]+)?)\s*$", texto(etiqueta))
        if not m:
            return None
        try:
            return Decimal(m.group(1).replace(",", "."))
        except InvalidOperation:
            return None

    def _importar_zoom(self, zoom, filas_actividad, resumen):
        cache = {}
        codigos_desconocidos = set()
        feriados = {}
        reservas = []

        for nombre_sala, filas in SALAS:
            sala, _ = SalaZoom.objects.get_or_create(
                nombre=nombre_sala,
                defaults={"orden": len(SalaZoom.objects.all()) + 1},
            )
            ReservaZoom.objects.filter(sala=sala).delete()

            cabecera = zoom[filas["fechas"] - 1]
            fechas = {
                j: v.date() if isinstance(v, datetime) else v
                for j, v in enumerate(cabecera)
                if isinstance(v, (datetime, date))
            }

            # (columna, etiqueta) -> lista de bloques contiguos
            for columna, fecha in fechas.items():
                celdas = []
                for fila_idx in range(filas["primer_bloque"], filas["ultimo_bloque"] + 1):
                    fila = zoom[fila_idx - 1]
                    if not fila:
                        continue
                    bloque = BLOQUES_POR_ETIQUETA.get(texto(fila[0]))
                    valor = texto(fila[columna]) if columna < len(fila) else ""
                    if not bloque or not valor:
                        continue
                    if valor.upper() == "FESTIVO":
                        feriados.setdefault(fecha, "")
                        continue
                    celdas.append((bloque, valor))

                # Fusiona bloques contiguos con la misma etiqueta.
                for bloque, etiqueta in celdas:
                    inicio, fin = bloque
                    if (
                        reservas
                        and reservas[-1].sala_id == sala.pk
                        and reservas[-1].fecha == fecha
                        and reservas[-1].etiqueta == etiqueta
                        and reservas[-1].hora_fin == inicio
                    ):
                        reservas[-1].hora_fin = fin
                        continue
                    actividad, soporte = self._resolver_codigo(etiqueta, cache)
                    if actividad is None:
                        codigos_desconocidos.add(etiqueta)
                    reservas.append(ReservaZoom(
                        sala=sala, fecha=fecha, hora_inicio=inicio, hora_fin=fin,
                        actividad=actividad, etiqueta=etiqueta, soporte=soporte,
                    ))

            resumen["salas"] = resumen.get("salas", 0) + 1

        # --- Filas especiales (solo existen bajo la sala 1) ---
        cabecera = zoom[SALAS[0][1]["fechas"] - 1]
        fechas_sala1 = {
            j: v.date() if isinstance(v, datetime) else v
            for j, v in enumerate(cabecera)
            if isinstance(v, (datetime, date))
        }

        asincronicas = 0
        fila_asinc = zoom[FILA_ASINCRONICAS - 1] if FILA_ASINCRONICAS <= len(zoom) else None
        if fila_asinc:
            for columna, fecha in fechas_sala1.items():
                etiqueta = texto(fila_asinc[columna]) if columna < len(fila_asinc) else ""
                if not etiqueta:
                    continue
                actividad, _ = self._resolver_codigo(etiqueta, cache)
                horas = self._horas_de_etiqueta(etiqueta)
                if actividad is None:
                    codigos_desconocidos.add(etiqueta)
                    continue
                dia, _ = DiaActividad.objects.get_or_create(
                    actividad=actividad,
                    fecha=fecha,
                    defaults={"tipo": DiaActividad.Tipo.EJECUCION},
                )
                dia.horas_asincronicas = horas
                dia.save(update_fields=["horas_asincronicas"])
                asincronicas += 1

        fila_otras = zoom[FILA_OTRAS - 1] if FILA_OTRAS <= len(zoom) else None
        if fila_otras:
            sala1 = SalaZoom.objects.get(nombre=SALAS[0][0])
            for columna, fecha in fechas_sala1.items():
                etiqueta = texto(fila_otras[columna]) if columna < len(fila_otras) else ""
                if not etiqueta:
                    continue
                reservas.append(ReservaZoom(
                    sala=sala1, fecha=fecha, hora_inicio=None, hora_fin=None,
                    actividad=None, etiqueta=etiqueta,
                    observacion="Otras actividades (sin bloque horario).",
                ))

        ReservaZoom.objects.bulk_create(reservas)

        for fecha in feriados:
            Feriado.objects.get_or_create(fecha=fecha)

        resumen["reservas"] = len(reservas)
        resumen["feriados"] = len(feriados)
        resumen["asincronicas"] = asincronicas

        if codigos_desconocidos:
            resumen["mensajes"].append(self.style.WARNING(
                f"{len(codigos_desconocidos)} códigos de Zoom sin curso asociado "
                f"(quedan como reserva sin actividad): "
                + "; ".join(sorted(codigos_desconocidos))
            ))


class _Rollback(Exception):
    """Aborta la transacción cuando se corre con --dry-run."""
